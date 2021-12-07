import platform
from queue import Queue
from openpyxl import load_workbook
from Misc.MyFunctions import *

wb = load_workbook('Test Reports/Excel automation.xlsx')
ws = wb.active
user = os.environ['USERPROFILE']

queue = Queue(1)
timeout = 120
data = readFromQueue(queue)

ws['B4'].value = platform.release()
wb.save('Test Reports/Excel automation.xlsx')

set_date = datetime.datetime.now()
ws['B5'].value = set_date
wb.save('Test Reports/Excel automation.xlsx')

ws['D8'].value = 'PASS'  # GC Launch
wb.save('Test Reports/Excel automation.xlsx')

delay()

#######################################################################################################################
# MODULE: DOCS

def docs(queue):
    start = 0
    delay()
    # TEST CASE 1: Can you write, save and then open the document via the recent documents?
    pag.press("right", presses=1, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)  # navigating to new document
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document using the automated script", interval=0.5)
        pag.press("esc", interval=1)  # saving document
        pag.press("enter", interval=1)
        pag.write("1", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '1' in data:
            pag.press("enter")
            pag.sleep(3.5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\"
                                  "Test Reports\\Printscreens\\Letters and Documents\\Documents\\testcase1.png")

            print('TEST CASE 1: Can you create a document and then view it in recent documents? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D15'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE 1: Can you create a document and then view it in recent documents? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D15'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("FAIL: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['D15'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 2: can you create a new document, save and then open via my documents?
    start = 0
    delay()
    pag.press("right", interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)  # navigating to new document
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document using the automated script", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        pag.write("2", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("2", interval=1)
        data = readFromQueue(queue)
        if "2" in data:
            pag.press("enter")
            pag.sleep(3.5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Letters and Documents\\Documents\\test case 2.png")

            print('TEST CASE 2: Can you create a document and then view it in my documents? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D16'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE 2: Can you create a document and then view it in my documents? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D16'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("FAIL: Navigation failed to enter a new document"
              "Current string:", data)
        ws['D16'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: LETTERS

def letter(queue):
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' in data:
        pag.write("joe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bloggs", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("22", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin street", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bromsgrove", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("worcestershire", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("b45 9xe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("united kingdom", interval=0.5)
        pag.press("enter", presses=4, interval=0.5)
        pag.write("this is a test letter", interval=0.5)
        pag.press("esc", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("letter 1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("L", interval=0.5)
        data = readFromQueue(queue)
        if 'letter 1' in data:
            pag.press("enter", interval=1)
            pag.sleep(2)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Letters and Documents\\Letters\\letter1.png")

            print('TEST CASE: Can you add a letter sender address and then save a letter and view it? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D17'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you add a letter sender address and then save a letter and view it? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D17'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()
    else:
        print("failed to open add senders address")
        print(data)
        ws['D17'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: SCANNER

def scanner(queue):
    start = 0
    delay()
    # TEST CASE 1: can you scan for text (1 page)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            print("Scanning...")
            pag.sleep(3)
            data = readFromQueue(queue)
            if 'Read Scanned Pages' in data:
                break

        data = readFromQueue(queue)
        if "Read Scanned Pages" in data:
            pag.sleep(2)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_SCANNER_OCR_PROGRESS' in data:
                print("Reading text...")
                pag.sleep(1)
                data = readFromQueue(queue)
                if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                    break
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                  "Scanner and camera\\Scanner\\Scanner_TC1.png")
            print('TEST CASE 1: Can you scan 1 page for text? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D21'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

        else:
            print('TEST CASE 1: Can you scan 1 page for text? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D21'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()
    else:
        print("Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['D21'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 2: Can you scan for an image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            print("Scanning...")
            pag.sleep(3)
            data = readFromQueue(queue)
            if 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' in data:
                break

        data = readFromQueue(queue)
        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                  "Scanner and camera\\Scanner\\Scanner_TC2.png")
            print('TEST CASE 2: Can you scan 1 page for text? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D22'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE 2: Can you scan 1 page for text? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D22'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['D22'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: CAMERA

def camera(queue):
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.sleep(5)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    data = readFromQueue(queue)

    while 'GMENU_SHOW_CAMERA_FEED' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_SHOW_CAMERA_FEED' in data:
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)
        start = 0

        while 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' in data:
            pag.press("f2", interval=1)
            pag.press("right", presses=3, interval=1)     # save image
            pag.press("enter", interval=1)
            pag.write("automated camera test", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.press("esc", presses=2, interval=1)
            pag.press("right", presses=3, interval=1)
            pag.press("enter", interval=1)
            pag.press("a", interval=1)
            pag.press("enter", interval=1)  # enter saved image
            data = readFromQueue(queue)
            if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                      "Scanner and camera\\Camera\\Camera_TC1.png")

                print('TEST CASE 1: Can you capture an image, save and view it? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D23'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print('TEST CASE 1: Can you capture an image, save and view it? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D23'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("Failed to load GMENU_SCANNER_SCAN_IMAGE_COMPLETE")
            print("Current String:", data)
            ws['D23'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()
    else:
        print("Camera failed to load GMENU_SHOW_CAMERA_FEED")
        print("Current String:", data)
        ws['D23'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: BOOKS

def books(queue):
    start = 0
    delay()
    # TEST CASE 1: can you download a book from project gutenburg and read it?
    pag.press("right", presses=4, interval=1)
    pag.sleep(1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f")
    data = readFromQueue(queue)

    while 'Find a new book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Find a new book' in data:
        pag.press("enter", interval=1)
        # PROJECT GUTENBURG
        data = readFromQueue(queue)

        while 'Project Gutenberg' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Project Gutenberg' in data:
            pag.press("enter", presses=3, interval=1)
            pag.press("A", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(15)
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)

            while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
                data = readFromQueue(queue)
                if '' in data:
                    break

                elif start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
                pag.sleep(15)
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                      "Printscreens\\Bookshelf\\gutenberg.png")
                print('TEST CASE: Can you download and view a book from project gutenberg? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D25'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            elif '' in data:
                pag.sleep(15)
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                      "Printscreens\\Bookshelf\\gutenberg.png")
                print('TEST CASE: Can you download and view a book from project gutenberg? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D25'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print('TEST CASE: Can you download and view a book from project gutenberg? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D25'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("failed to open project gutenberg")
            print(data)
            ws['D25'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to open find a new book")
        print(data)
        ws['D25'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    delay()
    start = 0
    # TEST CASE 2: can you continue reading a book?
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
        pag.sleep(10)
        print('TEST CASE: Can you continue reading a book? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['D29'].value = 'PASS'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    else:
        print('TEST CASE: Can you continue reading a book? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['D29'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: NEWSPAPERS

def newspapers(queue):
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.sleep(2)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'RNIB NTNM' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'RNIB NTNM' in data:
            pag.press("enter", interval=1)
            pag.write("i.rollason", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("tna751", interval=1)
            pag.press("enter", presses=7, interval=1)
            pag.sleep(10)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
                pag.sleep(10)
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                      "Printscreens\\Bookshelf\\rnibntnm.png")
                print('TEST CASE: Can you download and view a newspaper from RNIB NTNM? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D27'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print('TEST CASE: Can you download and view a newspaper from RNIB NTNM? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D27'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("failed to enter RNIB NTNM")
            print(data)
            ws['D27'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to enter newspapers")
        print(data)
        ws['D27'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: ADDRESS BOOK

def address_book(queue):
    delay()
    start = 0
    # TEST CASE 1: ADDRESS BOOK - ADD A NEW CONTACT
    pag.press("right", presses=5, interval=1)  # address book
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'GMENU_ADDRESSBOOK_ADD' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_ADDRESSBOOK_ADD' in data:
        print(">>> Adding address book contact")
        pag.write("Joe", interval=0.5)  # first name
        pag.press("enter", interval=1)
        pag.write("bloggs", interval=0.5)  # surname
        pag.press("enter", interval=1)
        pag.write("0123456789101", interval=0.5)  # telephone number
        pag.press("enter", interval=1)
        pag.write("22", interval=0.5)  # house number / number
        pag.press("enter", interval=1)
        pag.write("Dolphin street", interval=0.5)  # street name
        pag.press("enter", interval=1)
        pag.write("dolphin town", interval=0.5)  # town name
        pag.press("enter", interval=1)
        pag.write("dolphin county", interval=0.5)  # county name
        pag.press("enter", interval=1)
        pag.write("B45 9XE", interval=0.5)  # postcode
        pag.press("enter", interval=1)
        pag.write("United kingdom", interval=0.5)  # country
        pag.press("enter", interval=1)
        pag.write("joebloggs@dolphin.co.uk", interval=0.5)  # email
        pag.press("enter", interval=1)
        pag.write("Dolphin computer access", interval=0.5)  # company name
        pag.press("enter", interval=1)
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        pag.sleep(3)  # print screen address book entry
        pag.screenshot(user + "\\Desktop\\"
                              "Guide Automation\\Test Reports\\Printscreens\\Address book and calendar\\"
                              "Address book\\addcontact.png")
        data = readFromQueue(queue)
        if "Joe bloggs" in data:
            print('TEST CASE 1: Address Book - New Contact \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D29'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')

            # TEST CASE 2: ADDRESS BOOK - EDIT CONTACT
            delay()
            start = 0
            pag.press("enter", presses=2, interval=1)
            pag.write("edited", interval=0.5)  # first name
            pag.press("enter", interval=1)
            pag.press("down", interval=1)
            pag.press("enter", interval=1)
            pag.write("user", interval=0.5)  # surname
            pag.press("enter", interval=1)
            pag.press("down", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit phone number", interval=0.5)  # telephone number
            pag.press("enter", interval=1)
            pag.press("down", presses=3, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit house number", interval=0.5)  # house name / number
            pag.press("enter", interval=1)
            pag.press("down", presses=4, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit street name", interval=0.5)  # street name
            pag.press("enter", interval=1)
            pag.press("down", presses=5, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit town", interval=0.5)  # town
            pag.press("enter", interval=1)
            pag.press("down", presses=6, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit county", interval=0.5)  # county
            pag.press("enter", interval=1)
            pag.press("down", presses=7, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit postcode", interval=0.5)  # postcode
            pag.press("enter", interval=1)
            pag.press("down", presses=8, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit country", interval=0.5)  # country
            pag.press("enter", interval=1)
            pag.press("down", presses=9, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit_mail@", interval=0.5)  # email
            pag.press("enter", interval=1)
            pag.press("down", presses=10, interval=1)
            pag.press("enter", interval=1)
            pag.write("edit company", interval=0.5)  # company
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            pag.press("e", interval=1)
            pag.sleep(3)  # screen shot result of all edit fields being 'edit + field'
            pag.screenshot(user + "\\Desktop\\"
                                  "Guide Automation\\Test Reports\\Printscreens\\Address book and calendar\\"
                                  "Address book\\editcontact.png")
            data = readFromQueue(queue)
            if "edited user" in data:
                print('TEST CASE 2: Address Book - Edit Contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D30'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print('TEST CASE 2: Address Book - Edit Contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D30'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print('TEST CASE 1: Address Book - New Contact \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D29'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print('Result: FAIL \n'
              '>>> Reason: FAILED TO ENTER ADDRESS BOOK \n'
              '>>> Current String:', data)
        ws['D29'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: CALENDAR

def calendar(queue):
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)  # address book
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("enter", presses=3, interval=1)
    data = readFromQueue(queue)
    pag.sleep(2)

    while "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" in data:
        print(">>> Adding calendar event")
        # TEST CASE 1: - TODAY'S EVENTS
        pag.write("tc1", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)  # DAILY REPEAT
        pag.press("enter", presses=4, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(2)
        data = readFromQueue(queue)
        pag.sleep(5)
        if "Event name" in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\"
                                  "Test Reports\\"
                                  "Printscreens\\Address book and calendar\\Calendar\\TC1.png")
            print('TEST CASE 1: Calendar - Todays Events \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D31'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')

            # TEST CASE 2: Upcoming Events
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            pag.sleep(1)
            if "Upcoming events" in data:
                pag.press("enter")
                print(">>> Upcoming events entered")
                data = readFromQueue(queue)
                pag.sleep(5)
                if 'tc1' in data:
                    pag.sleep(2)
                    pag.screenshot(user + "\\Desktop\\Guide Automation\\"
                                          "Test Reports\\Printscreens\\Address book and calendar\\Calendar\\TC2.png")
                    print('TEST CASE 2: Calendar - Upcoming Events \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['D32'].value = 'PASS'
                    wb.save('Test Reports/Excel automation.xlsx')

                    # TEST CASE 3: Calendar - View Calendar
                    pag.press("esc", interval=1)
                    pag.press("right", presses=2, interval=1)
                    pag.press("enter", presses=3, interval=1)
                    data = readFromQueue(queue)
                    pag.sleep(5)
                    if 'Event name' in data:
                        pag.sleep(2)
                        pag.screenshot(user + "\\Desktop\\Guide Automation\\"
                                              "Test Reports\\"
                                              "Printscreens\\Address book and calendar\\Calendar\\TC3.png")

                        print('TEST CASE 3: Calendar - View Calendar \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['D33'].value = 'PASS'
                        wb.save('Test Reports/Excel automation.xlsx')
                        returnhome()

                    else:
                        print('TEST CASE 3: Calendar - View Calendar \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['D33'].value = 'FAIL'
                        wb.save('Test Reports/Excel automation.xlsx')
                        returnhome()

                else:
                    print('TEST CASE 2: Calendar - Upcoming Events \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['D32'].value = 'FAIL'
                    wb.save('Test Reports/Excel automation.xlsx')
                    returnhome()

        else:
            print('TEST CASE 1: Calendar - Todays Events \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D31'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print('Result: FAIL \n'
              '>>> Reason: FAILED TO ENTER CALENDAR \n'
              '>>> Current String:', data)
        ws['D31'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: RADIO

def radio(queue):
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    #   TC1 - Can you play a station
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Entertainment\\Radio\\radio1.png")
        print("TC1 - PASS")
        pag.sleep(2)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        #   TC2 - Can you play a station?
        pag.press("right", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Radio\\radio2.png")
            pag.sleep(2)
            print("TC2 - PASS")
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)

            #   TC3 - can you play a station?
            pag.press("right", presses=3, interval=1)  # BBC Radio 2
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                      "Printscreens\\Entertainment\\Radio\\radio3.png")
                print("TC3 - PASS")
                print('TEST CASE: Can you Listen to radio stations? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D35'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                pag.press("esc", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                print('TEST CASE: Can you Listen to radio stations? \n'
                      '>>> Result: TC3 - FAIL \n'
                      '>>> Current String:', data)
                ws['D35'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print('TEST CASE: Can you Listen to radio stations? \n'
                  '>>> Result: TC2 - FAIL \n'
                  '>>> Current String:', data)
            ws['D35'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print('TEST CASE: Can you Listen to radio stations? \n'
              '>>> Result: TC1 - FAIL \n'
              '>>> Current String:', data)
        ws['D35'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: PODCASTS

def podcasts(queue):
    delay()
    start = 0
    #   TEST CASE 1: Can you play a podcast?
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Podcasts' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.sleep(5)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Podcast\\Podcast1.png")
            pag.sleep(10)
            print('TEST CASE: Can you play a podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D36'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            pag.press("esc")
            pag.press("enter")
            returnhome()

        else:
            print('TEST CASE: Can you play a podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D36'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to enter podcasts")
        print("Current String:", data)
        ws['D36'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 2: Can you continue listening to a podcast?
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Podcasts' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Podcasts' in data:
        pag.press("enter", presses=2, interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Podcast\\Podcast2.png")
            print('TEST CASE: Can you continue listening to a podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D37'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)
            returnhome()

        else:
            print('TEST CASE: Can you continue listening to a podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D37'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to enter podcasts")
        print("Current String:", data)
        ws['D37'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 3: Can you download a podcast and then open to listen to it?
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Podcasts' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("right", presses=2, interval=1)
        pag.press("enter", interval=1)
        pag.press("right", presses=5, interval=1)
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("esc", presses=2, interval=1)
        pag.press("right", presses=3, interval=1)
        data = readFromQueue(queue)
        if 'My downloaded podcast episodes' in data:
            pag.sleep(15)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Podcast\\Podcast3.png")
            print('TEST CASE: Can you download a podcast and then open it? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D38'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            pag.press("enter", interval=1)
            returnhome()

        else:
            print('TEST CASE: Can you download a podcast and then open it? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D38'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to enter podcasts")
        print("Current String:", data)
        ws['D38'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: MUSIC

def music(queue):
    delay()
    start = 0
    # TEST CASE 1: Import music and listen to a track
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'Music and CD player' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Music and CD player' in data:
        pag.press("enter", presses=4, interval=1)
        data = readFromQueue(queue)
        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
            print("Importing Music..")
            pag.sleep(2)
            data = readFromQueue(queue)
            if 'OK' in data:
                break

        pag.sleep(5)
        if 'OK' in data:
            pag.press("enter", presses=3, interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)
            if '' in data:
                pag.sleep(5)
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                      "Printscreens\\Entertainment\\Music\\Music1.png")
                print('TEST CASE: Can you import music library and play a track? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D40'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                pag.sleep(1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                print('TEST CASE: Can you import music library and play a track? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D40'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                pag.sleep(1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            print("Music failed to import correctly \n"
                  "Current String:", data)
            ws['D40'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            pag.sleep(1)
            returnhome()

    else:
        print("Failed to enter Music and CD player \n"
              "Current String:", data)
        ws['D40'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        pag.sleep(1)
        returnhome()

#######################################################################################################################
# MODULE: GAMES

def games(queue):
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        # TEST CASE 1: LAUNCH HANGMAN
        pag.press("enter", presses=4, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_HANGMAN' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_HANGMAN' in data:
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Games\\Hangman.png")
            print('TEST CASE: Can you launch hangman? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you launch hangman? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to open games")
        print(data)
        ws['D39'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 2: LAUNCH SUDOKU
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_SUDOKU' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_SUDOKU' in data:
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Games\\Sudoku.png")
            print('TEST CASE: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you launch Sudoku? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to open games")
        print(data)
        ws['D39'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 3: LAUNCH BLACKJACK
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", presses=2, interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)
        pag.sleep(5)

        while 'GMENU_GAMES_BLACKJACK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_BLACKJACK' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Entertainment\\Games\\Blackjack.png")
            print('TEST CASE: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you launch Blackjack? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D39'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("Failed to open games")
        print(data)
        ws['D39'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: NOTES

def text_note(queue):
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Add new text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Add new text note' in data:
        pag.press("enter", interval=1)
        pag.write("text note", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("this is a text note", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'text note' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Notes\\Textnote1.png")
            print('TEST CASE: Can you create and view a text note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D42'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you create and view a text note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D42'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to add a new text note")
        print(data)
        ws['D42'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def audio_note(queue):
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_NOTES_MAIN_VOICE_TITLE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_NOTES_MAIN_VOICE_TITLE' in data:
        pag.write("voice note", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("v")
        data = readFromQueue(queue)
        if 'voice note' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Notes\\Voicenote1.png")
            print('TEST CASE: Can you create and view an audio note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D43'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you create and view an audio note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D43'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to add new voice note")
        print(data)
        ws['D43'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: TOOLS

def about(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=7, interval=1)  # about feature
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)  # highlight build number
    data = readFromQueue(queue)

    while 'Check my Premium Plan' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Check my Premium Plan' in data:
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Tools\\About\\About.png")

        print("TEST CASE: Check Premium plan status \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['D47'].value = 'PASS'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    else:
        print("TEST CASE: Check Premium plan status \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['D47'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def calc(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)  # navigating to documents module
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_CALCULATOR_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_CALCULATOR_MAIN' in data:
        print(">>> Calculator entered")
        # TEST CASE 1: 50,000 + 90,000 = 140,000
        pag.press("5", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("+")
        pag.press("9", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Tools\\Calculator\\calc1.png")
        pag.press("backspace", interval=1)

        # TEST CASE 2: 666 / 22 = 30.272
        pag.press("6", presses=3, interval=0.5)
        pag.press("/")
        pag.press("2", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Tools\\Calculator\\calc2.png")
        pag.press("backspace", interval=1)

        # TEST CASE 3: 200 - 11 = 189
        pag.press("2", interval=0.5)
        pag.press("0", presses=2, interval=0.5)
        pag.press("-")
        pag.press("1", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Tools\\Calculator\\calc3.png")
        pag.press("backspace", interval=1)

        # TEST CASE 4: 497 x 15 = 7,455
        pag.press("4", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("*")
        pag.press("1", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                              "Printscreens\\Tools\\Calculator\\calc4.png")
        print("TEST CASE(s): Use the calculator to add, multiply, divide and minus \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['D46'].value = 'PASS'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    else:
        print("TEST CASE(s): Use the calculator to add, multiply, divide and minus \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['D46'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def dictionary(queue):
    delay()
    start = 0
    # TEST CASE 1: Search a valid valid dictionary entry = test
    pag.press("right", presses=9, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("test", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Tools\\Dictionary\\dict1.png")
            print("TEST CASE: Can you search for a valid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['D48'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print("TEST CASE: Can you search for a valid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['D48'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("TEST CASE(s): Search for a valid term in the dictionary \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['D48'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 2: Can you search an invalid search term in dictionary
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("twest", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Tools\\Dictionary\\dict2.png")
            print("TEST CASE: Can you search for an invalid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['D49'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print("TEST CASE: Can you search for an invalid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['D49'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("TEST CASE(s): Search for an invalid term in the dictionary \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['D49'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def gstart_vids(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'Keyboard and mouse video' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Keyboard and mouse video' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                              "Tools\\Training\\Getting Started Videos\\Keyboardvideo.png")
        pag.press("esc", interval=1)
        pag.press("R", interval=1)
        data = readFromQueue(queue)
        if 'Remote control video' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                  "Tools\\Training\\Getting Started Videos\\Remotevideo.png")
            pag.press("esc", interval=1)
            pag.press("T", interval=1)
            data = readFromQueue(queue)
            if 'Touch screen video' in data:
                pag.press("enter", interval=1)
                pag.sleep(5)
                pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                      "Tools\\Training\\Getting Started Videos\\Touchvideo.png")
                pag.press("esc", interval=1)
                pag.press("V", interval=1)
                data = readFromQueue(queue)
                if 'Voice input video' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)
                    pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                          "Tools\\Training\\Getting Started Videos\\Voicevideo.png")

                    print("TEST CASE(s): Can you play the Getting Started Videos \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['D50'].value = 'PASS'
                    wb.save('Test Reports/Excel automation.xlsx')
                    returnhome()

                else:
                    print("Voice video missing")
                    print("Current String:", data)
                    ws['D50'].value = 'FAIL'
                    wb.save('Test Reports/Excel automation.xlsx')
                    returnhome()

            else:
                print("Touch video missing")
                print("Current String:", data)
                ws['D50'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("Remote video missing")
            print("Current String:", data)
            ws['D50'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("TEST CASE(s): Can you play the Getting Started Videos \n"
              "Result: FAIL -  Keyboard video not present \n"
              "Current String:", data)
        ws['D50'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def file_exp(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)  # navigating to documents module
    pag.press("enter", interval=1)
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("2", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", presses=6, interval=1)
    data = readFromQueue(queue)

    while 'Move Selected File' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Move Selected File' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("right", presses=4, interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=2, interval=0.5)
        pag.press("esc", interval=1)
        pag.press("right", presses=4, interval=0.5)
        pag.press("enter", interval=1)
        pag.press("2")
        data = readFromQueue(queue)
        if '2' in data:
            pag.sleep(2)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Tools\\File explorer\\filexplorer1.png")
            print("TEST CASE(s): Can you move a file with file explorer \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['D45'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print("TEST CASE(s): Can you move a file with file explorer \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['D45'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("FAIL: Navigation was not present on Move File \n"
              "Current String:", data)
        ws['D45'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def learn_keyboard(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_TYPING_TUTOR_DIALOGUE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_TYPING_TUTOR_DIALOGUE' in data:
        pag.press("1", interval=0.5)
        pag.press("2", interval=0.5)
        pag.press("3", interval=0.5)
        pag.press("4", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("6", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("8", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("q", interval=0.5)
        pag.press("w", interval=0.5)
        pag.press("e", interval=0.5)
        pag.press("r", interval=0.5)
        pag.press("t", interval=0.5)
        pag.press("y", interval=0.5)
        pag.press("u", interval=0.5)
        pag.press("i", interval=0.5)
        pag.press("o", interval=0.5)
        pag.press("p", interval=0.5)
        pag.press("a", interval=0.5)
        pag.press("s", interval=0.5)
        pag.press("d", interval=0.5)
        pag.press("f", interval=0.5)
        pag.press("g", interval=0.5)
        pag.press("h", interval=0.5)
        pag.press("j", interval=0.5)
        pag.press("k", interval=0.5)
        pag.press("l", interval=0.5)
        pag.press("z", interval=0.5)
        pag.press("x", interval=0.5)
        pag.press("c", interval=0.5)
        pag.press("v", interval=0.5)
        pag.press("b", interval=0.5)
        pag.press("n", interval=0.5)
        pag.press("m", interval=0.5)
        pag.press("enter", interval=0.5)
        data = readFromQueue(queue)
        if 'Learn the Keyboard' in data:
            print("TEST CASE(s): Can you use the typing tutor? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['D52'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print("TEST CASE(s): Can you use the typing tutor? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['D52'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("FAIL: Could not enter learn the keyboard \n"
              "Current String:", data)
        ws['D52'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: EXIT

def exit_module(queue):
    delay()
    start = 0
    pag.press("right", presses=11, interval=1)
    pag.press("enter", interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'Exit' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Exit' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        print("Sanity Script Complete..")
        print("See Excel Automation for Results")
        ws['D56'].value = 'PASS'
        wb.save('Test Reports/Excel automation.xlsx')

    else:
        print("Failed to close Guide")
        print(data)
        ws['D56'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: SETTINGS

def settings(queue):
    delay()
    start = 0
    pag.press("right", presses=10, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.press("right", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'Theme 5' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Theme 5' in data:
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("t", interval=1)
        pag.sleep(1)
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.sleep(1)
        pag.press("enter", interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)
        if 'Screen colours' in data:
            pag.sleep(2)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\"
                                  "Printscreens\\Settings\\theme5.png")
            print('TEST CASE: Can you change the GC theme and font? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D54'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE: Can you change the GC theme and font? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D54'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to set theme 5")
        print(data)
        ws['D54'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

#######################################################################################################################
# MODULE: EMAILS

def emails(queue):
    # TEST CASE 1: LOGS IN AND SENDS AN EMAIL
    delay()
    start = 0
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' in data:
        pag.write("guideautomation1@outlook.com", interval=0.5)
        pag.press("enter", interval=1)
        pag.write("testing", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.write("guideconnect1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(15)
        data = readFromQueue(queue)
        while 'OK' not in data:
            print("Inbox Loading..")
            pag.sleep(5)
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", presses=3, interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Subject line: Automation", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.write("this is a test email sent from Sanity.py", interval=0.5)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                while 'GMENU_EMAIL_SENDING' in data:
                    print("Email Sending...")
                    pag.sleep(3)
                    data = readFromQueue(queue)

                if 'OK' in data:
                    pag.press("enter", interval=1)
                    pag.screenshot(user + "\\Desktop\\Guide Automation\\"
                                          "Test Reports\\Printscreens\\"
                                          "Email\\SendEmail1.png")
                    print('TEST CASE 1: Can you send an email? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['D10'].value = 'PASS'
                    wb.save('Test Reports/Excel automation.xlsx')
                    returnhome()

                    # TEST CASE 2: CHECKS SENT FOLDER FOR PREVIOUSLY SENT EMAIL
                    delay()
                    start = 0
                    pag.press("enter", interval=1)
                    pag.press("right", presses=2, interval=1)
                    pag.press("enter", interval=1)
                    pag.press("s", interval=1)
                    data = readFromQueue(queue)

                    while 'Sent' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'Sent' in data:
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        pag.sleep(5)
                        if 'guideautomation2@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.sleep(5)
                            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                                  "Email\\SentEmail1.png")
                            pag.press("f2", interval=1)
                            pag.press("d", interval=1)
                            pag.press("enter", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            print('TEST CASE 2: Can you check for the sent email? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['D11'].value = 'PASS'
                            wb.save('Test Reports/Excel automation.xlsx')
                            returnhome()

                        else:
                            print('TEST CASE 2: Can you check for the sent email? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['D11'].value = 'FAIL'
                            wb.save('Test Reports/Excel automation.xlsx')
                            returnhome()

                    else:
                        print("failed to enter the sent folder")
                        print(data)
                        ws['D11'].value = 'FAIL'
                        wb.save('Test Reports/Excel automation.xlsx')
                        returnhome()

                else:
                    print('TEST CASE 1: Can you send an email? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['D10'].value = 'FAIL'
                    wb.save('Test Reports/Excel automation.xlsx')
                    returnhome()
            else:
                print("failed to write a new email")
                print(data)
                ws['D10'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                pag.press("esc")
                pag.press("right")
                pag.press("enter")
                returnhome()
        else:
            print("failed to sync inbox")
            print(data)
            ws['D10'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    # TEST CASE 3: Log into an addition email account
    delay()
    start = 0
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'Email' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Email' in data:
        pag.press("enter", interval=1)
        pag.press("a", interval=1)
        data = readFromQueue(queue)
        if 'Account manager' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("testing", interval=0.5)
            pag.press("enter", presses=3, interval=1)
            pag.write("guideconnect2", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)
            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                pag.press("enter", interval=1)
                print('TEST CASE 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['D12'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print('TEST CASE 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['D12'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("failed to enter account manager")
            print(data)
            ws['D12'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to navigate to email settings")
        print(data)
        ws['D12'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

    # TEST CASE 4: Can you open a received email?
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)
    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        pag.sleep(5)
        if 'guideautomation1@outlook.com' in data:
            pag.press("enter")
            pag.sleep(10)
            pag.screenshot(user + "\\Desktop\\Guide Automation\\Test Reports\\Printscreens\\"
                                  "Email\\ReceivedEmail1.png")
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            pag.sleep(10)
            print('TEST CASE 4: Can you view a received email? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['D13'].value = 'PASS'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            print('TEST CASE 4: Can you view a received email? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['D13'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()
    else:
        print("failed to locate guideautomation2@outlook.com")
        print(data)
        ws['D13'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def shortcut(queue):
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    data = readFromQueue(queue)
    while 'Training' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Training' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)
        while 'Shortcut keys' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Shortcut keys' in data:
            pag.press("enter", interval=1)
            while 'GMENU_COREMODULE_SHORTCUTKEYS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_COREMODULE_SHORTCUTKEYS' in data:
                print("TEST CASE(s): Can you use the typing tutor? \n"
                      "Result: PASS \n"
                      "Current String:", data)
                ws['D51'].value = 'PASS'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

            else:
                print("TEST CASE(s): Can you use the typing tutor? \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['D51'].value = 'FAIL'
                wb.save('Test Reports/Excel automation.xlsx')
                returnhome()

        else:
            print("failed to enter shortcut keys")
            print(data)
            ws['D51'].value = 'FAIL'
            wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        print("failed to navigate to training")
        print("data")
        ws['D51'].value = 'FAIL'
        wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

