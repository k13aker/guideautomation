import platform, logging
from queue import Queue
from openpyxl import load_workbook
from Misc.MyFunctions import *

radio_expected = [("Absolute Radio 80's", 2), ("Absolute Radio 90's", 2), ('Absolute\xa0Radio', 2),
                  ('Absolute\xa0Radio\xa060s', 2), ('Absolute\xa0Radio\xa070s', 2),
                  ('Absolute\xa0Radio\xa0Classic Rock', 2), ('BBC - World Service', 2),
                  ('BBC Local Radio Berkshire', 2), ('BBC Local Radio Bristol', 2),
                  ('BBC Local Radio Cambridge', 2), ('BBC Local Radio Cornwall', 2),
                  ('BBC Local Radio Coventry & Warwickshire', 2), ('BBC Local Radio Cumbria', 2),
                  ('BBC Local Radio Derby', 2), ('BBC Local Radio Devon', 2), ('BBC Local Radio Essex', 2),
                  ('BBC Local Radio Gloucestershire', 2), ('BBC local Radio Guernsey', 2),
                  ('BBC Local Radio Hereford & Worcester', 2), ('BBC Local Radio Humberside', 2),
                  ('BBC Local Radio Jersey', 2), ('BBC Local Radio Kent', 2), ('BBC Local Radio Lancashire', 2),
                  ('BBC Local Radio Leeds', 2), ('BBC Local Radio Leicester', 2), ('BBC Local Radio Lincolnshire', 2),
                  ('BBC Local Radio London', 2), ('BBC Local Radio Manchester', 2), ('BBC Local Radio Merseyside', 2),
                  ('BBC Local Radio Newcastle', 2), ('BBC Local Radio Norfolk', 2), ('BBC Local Radio Northampton', 2),
                  ('BBC Local Radio Oxford', 2), ('BBC Local Radio Sheffield', 2), ('BBC Local Radio Shropshire', 2),
                  ('BBC Local Radio Solent', 2), ('BBC Local radio Somerset', 2), ('BBC Local Radio Stoke', 2),
                  ('BBC Local Radio Suffolk', 2), ('BBC Local Radio Sussex', 2), ('BBC Local Radio Tees', 2),
                  ('BBC Local Radio Three Counties', 2), ('BBC Local Radio Wiltshire', 2), ('BBC Local Radio York', 2),
                  ('BBC Radio 1', 2), ('BBC Radio 2', 2), ('BBC Radio 3', 2), ('BBC Radio 4', 2), ('BBC Radio 5', 2),
                  ('BBC Radio 6', 2), ('BBC World Service', 2), ('BBC\xa0World\xa0Service News', 2),
                  ('British Comedy Radio GB', 2), ('Classic Fm', 2), ('Heart London', 2), ('Lbc London', 2),
                  ('LBC News', 2), ('Nottingham & Derby roots radio', 2), ('R.T.E. Raidió na Gaeltachta', 2),
                  ('RNIB Connect radio', 2), ('Smooth Scotland', 2), ('Talksport', 2), ('UCB UK', 2)]

books_expected = [('Project Gutenberg', 2), ('RNIB Reading Services', 2), ('Calibre Audio', 2)]
book = []

wb = load_workbook('Test Reports/Automated Test cases.xlsx')
ws = wb.active
user = os.environ['USERPROFILE']

queue = Queue(1)
timeout = 120
data = readFromQueue(queue)

ws['B3'].value = platform.release()
wb.save('Test Reports/Automated Test cases.xlsx')

set_date = datetime.datetime.now()
ws['B5'].value = set_date
wb.save('Test Reports/Automated Test cases.xlsx')

logging.basicConfig(filename="regression.log", level=logging.DEBUG)
delay()

#######################################################################################################################
                                           # MODULE: EMAIL START

def email_send1(queue):     # STATUS: TESTED
    # Logs In And Sends An Email
    delay()
    start = 0
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' in data:
        pag.write("guideautomation1@outlook.com", interval=0.5)
        pag.press("enter", interval=1)
        pag.write("testing", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.write("guideconnect1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(15)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", presses=3, interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Automation", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.write("test email", interval=0.5)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SENDING' in data:
                    data = readFromQueue(queue)

                if 'OK' in data:
                    pag.press("enter", interval=1)
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                elif 'GMENU_EMAIL_SENT_SUCCESS' in data:
                    pag.press("enter", interval=1)
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Emails 1: Failed to compose and send an email")
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 1: Failed to compose a new email")
                print("Emails 1: Failed to compose new email")
                print(data)
                ws['I156'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 1: Failed to sign into email address")
            print("Emails 1: failed to sign into email address")
            print(data)
            ws['I156'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 1: Failed to add email address")
        print("Emails 1: failed to add email address")
        print(data)
        ws['I156'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send2(queue):     # STATUS: TESTED
    # Checks sent folder for 'Sent' email
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    pag.press("s", interval=1)
    data = readFromQueue(queue)

    while 'Sent' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Sent' in data:
        pag.press("enter", interval=1)

        while 'guideautomation2@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation2@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.sleep(5)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)

            print('Emails 2: Check for the sent email via the Sent folder \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I157'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails 2: Failed to find sent email in Sent folder")
            print('Emails 2: Check for the sent email via the Sent folder \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I157'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 2: Failed to enter the sent folder")
        print("Emails 2: failed to enter the sent folder")
        print(data)
        ws['I157'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send3(queue):     # STATUS: TESTED
    # Log into an addition email account
    delay()
    start = 0
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'Email' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Email' in data:
        pag.press("enter", interval=1)
        pag.press("a", interval=1)
        data = readFromQueue(queue)

        if 'Account manager' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("testing", interval=0.5)
            pag.press("enter", presses=3, interval=1)
            pag.write("guideconnect2", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            pag.press("enter", interval=1)

            while 'Write new email' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Write new email' in data:
                print('Emails 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I158'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Emails 3: Failed to add a secondary email account via account manager")
                print('Emails 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I158'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 3: Failed to enter account manager")
            print("Emails 3: failed to enter account manager")
            print(data)
            ws['I158'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 3: Failed to navigate to email settings")
        print("Emails 3: failed to navigate to email settings")
        print(data)
        ws['I158'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send4(queue):     # STATUS: TESTED
    # Can you open a received email?
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        pag.sleep(5)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("enter")
            pag.sleep(10)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(5)

            print('Emails 4: Can you view a received email? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I159'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails 4: Failed to view and delete received email")
            print('Emails 4: Can you view a received email? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I159'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 4: Failed to locate guideautomation2@outlook.com")
        print("Emails 4: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I159'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send5(queue):     # STATUS: TESTED
    # Send a new email via address book
    delay()
    start = 0
    pag.press("enter", presses=3, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Select from address book' not  in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Select from address book' in data:
        pag.press("enter", interval=1)
        pag.press("e", interval=1)
        data = readFromQueue(queue)

        while 'Email Testing' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Email Testing' in data:
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.write("test", interval=0.5)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SENDING' in data:
                    data = readFromQueue(queue)
                    pag.sleep(3)

                if 'OK' in data:
                    returnhome()
                    pag.press("enter", interval=1)
                    pag.press("g", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation2@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation2@outlook.com' in data:
                        pag.press("enter", interval=1)
                        pag.press("i", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'GMENU_EMAIL_VIEW_MESSAGE' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_EMAIL_VIEW_MESSAGE' in data:
                                pag.sleep(10)
                                print('Emails 5: Send and view an email via Address book contact \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I160'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Emails 5: Failed to send and view and email via address contact")
                                print('Emails 5: Send and view an email via Address book contact \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I160'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Emails 5: Failed to find email sent via address book")
                            print("Emails 5: failed to find email sent via address book")
                            print(data)
                            ws['I160'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Emails 5: Failed to locate guideautomation2@outlook.com")
                        print("Emails 5: failed to locate guideautomation2@outlook.com")
                        print(data)
                        ws['I160'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Emails 5: Failed to send email")
                    print("Emails 5: failed to send email")
                    print(data)
                    ws['I160'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 5: Failed to open body of a new email")
                print("Emails 5: failed to open new message body")
                print(data)
                ws['I160'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 5: Failed to locate contact (Email Testing")
            print("Emails 5: failed to locate contact - Email Testing")
            print(data)
            ws['I160'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 5: Failed to locate select from address book")
        print("Emails 5: failed to locate select from address book")
        print(data)
        ws['I160'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send6(queue):     # STATUS: TESTED
    # add and remove recipients from both email address and contacts
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", presses=2, interval=2)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add or view recipients' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add or view recipients' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'No recipients' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'No recipients' in data:
                    pag.press("f2", interval=1)
                    pag.press("enter", presses=3, interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SEND_ACTION_RECIPIENTS_ACTION_NEW_ENTER' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_EMAIL_SEND_ACTION_RECIPIENTS_ACTION_NEW_ENTER' in data:
                        pag.write("test", interval=0.5)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while '(To) test' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if '(To) test' in data:
                            pag.press("f2", interval=1)
                            pag.press("enter", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", presses=2, interval=1)
                            data = readFromQueue(queue)

                            while 'No recipients' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'No recipients' in data:
                                pag.press("f2", interval=1)
                                pag.press("enter", presses=2, interval=1)
                                pag.press("s", interval=1)
                                pag.press("enter", interval=1)
                                pag.press("e", interval=1)
                                pag.press("enter", interval=1)
                                data = readFromQueue(queue)

                                while '(To) guideautomation2@outlook.com' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if '(To) guideautomation2@outlook.com' in data:
                                    print('Emails 6: Add and remove recipients to email from contact/email address\n'
                                          '>>> Result: PASS \n'
                                          '>>> Current String:', data)
                                    ws['I161'].value = 'PASS'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()

                                else:
                                    logging.critical("Emails 6: Failed to add/remove recipients from contact/email")
                                    print('Emails 6: Add and remove recipients to email from contact/email address\n'
                                          '>>> Result: FAIL \n'
                                          '>>> Current String:', data)
                                    ws['I161'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()

                            else:
                                logging.critical("Emails 6: Failed to remove a recipient")
                                print("Emails 6: failed to remove a recipient")
                                print(data)
                                ws['I161'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                pag.press("esc", presses=2, interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", interval=1)
                                returnhome()

                        else:
                            logging.critical("Emails 6: Failed to locate test(Added via email) in recipients")
                            print("Emails 6: test not present in recipients")
                            print(data)
                            ws['I161'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", presses=2, interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Emails 6: Failed to add recipient via email address")
                        print("Emails 6: failed to add recipient via email address")
                        print(data)
                        ws['I161'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=3, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:

                    logging.critical("Emails 6: Unexpected Recipient was present")
                    print("Emails 6: unexpected recipient was present")
                    print(data)
                    ws['I161'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:

                logging.critical("Emails 6: Failed to locate add or view recipients")
                print("Emails 6: failed to locate add or view recipients")
                print(data)
                ws['I161'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:

            logging.critical("Emails 6: Failed to enter body of the new email")
            print("Emails 6: failed to enter body of the new email")
            print(data)
            ws['I161'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 6: Failed to open recipient address")
        print("Emails 6: failed to add recipient address")
        print(data)
        ws['I161'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send7(queue):     # STATUS: TESTED
    # Change email subject line
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_SUBJECT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_SUBJECT' in data:
            pag.write("testing", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("c", interval=1)
            data = readFromQueue(queue)

            while 'Change subject' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Change subject' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SEND_ACTION_CHANGE_SUBJECT' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_SEND_ACTION_CHANGE_SUBJECT' in data:
                    pag.write("changed", interval=0.5)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                        print('Emails 7: Change the email Subject \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I162'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 7: Failed to change the email subject line")
                        print('Emails 7: Change the email Subject \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I162'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 7: Failed to open change subject line")
                    print("Emails 7: failed to open change subject")
                    print(data)
                    ws['I162'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 7: Failed to locate change subject line")
                print("Emails 7: failed to locate change subject")
                print(data)
                ws['I162'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 7: Failed to add a new subject line")
            print("Emails 7: failed to add a new subject")
            print(data)
            ws['I162'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 7: Failed to open recipient address")
        print("Emails 7: failed to add recipient address")
        print(data)
        ws['I162'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send8(queue):     # STATUS: TESTED
    # Complete a spell check inside of an email
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
            pag.write("dis iz a twest", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)

            while 'Spell Check' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Spell Check' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_NEW_MESSAGE_SPELLCHECK' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE_SPELLCHECK' in data:
                    pag.press("d", presses=3, interval=1)
                    pag.press("enter", interval=1)
                    pag.press("t", interval=1)
                    data = readFromQueue(queue)

                    while 'test' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'test' in data:
                        pag.press("enter", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                            print('Emails 8: Spell check an email  \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I163'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                        else:
                            logging.critical("Emails 8: Failed to spell check an email")
                            print('Emails 8: Spell check an email  \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I163'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Emails 8: Failed to change twest to test")
                        print("Emails 8: failed to spell check twest to test")
                        print(data)
                        ws['I163'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 8: Failed to open spell checker")
                    print("Emails 8: failed to open spell checker")
                    print(data)
                    ws['I163'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 8: Failed to locate spell checker")
                print("Emails 8: failed to locate spell check")
                print(data)
                ws['I163'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 8: Failed to enter body of the email via send to address contact")
            print("Emails 8: failed to enter body of email via send to address book contact")
            print(data)
            ws['I163'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 8: Failed to open recipient address")
        print("Emails 8: failed to add recipient address")
        print(data)
        ws['I163'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send9(queue):  # STATUS: TESTED
    # Save an email to drafts via F2
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("F2save2drafts@drafts.com", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("save to drafts", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Save to Drafts' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Save to Drafts' in data:
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.press("f", interval=1)
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                data = readFromQueue(queue)

                if 'Drafts' in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'F2save2drafts@drafts.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'F2save2drafts@drafts.com' in data:
                        print('Emails 9: Save a draft via F2 actions menu  \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I164'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("f2", interval=1)
                        pag.press("d", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", presses=2, interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 9: Failed to save draft via F2 actions menu")
                        print('Emails 9: Save a draft via F2 actions menu  \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I164'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Emails 9: Failed to locate drafts folder")
                    print("Emails 9: failed to locate drafts folder")
                    print(data)
                    ws['I164'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 9: Failed to save draft and return to body of the email")
                print("Emails 9: failed to save to drafts and return to body of email")
                print(data)
                ws['I164'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 9: Failed to locate save to drafts")
            print("Emails 9: failed to locate save to drafts")
            print(data)
            ws['I164'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 9: Failed to open recipient address")
        print("Emails 9: failed to add recipient address")
        print(data)
        ws['I164'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send10(queue):        # STATUS: TESTED
    # Save email to draft via ESC
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("ESCsave2drafts@drafts.com", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("draft email via ESC", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'ESCsave2drafts@drafts.com'not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'ESCsave2drafts@drafts.com' in data:
            print('Emails 10: Save a draft via ESC method  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I165'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            returnhome()

        else:
            logging.critical("Emails 10: Failed to save a draft via ESC")
            print('Emails 10: Save a draft via ESC method  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I165'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 10: Failed to open recipient address")
        print("Emails 10: failed to add recipient address")
        print(data)
        ws['I165'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send11(queue):        # STATUS: TESTED
    # Adds and removes an attachment from an email
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("guideautomation2@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("attachments", interval=0.5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("test", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("a", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Add or view attachments' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Add or view attachments' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'No Attachments' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No Attachments' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("f2", interval=1)
                    pag.press("r", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)

                    while 'No Attachments' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'No Attachments' in data:
                        print('Emails 11: Adds and removes an attachment from an email  \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I166'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 11: Failed to add and remove an attachment from an email")
                        print('Emails 11: Adds and removes an attachment from an email  \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I166'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 11: Failed to add attachment to an email")
                    print("Emails 11: failed to add attachment to an email")
                    print(data)
                    ws['I166'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 11: No Attachments Prompt is missing")
                print("Emails 11: Expected No Attachments prompt is missing")
                print(data)
                ws['I166'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 11: Failed to locate add or view attachments")
            print("Emails 11: failed to locate add or view attachments")
            print(data)
            ws['I166'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 11: Failed to open recipient address")
        print("Emails 11: failed to add recipient address")
        print(data)
        ws['I166'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send12(queue):    # STATUS: TESTED
    # Add all test data attachments to an email and check its received correctly
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("guideautomation2@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("attachments", interval=0.5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("test", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("a", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Add or view attachments' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Add or view attachments' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)

            while 'testdata' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'testdata' in data:
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("j", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("p", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("p", presses=2, interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("r", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("x", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("right", interval=1)
                    data = readFromQueue(queue)

                    while 'JPEG.jpg' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'JPEG.jpg' in data:
                        pag.press("right", interval=1)
                        data = readFromQueue(queue)

                        while 'PDF.pdf' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'PDF.pdf' in data:
                            pag.press("right", interval=1)
                            data = readFromQueue(queue)

                            while 'PNG.png' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'PNG.png' in data:
                                pag.press("right", interval=1)
                                data = readFromQueue(queue)

                                while 'RTF.rtf' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if 'RTF.rtf' in data:
                                    pag.press("right", interval=1)
                                    data = readFromQueue(queue)

                                    while 'TXT.txt' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'TXT.txt' in data:
                                        pag.press("right", interval=1)
                                        data = readFromQueue(queue)

                                        while 'XDOC.docx' not in data:
                                            data = readFromQueue(queue)
                                            if start < timeout:
                                                pag.sleep(10)
                                                start += 10
                                                if start == timeout:
                                                    print(data, ": Timed out after", timeout, "Seconds")
                                                    returnhome()
                                                    break

                                        if 'XDOC.docx' in data:
                                            pag.press("esc", interval=1)
                                            pag.press("f2", interval=1)
                                            pag.press("enter", interval=1)
                                            data = readFromQueue(queue)

                                            while 'GMENU_EMAIL_SENDING' in data:
                                                data = readFromQueue(queue)
                                                pag.sleep(2)

                                            if 'OK' in data:
                                                pag.press("enter", interval=1)
                                                pag.sleep(2)
                                                pag.press("esc", interval=1)
                                                pag.press("g", interval=1)
                                                data = readFromQueue(queue)

                                                while 'guideautomation2@outlook.com' not in data:
                                                    data = readFromQueue(queue)
                                                    if start < timeout:
                                                        pag.sleep(10)
                                                        start += 10
                                                        if start == timeout:
                                                            print(data, ": Timed out after", timeout, "Seconds")
                                                            returnhome()
                                                            break

                                                if 'guideautomation2@outlook.com' in data:
                                                    pag.press("enter", interval=1)
                                                    pag.press("i", interval=1)
                                                    pag.sleep(60)
                                                    pag.press("enter", interval=1)
                                                    data = readFromQueue(queue)

                                                    while 'guideautomation1@outlook.com' not in data:
                                                        data = readFromQueue(queue)
                                                        if start < timeout:
                                                            pag.sleep(10)
                                                            start += 10
                                                            if start == timeout:
                                                                print(data, ": Timed out after", timeout, "Seconds")
                                                                returnhome()
                                                                break

                                                    if 'guideautomation1@outlook.com' in data:
                                                        pag.press("f2", interval=1)
                                                        pag.press("v", interval=1)
                                                        pag.press("enter", interval=1)
                                                        data = readFromQueue(queue)

                                                        while 'DOC.doc' not in data:
                                                            data = readFromQueue(queue)
                                                            if start < timeout:
                                                                pag.sleep(10)
                                                                start += 10
                                                                if start == timeout:
                                                                    print(data, ": Timed out after", timeout, "Seconds")
                                                                    returnhome()
                                                                    break

                                                        if 'DOC.doc' in data:  # might need a condition for Archive
                                                            pag.press("esc", interval=1)
                                                            pag.press("f2", interval=1)
                                                            pag.press("m", interval=1)
                                                            pag.press("enter", presses=3, interval=1)
                                                            pag.sleep(5)
                                                            print('Emails 12: Sends all Attachments Types via Email  \n'
                                                                  '>>> Result: PASS \n'
                                                                  '>>> Current String:', data)
                                                            ws['I167'].value = 'PASS'
                                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                                            returnhome()

                                                        else:
                                                            logging.critical(
                                                                "Emails 12: Failed to send attachments via email")
                                                            print('Emails 12: Sends all Attachments Types via Email  \n'
                                                                  '>>> Result: FAIL \n'
                                                                  '>>> Current String:', data)
                                                            ws['I167'].value = 'FAIL'
                                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                                            returnhome()
                                                    else:
                                                        logging.critical(
                                                            "Emails 12: Failed to find email from guide automation 1"
                                                                        )
                                                        print("Emails 12: failed to find email"
                                                              " from guideautomation1@outlook.com")
                                                        print(data)
                                                        ws['I167'].value = 'FAIL'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()

                                                else:
                                                    logging.critical("Emails 12: Failed to locate "
                                                                     "guide automation 2")
                                                    print("Emails 12: failed to locate guideautomation2@outlook.com")
                                                    print(data)
                                                    ws['I167'].value = 'FAIL'
                                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                                    returnhome()
                                        else:
                                            logging.critical("Emails 12: Failed to locate attached DOCX")
                                            print("Emails 12: failed to locate DOCX")
                                            print(data)
                                            ws['I167'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            pag.press("esc", presses=2, interval=1)
                                            pag.press("right", interval=1)
                                            pag.press("enter", interval=1)
                                            returnhome()

                                    else:
                                        logging.critical("Emails 12: Failed to locate attached TXT")
                                        print("Emails 12: failed to locate TXT")
                                        print(data)
                                        ws['I167'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        pag.press("esc", presses=2, interval=1)
                                        pag.press("right", interval=1)
                                        pag.press("enter", interval=1)
                                        returnhome()

                                else:
                                    logging.critical("Emails 12: Failed to locate attached RTF")
                                    print("Emails 12: failed to locate RTF")
                                    print(data)
                                    ws['I167'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()
                            else:
                                logging.critical("Emails 12: Failed to locate attached PNG")
                                print("Emails 12: failed to locate PNG")
                                print(data)
                                ws['I167'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                pag.press("esc", presses=2, interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", interval=1)
                                returnhome()
                        else:
                            logging.critical("Emails 12: Failed to locate attached PDF")
                            print("Emails 12: failed to locate PDF")
                            print(data)
                            ws['I167'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", presses=2, interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()
                    else:
                        logging.critical("Emails 12: Failed to locate attached JPG")
                        print("Emails 12: failed to locate JPEG")
                        print(data)
                        ws['I167'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()
                else:
                    logging.critical("Emails 12: Failed to locate attached DOC")
                    print("Emails 12: failed to locate DOC")
                    print(data)
                    ws['I167'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()
            else:
                logging.critical("Emails 12: Failed to locate test data folder")
                print("Emails 12: failed to locate test data folder")
                print(data)
                ws['I167'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=3, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 12: Failed to locate add or view attachments")
            print("Emails 12: failed to locate add or view attachments")
            print(data)
            ws['I167'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 12: Failed to open recipient address")
        print("Emails 12: failed to add recipient address")
        print(data)
        ws['I167'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        # wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive1(queue):      # STATUS: TESTED
    # Manually check for new emails
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        pag.press("f2", interval=2)
        pag.sleep(5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        while 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK' in data:
            data = readFromQueue(queue)

        if 'OK' in data:
            pag.sleep(5)
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK_COMPLETE' in data:
            pag.sleep(5)
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 1: Failed to manually check emails")
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 1: Failed to locate guideautomation2@outlook.com")
        print("Email Receive 1: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I168'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive2(queue):      # STATUS: TESTED
    # reply to an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            if 'Reply' in data:
                pag.press("enter", interval=1)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("reply email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 2: Reply to an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I169'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 2: Failed to reply to an email")
                                print('Email Receive 2: Reply to an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I169'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 2: Failed to locate guide automation1")
                            print("Email Receive 2: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I169'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 2: Failed to send an email")
                        print("Email Receive 2: failed to send email")
                        print(data)
                        ws['I169'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Receive 2: Failed to enter body of an email")
                    print("Email Receive 2: failed to enter body of email")
                    print(data)
                    ws['I169'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 2: Failed to locate reply in actions menu")
                print("Email Receive 2: failed to locate reply in actions menu")
                print(data)
                ws['I169'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 2: Failed to locate guide automation 1 in inbox")
            print("Email Receive 2: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I169'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 2: Failed to locate guide automation 2")
        print("Email Receive 2: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I169'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive3(queue):  # STATUS: TESTED
    # reply to all recipients in email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("r", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'Reply All' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Reply All' in data:
                pag.press("enter", interval=1)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("reply all email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 3: Reply All to an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I170'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 3: Failed to reply All to an email")
                                print('Email Receive 3: Reply All to an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I170'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 3: Failed to locate guide automation1")
                            print("Email Receive 3: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I170'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 3: Failed to send an email")
                        print("Email Receive 3: failed to send email")
                        print(data)
                        ws['I170'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Receive 3: Failed to enter body of an email")
                    print("Email Receive 3: failed to enter body of email")
                    print(data)
                    ws['I170'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 3: Failed to locate reply all in actions menu")
                print("Email Receive 3: failed to locate reply all in actions menu")
                print(data)
                ws['I170'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Email Receive 3: Failed to locate guide automation 1 in inbox")
            print("Email Receive 3: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I170'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 3: Failed to locate guide automation 2")
        print("Email Receive 3: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I170'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive4(queue):  # STATUS: TESTED
    # forward an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("f", interval=1)
            data = readFromQueue(queue)

            while 'Forward' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Forward' in data:
                pag.press("enter", presses=2, interval=1)
                pag.sleep(10)
                pag.write("guideautomation1@outlook.com", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("forward email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 4: Forward an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I171'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 4: Failed to forward an email")
                                print('Email Receive 4: Forward an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I171'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 4: Failed to locate guide automation1")
                            print("Email Receive 4: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I171'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 4: Failed to send an email")
                        print("Email Receive 4: failed to send email")
                        print(data)
                        ws['I171'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Email Receive 4: Failed to enter body of an email")
                    print("Email Receive 4: failed to enter body of email")
                    print(data)
                    ws['I171'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 4: Failed to locate forward in actions menu")
                print("Email Receive 4: failed to locate forward in actions menu")
                print(data)
                ws['I171'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Email Receive 4: Failed to locate guide automation 1 in inbox")
            print("Email Receive 4: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I171'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 4: Failed to locate guide automation 2")
        print("Email Receive 4: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I171'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive5(queue):  # STATUS: TESTED
    # Add email contact to address book
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add sender to address book' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add sender to address book' in data:
                pag.press("enter", interval=1)
                pag.write("email", interval=0.5)
                pag.press("enter", interval=1)
                pag.write("contact", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                returnhome()
                pag.press("a", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("e", interval=1)
                data = readFromQueue(queue)

                while 'email contact' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'email contact' in data:
                    print('Email Receive 5: Add email contact to address book \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I172'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Receive 5: Failed to add a email contact to address book")
                    print('Email Receive 5: Add email contact to address book \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I172'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 5: Failed to locate Add contact in actions menu")
                print("Email Receive 5: failed to locate Add contact in actions menu")
                print(data)
                ws['I172'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 5: Failed to locate guide automation 1 in inbox")
            print("Email Receive 5: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I172'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 5: Failed to locate guide automation 2")
        print("Email Receive 5: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I172'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive6(queue):  # STATUS: TESTED
    # Move email to a different folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=2)
            pag.press("m", interval=2)
            data = readFromQueue(queue)

            while 'Move to folder' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Move to folder' in data:
                pag.press("enter", interval=1)
                pag.press("n", interval=1)
                data = readFromQueue(queue)

                while 'Notes' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Notes' in data:
                    pag.press("enter", presses=2, interval=1)
                    pag.sleep(3)
                    pag.press("esc", interval=1)
                    pag.press("f", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("n", interval=1)
                    pag.press("enter", interval=1)
                    pag.sleep(3)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        print('Email Receive 6: Move email to a different folder \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I173'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Email Receive 6: Failed to move an email to Notes folder")
                        print('Email Receive 6: Move email to a different folder \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I173'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

            else:
                logging.critical("Email Receive 6: Failed to locate Move email in actions menu")
                print("Email Receive 6: failed to locate Add contact in actions menu")
                print(data)
                ws['I173'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 6: Failed to locate guide automation 1 in inbox")
            print("Email Receive 6: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I173'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 6: Failed to locate guide automation 2")
        print("Email Receive 6: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I173'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive7(queue):  # STATUS: TESTED
    # View email attachments (JPG, PDF, DOC - All viewers)
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Folders' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Folders' in data:
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                pag.press("f2", interval=1)
                pag.press("v", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("enter", interval=1)

                    while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                        pag.press("esc", interval=1)
                        pag.press("j", interval=1)
                        data = readFromQueue(queue)

                        while 'JPEG.jpg' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'JPEG.jpg' in data:
                            pag.press("enter", interval=1)

                            while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                                pag.press("esc", interval=1)
                                pag.press("p", interval=1)
                                data = readFromQueue(queue)

                                while 'PDF.pdf' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if 'PDF.pdf' in data:
                                    pag.press("enter", interval=1)

                                    while 'GMENU_ENTERTAINMENT_PDFVIEWER' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'GMENU_ENTERTAINMENT_PDFVIEWER' in data:
                                        print('Email Receive 7: View received attachments via email \n'
                                              '>>> Result: PASS \n'
                                              '>>> Current String:', data)
                                        ws['I174'].value = 'PASS'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()

                                    else:
                                        logging.critical("Email Receive 7: Failed to view attached emails")
                                        print('Email Receive 7: View received attachments via email \n'
                                              '>>> Result: FAIL \n'
                                              '>>> Current String:', data)
                                        ws['I174'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()

                                else:
                                    logging.critical("Email Receive 7: Failed to locate attached PDF")
                                    print("Email Receive 7: failed to locate attached PDF")
                                    print(data)
                                    ws['I174'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                        else:
                            logging.critical("Email Receive 7: Failed to locate attached JPEG")
                            print("Email Receive 7: failed to locate attached JPEG")
                            print(data)
                            ws['I174'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                else:
                    logging.critical("Email Receive 7: Failed to locate attached DOC")
                    print("Email Receive 7: failed to locate attached DOC")
                    print(data)
                    ws['I174'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 7: Failed to locate attachment email in archive folder")
                print("Emails Receive 7: failed to locate attachment email via archive folder")
                print(data)
                ws['I174'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails Receive 7: Failed to locate folder")
            print("Emails Receive 7: failed to locate folders")
            print(data)
            ws['I174'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails Receive 7: Failed to locate guide automation 2")
        print("Email Receive 7: failed to locate guide automation 2")
        print(data)
        ws['I174'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive8(queue):      # STATUS: TESTED
    # Blocks a sender via actions menu
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        pag.press("k", interval=1)
        data = readFromQueue(queue)

        while 'Kieran.Baker@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Kieran.Baker@yourdolphin.com' in data:
            pag.press("f2", interval=1)
            pag.press("b", interval=1)
            data = readFromQueue(queue)

            while 'Block sender' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Block sender' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'Kieran.Baker@yourdolphin.com' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Kieran.Baker@yourdolphin.com' in data:
                    pag.press("enter", presses=2, interval=1)
                    print('Email Receive 8: Block email contact via actions menu \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I175'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Receive 8: Failed to block email contact via actions menu")
                    print('Email Receive 8: Block email contact via actions menu \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I175'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 8: Failed to locate blocked sender")
                print("Email Receive 8: failed to locate Add contact in actions menu")
                print(data)
                ws['I175'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Email Receive 8: Failed to locate kieran.baker@yourdolphin.com in inbox")
            print("Email Receive 8: failed to locate 'kieran.baker@yourdolphin.com' in inbox")
            print(data)
            ws['I175'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 8: Failed to locate guide automation 2")
        print("Email Receive 8: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I175'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive9(queue):  # STATUS: TESTED
    # View details of a received email
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if 'GMENU_EMAIL_VIEW_DETAILS' in data:
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'Date' in data:
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 9: Failed to view details of a received email")
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 9: Failed to locate details in the actions menu")
        print("Email Receive 9: failed to locate details via actions menu")
        print(data)
        ws['I176'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive10(queue):     # STATUS: TESTED
    # Mark email as Read / Unread
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("m", presses=2, interval=1)
    data = readFromQueue(queue)

    if 'Mark as read' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("m", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'Mark as unread' in data:
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails Receive 10: Failed to set email to read / unread")
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    elif 'Mark as unread' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("m", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'Mark as read' in data:
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails Receive 10: Failed to set email to read / unread")
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 10: Failed to locate an unopened email")
        print("Email Receive 10: failed to locate unread email")
        print(data)
        ws['I177'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive11(queue):     # STATUS: TESTED
    # Search for specific email
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("s", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_VIEW_ACTION_SEARCH' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_VIEW_ACTION_SEARCH' in data:
        pag.write("support@yourdolphin.com", interval=0.5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'support@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'support@yourdolphin.com' in data:
            print('Email Receive 11: Search for specific email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I178'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 11: Failed to search for 'Support'")
            print('Email Receive 11: Search for specific email \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I178'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 11: Failed to open search email function")
        print("Email Receive 11: failed to open search email function")
        print(data)
        ws['I178'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive12(queue):     # STATUS: TESTED
    # Delete an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation1@outlook.com' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        pag.sleep(5)

    if 'guideautomation1@outlook.com' not in data:
        print('Email Receive 12: Delete email \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I179'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Email Receive 12: Failed to Delete Email(s) from Guide Automation 1")
        print('Email Receive 12: Delete email \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I179'].value = 'fail'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive13(queue):     # STATUS: TESTED
    # Save attachments from email and open via My saved attachments
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'Archive' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Archive' in data:
            pag.press("enter", interval=1)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                pag.press("f2", interval=1)
                pag.press("v", interval=1)
                pag.press("enter", interval=1)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("f2", interval=1)
                    pag.press("s", interval=1)
                    pag.press("enter", presses=3, interval=1)
                    pag.press("j", interval=1)
                    data = readFromQueue(queue)

                    if 'JPEG.jpg' in data:
                        pag.press("f2", interval=1)
                        pag.press("s", interval=1)
                        pag.press("enter", presses=3, interval=1)
                        pag.press("p", interval=1)
                        data = readFromQueue(queue)

                        if 'PDF.pdf' in data:
                            pag.press("f2", interval=1)
                            pag.press("s", interval=1)
                            pag.press("enter", presses=3, interval=1)
                            pag.press("esc", presses=3, interval=1)

                            while 'My saved attachments' not in data:
                                pag.press("m", interval=1)
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'My saved attachments' in data:
                                pag.press("enter", interval=1)
                                pag.press("d", interval=1)
                                data = readFromQueue(queue)
                                if 'DOC' in data:
                                    pag.press("enter", interval=1)

                                    while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                                        pag.press("esc", interval=1)
                                        pag.press("j", interval=1)
                                        data = readFromQueue(queue)
                                        if 'JPEG' in data:
                                            pag.press("enter", interval=1)

                                            while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                                                data = readFromQueue(queue)
                                                if start < timeout:
                                                    pag.sleep(10)
                                                    start += 10
                                                    if start == timeout:
                                                        print(data, ": Timed out after", timeout, "Seconds")
                                                        returnhome()
                                                        break

                                            if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                                                pag.press("esc", interval=1)
                                                pag.press("p", interval=1)
                                                data = readFromQueue(queue)
                                                if 'PDF' in data:
                                                    pag.press("enter", interval=1)

                                                    while 'GMENU_ENTERTAINMENT_PDFVIEWER' not in data:
                                                        data = readFromQueue(queue)
                                                        if start < timeout:
                                                            pag.sleep(10)
                                                            start += 10
                                                            if start == timeout:
                                                                print(data, ": Timed out after", timeout, "Seconds")
                                                                returnhome()
                                                                break

                                                    if 'GMENU_ENTERTAINMENT_PDFVIEWER' in data:
                                                        print('Email Receive 13: Open Saved attachments \n'
                                                              '>>> Result: PASS \n'
                                                              '>>> Current String:', data)
                                                        ws['I180'].value = 'PASS'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()

                                                    else:
                                                        logging.critical("Email Receive 13: "
                                                                         "Failed to open all saved attachments")
                                                        print('Email Receive 13: Open Saved attachments \n'
                                                              '>>> Result: FAIL \n'
                                                              '>>> Current String:', data)
                                                        ws['I180'].value = 'FAIL'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()
                                                else:
                                                    logging.critical("Email Receive 13:"
                                                                     "Failed to locate saved attachment PDF")
                                                    print("Email Receive 13: failed to locate saved attachment PDF")
                                                    print(data)
                                                    ws['I180'].value = 'FAIL'
                                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                                    returnhome()
                                            else:
                                                logging.critical("Email Receive 13: Failed to open image viewer")
                                                print("Email Receive 13: failed to open Image Viewer")
                                                print(data)
                                                ws['I180'].value = 'FAIL'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()
                                        else:
                                            logging.critical("Email Receive 13: Failed to locate"
                                                             "saved attachment JPEG")
                                            print("Email Receive 13: failed to locate saved attachment JPEG")
                                            print(data)
                                            ws['I180'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            returnhome()
                                    else:
                                        logging.critical("Email Receive 13: Failed to open Doc Viewer")
                                        print("Email Receive 13: failed to open Doc Viewer")
                                        print(data)
                                        ws['I180'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()
                                else:
                                    logging.critical("Email Receive 13: Failed to locate"
                                                     "saved attachment DOC")
                                    print("Email Receive 13: failed to locate saved attachment DOC")
                                    print(data)
                                    ws['I180'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()
                            else:
                                logging.critical("Email Receive 13: Failed to locate"
                                                 "my saved attachments")
                                print("Email Receive 13: failed to locate my saved attachments")
                                print(data)
                                ws['I180'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()
                        else:
                            logging.critical("Email Receive 13: Failed to locate"
                                             "attached PDF")
                            print("Email Receive 13: failed to locate attached PDF")
                            print(data)
                            ws['I180'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()
                    else:
                        logging.critical("Email Receive 13: Failed to locate attached JPEG")
                        print("Email Receive 13: failed to locate attached JPEG")
                        print(data)
                        ws['I180'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Email Receive 13: Failed to locate attached DOC")
                    print("Email Receive 13: failed to locate attached DOC file")
                    print(data)
                    ws['I180'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Email Receive 13: Failed to locate "
                                 "guide automation 1 email with attachments")
                print("Email Receive 13: failed to locate guide automation1 email with attachments")
                print(data)
                ws['I180'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Email Receive 13: Failed to locate archive folder")
            print("Email Receive 13: failed to locate Archive folder")
            print(data)
            ws['I180'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Email Receive 13: Failed to locate 'Guideautomation2@outlook.com")
        print("Email Receive 13: failed to locate 'guideautomation2@outlook.com'")
        print(data)
        ws['I180'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments1(queue):      # STATUS: TESTED
    # Sort documents
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'DOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'DOC' in data:
            pag.press("f2", interval=1)
            pag.press("s", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)

            while 'Sort by date' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Sort by date' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                data = readFromQueue(queue)

                while 'PDF' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'PDF' in data:
                    print('Email Attachments 1: Sort attachments \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I181'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Attachments 1: Failed to sort attachments by date")
                    print('Email Attachments 1: Sort attachments \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I181'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Attachments 1: Failed to locate sort by date")
                print("Email Attachments 1: failed to locate sort by date")
                print(data)
                ws['I181'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 1: Failed to locate saved attachment DOC")
            print("Email Attachments 1: failed to locate saved attachment DOC")
            print(data)
            ws['I181'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 1: Failed to locate my saved attachments folder")
        print("Email Attachments 1: failed to locate my saved attachments folder")
        print(data)
        ws['I181'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments2(queue):      # STATUS: TESTED
    # Search specific saved document
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("jpeg", interval=0.5)
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)

        while 'JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'JPEG' in data:
            print('Email Attachments 2: Search attachments \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I182'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Attachments 2: Failed to search attachments for JPEG")
            print('Email Attachments 2: Search attachments \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I182'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 2: Failed to locate my saved attachments folder")
        print("Email Attachments 2: failed to locate my saved attachments folder")
        print(data)
        ws['I182'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments3(queue):      # STATUS: TESTED
    # rename saved email attachment
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)

        while 'JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'renamed JPEG' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'renamed JPEG' in data:
                print('Email Attachments 3: Rename saved document \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I183'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Attachments 3: Failed to rename a saved attachment")
                print('Email Attachments 3: Rename saved document \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I183'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 3: Failed to locate saved attachment JPG")
            print("Email Attachments 3: failed to locate saved attachment JPG")
            print(data)
            ws['I183'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 3: Failed to locate my saved attachments folder")
        print("Email Attachments 3: failed to locate my saved attachments folder")
        print(data)
        ws['I183'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments4(queue):      # STATUS: TESTED
    # delete saved attachments
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'renamed JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'renamed JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'Go to folder - GuideConnect' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Go to folder - GuideConnect' in data:
                print('Email Attachments 4: Delete saved document \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I184'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Attachments 4: Failed to delete saved document rename JPG")
                print('Email Attachments 4: Delete saved document \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I184'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 4: Failed to locate saved attachment renamed JPG")
            print("Email Attachments 4: failed to locate saved attachment JPG")
            print(data)
            ws['I184'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 4: Failed to locate my saved attachments folder")
        print("Email Attachments 4: failed to locate my saved attachments folder")
        print(data)
        ws['I184'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_folders1(queue):       # STATUS: TESTED
    # Add a custom folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Folders' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Folders' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_FOLDERS_ACTION_CREATE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_FOLDERS_ACTION_CREATE' in data:
            pag.write("added", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'added' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'added' in data:
                print('Email Folders 1: Create a new folder \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I185'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Folders 1: Failed to create new folder 'added'")
                print('Email Folders 1: Create a new folder \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I185'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Folders 1: Failed to open create a new folder")
            print("Email Folders 1: failed to open create new folder")
            print(data)
            ws['I185'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Folders 1: Failed to locate Folders folder")
        print("Email Folders 1: failed to locate Folders folder")
        print(data)
        ws['I185'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_folders2(queue):      # STATUS: TESTED
    # delete a custom folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Folders' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Folders' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'added' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'added' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            if 'added' not in data:
                print('Email Folders 2: delete a folder \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I186'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Folders 2: Failed to delete a custom folder")
                print('Email Folders 2: delete a folder \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I186'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Folders 2: Failed to locate added folder 'added' ")
            print("failed to locate added folder 'added'")
            print(data)
            ws['I186'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Folders 2: Failed to locate Folders folder")
        print("Email Folders 2: failed to locate Folders folder")
        print(data)
        ws['I186'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked1(queue):      # STATUS: TESTED
    # edit an existing blocked contact
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'Kieran.Baker@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Kieran.Baker@yourdolphin.com' in data:
            pag.press("enter", interval=1)
            pag.write("changed", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'changed' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'changed' in data:
                print('Email Blocked 1: Edit an existing blocked contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I187'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 1: Failed to edit an existing blocked contact")
                print('Email Blocked 1: Edit an existing blocked contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I187'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 1: Failed to locate previous blocked contact"
                             "via inbox actions menu")
            print("Email Blocked 1: failed to locate added entry from inbox actions menu")
            print(data)
            ws['I187'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 1: Failed to locate Blocked senders folder")
        print("Email Blocked 1: failed to locate Blocked senders folder")
        print(data)
        ws['I187'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked2(queue):      # STATUS: TESTED
    # remove a blocked contact
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'changed' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'changed' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            pag.press("enter", presses=3, interval=1)
            data = readFromQueue(queue)

            while 'No items found' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No items found' in data:
                print('Email Blocked 2: Remove a blocked contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I188'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 1: Failed to remove an existing blocked contact")
                print('Email Blocked 1: Remove a blocked contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I188'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 2: Failed to locate previous blocked contact"
                             "via inbox actions menu")
            print("Email Blocked 2: failed to locate added entry from inbox actions menu")
            print(data)
            ws['I188'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 2: Failed to locate Blocked senders folder")
        print("Email Blocked 2: failed to locate Blocked senders folder")
        print(data)
        ws['I188'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked3(queue):      # STATUS: NOT TESTED
    # Add a blocked sender
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_BLOCKED_SENDERS_ACTION_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_BLOCKED_SENDERS_ACTION_ADD' in data:
            pag.write("guideautomation1@outlook.com", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                print('Email Blocked 3: Add a blocked contact via blocked contacts \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I189'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 3: Failed to add a blocked contact "
                                 "via blocked contacts feature")
                print('Email Blocked 3: Add a blocked contact via blocked contacts \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I189'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 3: Failed to open add blocked sender via blocked senders")
            print("Email Blocked 3: Failed to open add blocked sender via blocked senders")
            print(data)
            ws['I189'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 3: Failed to locate Blocked senders folder")
        print("Email Blocked 3: failed to locate Blocked senders folder")
        print(data)
        ws['I189'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked4(queue):      # STATUS: NOT TESTED
    # Check blocked senders folder for blocked emails
    start = 0
    delay()
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.write("guideautomation2@outlook.com", interval=0.5)
        pag.press("enter", interval=1)
        pag.write("blocked", interval=0.5)
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_SENDING' in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            pag.press("g", interval=1)
            data = readFromQueue(queue)
            if 'guideautomation2@outlook.com' in data:
                pag.press("enter", interval=1)
                pag.press("f", interval=1)
                pag.press("enter", interval=1)
                pag.press("b", interval=1)
                data = readFromQueue(queue)
                if 'Blocked Senders' in data:
                    pag.sleep(30)
                    pag.press("enter", interval=1)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.press("del", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", presses=2, interval=1)
                        data = readFromQueue(queue)
                        if 'No emails' in data:
                            print('Email Blocked 4: Blocked senders emails go to blocked folder \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I190'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Email Blocked 4: Failed to locate blocked sender email"
                                             "in blocked senders folder")
                            print('Email Blocked 4: Blocked senders emails go to blocked folder \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I190'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Blocked 4: Failed to locate expected blocked email in"
                                         "the blocked senders folder")
                        print("Email Blocked 4: failed to locate expected blocked email in blocked folder")
                        print(data)
                        ws['I190'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Blocked 4: failed to locate blocked senders folder")
                    print("Email Blocked 4: failed to locate blocked senders folder")
                    print(data)
                    ws['I190'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Blocked 4: Failed to locate guide automation 2 in email accounts")
                print("Email Blocked 4: failed to locate guide automation 2")
                print(data)
                ws['I190'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 4: Failed to send an email")
            print("Email Blocked 4: failed to send email")
            print(data)
            ws['I190'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 4: Failed to enter recipient email address")
        print("Email Blocked 4: failed to enter recipient email address")
        print(data)
        ws['I190'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def emailClean1(queue):
    delay()
    start = 0
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'guideautomation1@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)
        if 'Sent' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)

            while 'No emails' not in data:
                data = readFromQueue(queue)
                pag.sleep(3)
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

            if 'No emails' in data:
                logging.info("emailClean1: Email Sent folder cleaned")
                print("emailClean1: Email Sent folder cleaned")
                returnhome()

            else:
                logging.info("emailClean1: Failed to clean email sent folder")
                print("emailClean1: failed to clean email sent folder")
                returnhome()

        else:
            logging.critical("emailClean1: Failed to locate sent items folder")
            print("emailClean1: failed to locate sent items folder")
            returnhome()
    else:
        logging.critical("emailClean1: Failed to locate guideautomation1")
        print("emailClean1: failed to locate 'guideautomation1@outlook.com'")
        returnhome()


def emailClean2(queue):
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)
    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Archive' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)

            while 'No emails' not in data:
                data = readFromQueue(queue)
                pag.sleep(3)
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

            if 'No emails' in data:
                pag.press("esc", interval=1)
                pag.press("i", interval=1)
                data = readFromQueue(queue)
                if 'Inbox' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)
                    data = readFromQueue(queue)

                    while 'Kieran.Baker@yourdolphin.com' not in data:
                        data = readFromQueue(queue)
                        pag.press("f2", interval=1)
                        pag.press("d", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", presses=2, interval=1)
                        pag.sleep(3)
                        data = readFromQueue(queue)

                    if 'Kieran.Baker@yourdolphin.com' in data:
                        pag.press("esc", interval=1)
                        pag.press("n", interval=1)
                        data = readFromQueue(queue)
                        if 'Notes' in data:
                            pag.press("enter", interval=1)
                            pag.sleep(5)
                            data = readFromQueue(queue)

                            while 'No emails' not in data:
                                data = readFromQueue(queue)
                                pag.press("f2", interval=1)
                                pag.press("d", interval=1)
                                pag.press("enter", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)
                                data = readFromQueue(queue)
                                pag.sleep(3)

                            if 'No emails' in data:
                                pag.press("esc", interval=1)
                                pag.press("s", interval=1)
                                data = readFromQueue(queue)
                                if 'Sent' in data:
                                    pag.press("enter", interval=1)
                                    pag.sleep(5)
                                    data = readFromQueue(queue)

                                    while 'No emails' not in data:
                                        data = readFromQueue(queue)
                                        pag.sleep(3)
                                        pag.press("f2", interval=1)
                                        pag.press("d", interval=1)
                                        pag.press("enter", interval=1)
                                        pag.press("right", interval=1)
                                        pag.press("enter", presses=2, interval=1)

                                    if 'No emails' in data:
                                        logging.info("emailClean2: All email folders Cleaned")
                                        print("emailClean2: All email folders Cleaned")
                                        returnhome()

                                    else:
                                        logging.critical(
                                        "emailClean2: Failed to clean Sent items folder")
                                        print("emailClean2: Failed to clean Sent items folder")
                                        returnhome()

                                else:
                                    logging.critical(
                                    "emailClean2: Failed to locate sent items folder")
                                    print("emailClean2: failed to locate Sent items folder")
                                    returnhome()

                            else:
                                logging.critical("emailClean2: Failed to clear notes folder")
                                print("emailClean2: failed to clear notes folder")
                                returnhome()

                        else:
                            logging.critical("emailClean2: Failed to locate notes folder")
                            print("emailClean2: failed to locate notes folder")
                            returnhome()

                    else:
                        logging.critical("emailClean2: Failed to clean inbox folder")
                        print("emailClean2: Failed to clean inbox folder")
                        returnhome()

                else:
                    logging.critical("emailClean2: Failed to locate inbox folder")
                    print("emailClean2: failed to locate inbox folder")
                    returnhome()
            else:
                logging.critical("emailClean2: Failed to clean archive folder")
                print("emailClean2: failed to clean archive folder")
                returnhome()
        else:
            logging.critical("emailClean2: Failed to locate Archive folder")
            print("emailClean2: failed to locate Archive folder")
            returnhome()
    else:
        logging.critical("emailClean2: Failed to locate guideautomation2")
        print("emailClean2:failed to locate 'guideautomation2@outlook.com'")
        returnhome()

                                            # MODULE: EMAIL END
#######################################################################################################################

                                     # MODULE: DOCUMENTS & LETTERS START

def Docs1(queue):  # STATUS: TESTED
    # Can you save a document using ESC and view it recent documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document", interval=0.5)
        pag.press("esc", interval=1)  # saving document
        pag.press("enter", interval=1)
        pag.write("1", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '1' in data:
            pag.press("enter", interval=1)
            pag.sleep(3)
            print('Docs1: Can you save a document using ESC and view it recent documents \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I71'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs1: Can you save a document using ESC and view it in recent documents")
            print('Docs1: Can you save a document using ESC and view it in recent documents \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I71'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 1: Failed to open a new document")
        print("Docs 1: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I71'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs2(queue):  # STATUS: TESTED
    # Can you save a document using F2 save and view it in my documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("2", interval=0.5)
        pag.press("enter", interval=1, presses=2)
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        pag.press("2", interval=1)
        data = readFromQueue(queue)
        while '2' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if '2' in data:
            pag.press("enter", interval=1)
            pag.sleep(3)
            print('Docs2: Can you save a document using F2 save and view it in my documents \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I72'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs2: Failed to save and view document from my documents")
            print('Docs2: Can you save a document using F2 save and view it in my documents \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I72'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 2: Failed to open a new document")
        print("Docs 2: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I72'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs3(queue):    # STATUS: TESTED
    # Spell checks a document
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("thi iz a speling chek", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'GMENU_DOCUMENTS_SPELLCHECK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DOCUMENTS_SPELLCHECK' in data:
            pag.press("enter", interval=1, presses=5)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Docs3: Can you spell check a created document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I73'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Docs 3: Failed to sepll check a document")
                print('Docs3: Can you spell check a created document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I73'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Docs 3: Failed to spell check document")
            print("Docs3: Failed to Spell check document \n"
                  ">>> Current String:", data)
            ws['I73'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.sleep(3)
            pag.press("esc", presses=2)
            pag.press("right")
            pag.press("enter")
            returnhome()

    else:
        logging.critical("Docs 3: Failed to open a new document")
        print("Docs3: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I73'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs4(queue):   # STATUS: TESTED
    # Looks up a dictionary word from document
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.press("f2", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DICTIONARYMODULE_MAIN' in data:
            pag.write("document", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' in data:
                pag.sleep(5)
                print('Docs4: Can you look up a word in a created document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I74'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs 4: Failed to look up word from created document")
                print('Docs4: Can you look up a word in a created document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I74'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 4: Failed to open look up word")
            print("Docs4: Failed to open look up word \n"
                  ">>> Current String:", data)
            ws['I74'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 4: Failed to open a new document")
        print("Docs4: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I74'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs5(queue):   # STATUS: TESTED
    # search for a document from my documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.write("2", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(3)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    while '2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '2' in data:
        print('Docs5: Can you search for a specific document from my documents? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I75'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Docs 5: Failed to search for specific document in my documents")
        print('Docs5: Can you search for a specific document from my documents? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I75'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs6(queue):   # STATUS: TESTED
    # Can you rename a document?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("2", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.write("rename", interval=0.5)
    pag.press("enter", interval=2, presses=2)
    pag.press("r", interval=1)
    data = readFromQueue(queue)

    while 'rename2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'rename2' in data:
        print('Docs6: Can you rename a document from my documents? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I76'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Docs 6: Failed to rename a document from my documents")
        print('Docs6: Can you rename a document from my documents? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I76'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs7(queue):   # STATUS: TESTED
    # Can you delete a document?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("r", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Delete selected document' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Delete selected document' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=2, presses=2)
        pag.sleep(5)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'Go to folder - GuideConnect' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Go to folder - GuideConnect' in data:
            print('Docs7: Can you delete a document from my documents? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I77'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs7: Failed to delete a document from my documents")
            print('Docs7: Can you delete a document from my documents? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I77'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs7: Failed to navigate to delete document")
        print('Docs7: Failed to navigate to delete document \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I77'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs8(queue):   # STATUS: TESTED
    # Opens a DOC file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'DOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'DOC' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs8: Can you view a .DOC file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I78'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs8: Failed to view DOC file")
                print('Docs8: Can you view a .DOC file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I78'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs8: Failed to locate DOC file from test data")
            print("Docs 8: failed to locate DOC.doc test file")
            print(data)
            ws['I78'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs8: Failed to find test data folder")
        print("Docs 8: failed to find test data folder")
        print(data)
        ws['I78'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs9(queue):   # STATUS: TESTED
    # Opens a PDF file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'PDF' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'PDF' in data:
            pag.press("enter", interval=1)

            while 'GMENU_DOCUMENTS_PDF_VIEWER' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_PDF_VIEWER' in data:
                pag.sleep(10)
                print('Docs9: Can you view a .PDF file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I79'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs9: Failed to open a PDF file")
                print('Docs9: Can you view a .PDF file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I79'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs9: Failed to locate PDF test file")
            print("Docs 9: failed to locate PDF.pdf test file")
            print(data)
            ws['I79'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs9: Failed to find test data folder")
        print("Docs 9: failed to find test data folder")
        print(data)
        ws['I79'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs10(queue):   # STATUS: TESTED
    # Opens a RTF file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'RTF' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'RTF' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs10: Can you view a .RTF file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I80'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs10: Failed to view a RTF File")
                print('Docs10: Can you view a .RTF file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I80'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs10: Failed to locate RTF test file")
            print("Docs 10: failed to locate RTF.rtf test file")
            print(data)
            ws['I80'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs10: Failed to find test data folder")
        print("Docs 10: failed to find test data folder")
        print(data)
        ws['I80'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs11(queue):  # STATUS: TESTED
    # Opens a DOCX file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("x", interval=1)
        data = readFromQueue(queue)

        while 'XDOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'XDOC' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs11: Can you view a .DOCX file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I81'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs11: Failed to view a DOCX file")
                print('Docs11: Can you view a .DOCX file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I81'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 11: Failed to locate DOCX test file")
            print("Docs 11: failed to locate XDOC.docx test file")
            print(data)
            ws['I81'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 11: Failed to find test data folder")
        print("Docs 11: failed to find test data folder")
        print(data)
        ws['I81'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs12(queue):  # STATUS: TESTED
    # Opens DOC from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)
    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'DOC' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs12: Can you open a DOC file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I82'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs12: Failed to open DOC file from USB")
                    print('Docs12: Can you open a DOC file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I82'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs12: Failed to DOC file from the USB")
                print("Docs12: Failed to find DOC.doc from the USB")
                print(data)
                ws['I82'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs12: Failed to find correct Regression USB")
            print("Docs12: Failed to find correct Regression USB")
            print(data)
            ws['I82'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs12: Failed to enter open my devices")
        print("Docs12: failed to enter open from device")
        print(data)
        ws['I82'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs13(queue):  # STATUS: TESTED
    # Opens PDF from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("p", interval=1)
            data = readFromQueue(queue)

            while 'PDF' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'PDF' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_PDF_VIEWER' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_PDF_VIEWER' in data:
                    pag.sleep(10)
                    print('Docs13: Can you open a PDF file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I83'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs13: Failed to open PDF from USB")
                    print('Docs13: Can you open a PDF file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I83'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 13: Failed to find PDF on USB")
                print("Docs 13: Failed to find PDF.pdf from USB")
                print(data)
                ws['I83'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 13: Failed to find correct Regression USB")
            print("Docs 13: Failed to find correct Regression USB")
            print(data)
            ws['I83'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 13: Failed to enter open from device")
        print("Docs 13: failed to enter open from device")
        print(data)
        ws['I83'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs14(queue):  # STATUS: TESTED
    # Opens RTF from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'RTF' in data:
                pag.press("enter", interval=1)

                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs14: Can you open a RTF file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I84'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs14: Failed to find RTF file from USB")
                    print('Docs14: Can you open a RTF file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I84'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 14: Failed to find RTF from USB")
                print("Docs 14: Failed to find RTF.rtf from USB")
                print(data)
                ws['I84'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 14: Failed to find correct Regression USB")
            print("Docs 14: Failed to find correct Regression USB")
            print(data)
            ws['I84'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 14: Failed to enter open from device")
        print("Docs 14: failed to enter open from device")
        print(data)
        ws['I84'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs15(queue):  # STATUS: TESTED
    # Opens DOCX from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("x", interval=1)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs15: Can you open a DOCX file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I85'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs 15: Failed to open DOCX from USB")
                    print('Docs15: Can you open a DOCX file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I85'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 15: Failed to find DOCX from USB")
                print("Docs15: Failed to find XDOC.docx from USB")
                print(data)
                ws['I85'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 15: Failed to find correct Regression USB")
            print("Docs15: Failed to find correct Regression USB")
            print(data)
            ws['I85'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 15: Failed to enter open from device")
        print("Docs15: failed to enter open from device")
        print(data)
        ws['I85'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter1(queue):  # STATUS: TESTED
    # Can you manually enter your sender address send a letter to a address book contact?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' in data:
        pag.write("joe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bloggs", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("22", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin street", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bromsgrove", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("worcestershire", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("b45 9xe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("united kingdom", interval=0.5)
        pag.press("enter", presses=4, interval=0.5)
        pag.write("this is a test letter", interval=0.5)
        pag.press("esc", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("letter 1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("L", interval=0.5)
        data = readFromQueue(queue)
        if 'letter 1' in data:
            pag.press("enter", interval=1)
            pag.sleep(2)
            print('Letter1: Can you manually enter your sender address and'
                  ' send a letter to a address book contact? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I86'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Letter 1: Failed to add manual sender address and compose a letter")
            print('Letter1: Can you manually enter your sender address and'
                  ' send a letter to a address book contact? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I86'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letter 1: Failed to open add senders address")
        print("Letter 1: failed to open add senders address")
        print(data)
        ws['I86'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter2(queue):  # STATUS: TESTED
    # Can you create a letter using manually enter address
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'Type address manually' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Type address manually' in data:
        pag.press("enter", interval=1)

        while 'GMENU_DOCUMENTS_RECIPIENTS_ADDRESS_FIRST_NAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DOCUMENTS_RECIPIENTS_ADDRESS_FIRST_NAME' in data:
            pag.write("joe", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("bloggs", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("dolphin", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("22", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("dolphin street", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("bromsgrove", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("b45 9xe", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("united kingdom", interval=0.5)
            pag.press("enter", interval=0.5, presses=2)
            pag.write("this is a letter 2", interval=0.5)
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)
            pag.write("manual", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("m", interval=1)
            pag.press("enter", interval=1)
            pag.press("m", interval=1)
            data = readFromQueue(queue)

            while 'manual' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'manual' in data:
                print('Letter2: Can you manually enter a letter address? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I87'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 2: Failed to manually enter a letter address")
                print('Letter2: Can you manually enter a letter address? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I87'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letter 2: Failed to enter type manual address")
            print("Letters 2: failed to enter type manual address")
            print(data)
            ws['I87'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letters 2: Failed to navigate to type manual address")
        print("Letters 2: failed to navigate to type manual address")
        print(data)
        ws['I87'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter3(queue):  # STATUS: TESTED
    # can you change the sender address via f2 menu?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)

    while 'Change address' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Change address' in data:
        pag.press("enter", interval=1, presses=3)
        data = readFromQueue(queue)

        while 'Dolphin Computer Access Ltd.' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Dolphin Computer Access Ltd.' in data:
            pag.press("enter", interval=2, presses=2)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Letter3: Can you change the sender address via f2 menu? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I88'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 3: Failed to change sender address via F2 menu")
                print('Letter3: Can you change the sender address via f2 menu? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I88'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letters 3: Failed to find dolphin in address book")
            print("Letters 3: Failed to find Dolphin in address book")
            print(data)
            ws['I88'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letters 3: Failed to navigate to change address")
        print("Letters 3: Failed to navigate to change address")
        print(data)
        ws['I88'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter4(queue):  # STATUS: TESTED
    # can you change the recipient address via f2 menu?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)

    while 'Change address' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Change address' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1, presses=2)
        data = readFromQueue(queue)

        while 'Dolphin Computer Access Ltd.' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Dolphin Computer Access Ltd.' in data:
            pag.press("enter", interval=2, presses=2)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Letter4: Can you change the recipient address via f2 menu? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I89'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 4: Failed to change recipients address via F2 menu")
                print('Letter4: Can you change the recipient address via f2 menu? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I89'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letter 4: Failed to find dolphin in address book")
            print("Letter4: Failed to find Dolphin in address book")
            print(data)
            ws['I89'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letter 4: Failed to navigate to change address")
        print("Letter4: Failed to navigate to change address")
        print(data)
        ws['I89'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


                                    # MODULE: DOCUMENTS AND LETTERS END
########################################################################################################################

                                    # MODULE: WEBSITES START (NEEDED)

                                    # MODULE: WEBSITES END (NEEDED)

########################################################################################################################

                                    # MODULE: SCANNER AND CAMERA START

def scanner_text1(queue):  # STATUS: TESTED
    # Can you select a scanner to use?
    start = 0
    delay()
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("s", interval=1)
    data = readFromQueue(queue)

    while 'Scanner' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scanner' in data:
        pag.press("enter", interval=1, presses=2)
        pag.sleep(5)
        data = readFromQueue(queue)
        if 'No devices found.  Please make sure that your scanner' \
           ' is connected to the computer and switched on.' in data:
            logging.critical("Scanner 1: Failed to find scanner attached to system")
            print("Scanner 1: No Scanner attached to the system")
            print(data)
            ws['I143'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'No devices found.  Please make sure that your scanner' \
             ' is connected to the computer and switched on.' not in data:
            data = readFromQueue(queue)
            scanner = data
            if scanner == data:
                pag.press("enter", interval=1)
                print(f"Scanner found: {scanner}")
                print("Scanner 1: Can you select your Scanner \n"
                      "Result: PASS \n"
                      "Current String:", data)
                ws['I143'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 1: Failed to find your attached scanner")
                print("Scanner 1: Can you select your Scanner \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I143'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 1: Failed to select a scanner")
            print("Scanner 1: Can you select your Scanner \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I143'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 1: Failed to navigate to scanner settings")
        print("Scanner 1: failed to navigate to Scanner settings")
        print(data)
        ws['I143'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text2(queue):    # STATUS: TESTED
    # can you scan for text (1 page)
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'Read Scanned Pages' in data:
                break

        data = readFromQueue(queue)
        if "Read Scanned Pages" in data:
            pag.sleep(2)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_SCANNER_OCR_PROGRESS' in data:
                data = readFromQueue(queue)
                if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                    break
            pag.sleep(5)

            print('Scanner Text 2: Can you scan 1 page for text? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I144'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

        else:
            logging.critical("Scanner Text 2: Failed to scan a single page for text")
            print('Scanner Text 2: Can you scan 1 page for text? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I144'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Scanner Text 2: Failed to locate scan for text")
        print("Scanner Text 2: Failed to locate Scan for text")
        print(data)
        ws['I144'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text3(queue):   # STATUS: TESTED
    # Can you scan for text (2 pages)
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'Read Scanned Pages' in data:
                pag.press("right", interval=1)
                data = readFromQueue(queue)
                break

            elif 'GMENU_SCANNER_SCAN_TEXT_COMPLETE' in data:
                pag.press("right", interval=1)
                data = readFromQueue(queue)
                break

        if "Scan Another Page" in data:
            pag.press("enter", interval=1)
            while 'Read Scanned Pages' not in data:
                data = readFromQueue(queue)

            if 'Read Scanned Pages' in data:
                pag.sleep(3)
                pag.press("enter", interval=1)

                while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                    pag.sleep(15)
                    print('Scanner Text 3: Can you scan 2 pages for text? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I145'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

                else:
                    logging.critical("Scanner Text 3: Failed to scan Multiple pages for text")
                    print('Scanner Text 3: Can you scan 2 pages for text? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I145'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Scanner Text 3: Failed to locate read scanned pages")
                print("Scanner Text 3: Failed to locate Read Scanned Pages")
                print(data)
                ws['I145'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Scanner Text 3: Failed to locate read scanned pages")
            print("Scanner Text 3: Failed to locate Read Scanned Pages")
            print(data)
            ws['I145'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Scanner Text 3: Failed to locate scan for text")
        print("Scanner Text 3: Failed to locate Scan for text")
        print(data)
        ws['I145'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text4(queue):    # STATUS: TESTED
    # Can you scan a single document after scanning multiple?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Read Scanned Pages' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if "Read Scanned Pages" in data:
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                pag.sleep(10)
                print('Scanner Text 4: Can you scan 1 page for text after scanning multiple? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I146'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Scanner Text 4: Failed to scan a single page after multiple pages")
                print('Scanner Text 4: Can you scan 1 page for text after scanning multiple?  \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I146'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner Text 4: Failed to locate read scanned pages")
            print("Scanner Text 4: failed to locate read scanned pages")
            print(data)
            ws['I146'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner Text 4: Failed to locate scan for text")
        print("Scanner Text 4: Failed to locate scan for text")
        print(data)
        ws['I146'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text5(queue):    # STATUS: TESTED
    # Can you save a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Read Scanned Pages' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if "Read Scanned Pages" in data:
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.write("5scanner", interval=0.5)
                pag.press("enter", interval=1, presses=2)
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.press("esc", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'My scanned documents' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'My scanned documents' in data:
                    pag.press("enter", interval=1)
                    pag.press("5", interval=0.5)
                    data = readFromQueue(queue)

                    while '5scanner' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if '5scanner' in data:
                        print('Scanner 5: Can you save a scanned document? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I147'].value = 'PASS'
                        wb.save('Test Reports/Automation Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Scanner Text 5: Failed to save a scanned document")
                        print('Scanner 5: Can you save a scanned document? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I147'].value = 'FAIL'
                        wb.save('Test Reports/Automation Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Scanner Text 5: Failed to navigate to my scanned documents")
                    print("Scanner 5: Failed to navigate to my scanned documents")
                    print(data)
                    ws['I147'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Scanner Text 5: Failed to open scanned document")
                print("Scanner 5: Failed to open scanned document")
                print(data)
                ws['I147'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Scanner Text 5: Failed to find read scanned pages")
            print("Scanner 5: Failed to find Read scanned pages")
            print(data)
            ws['I147'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner Text 5: Failed to locate scan for text ")
        print("Scanner 5: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I147'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def scan_image1(queue):  # STATUS: TESTED
    # Can you scan for an image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            print('Scan Image 1: Can you scan for an image? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I148'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Scan Image 1: Failed to scan for an image")
            print('Scan Image 1: Can you scan for an image? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I148'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 1: Failed to locate scan for image")
        print("Scan image 1: Failed to locate scan for image")
        print(data)
        ws['I148'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scan_image2(queue):   # STATUS: TESTED
    # can you save a scanned image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("s", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.write("scanned image", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("esc", presses=2)
            pag.sleep(3)
            pag.press("m", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)
            if 'scanned image' in data:
                print('Scan Image 2: Can you save a scanned image? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I149'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scan Image 2: Failed to save scanned image")
                print('Scan Image 2: Can you save a scanned image? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I149'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scan Image 2: Failed to load scanned image")
            print("Scan image 2: Failed to Succesfully load scanned image")
            print(data)
            ws['I149'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 2: Failed to locate Scan for image")
        print("Scan image 2: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I149'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scan_image3(queue):  # STATUS: TESTED
    # can you scan for text inside of a scanned image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                print('Scan Image 3: Can you scan for text in a scanned image? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I150'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Scan Image 3: Failed to scan for text inside a scanned image")
                print('Scan Image 3: Can you scan for text in a scanned image? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I150'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Scan Image 3: Failed to load scanned image")
            print("Scan image 3: Failed to Succesfully load scanned image")
            print(data)
            ws['I150'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 3: Failed to locate scan for image")
        print("Scan image 3: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I150'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner9(queue):  # STATUS: TESTED
    # Can you rename a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)
    if 'My scanned documents' in data:
        pag.press("enter", interval=1)
        pag.press("5", interval=1)
        data = readFromQueue(queue)

        if '5scanner' in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'renamed 5scanner' in data:
                print('Scanner9: Can you rename a saved scanned document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I151'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 9: Failed to rename a scanned document ")
                print('Scanner9: Can you rename a saved scanned document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I151'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 9: Failed to find a saved scanned document")
            print("Scanner9: Failed to find a saved scanned document")
            print(data)
            ws['I151'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 9: Failed to locate my scanned documents")
        print("Scanner 9: failed to locate my scanned documents")
        print(data)
        ws['I151'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner10(queue):  # STATUS: TESTED
    # Can you delete a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)
    if 'My scanned documents' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'renamed 5scanner' in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            if 'Go to folder - GuideConnect' in data:
                print('Scanner 10: Can you delete a saved scanned document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I152'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 10: Failed to delete a saved scanned document")
                print('Scanner 10: Can you delete a saved scanned document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I152'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 10: Failed to find a saved scanned document")
            print("Scanner 10: Failed to find a saved scanned document")
            print(data)
            ws['I152'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 10: Failed to locate my scanned documents")
        print("Scanner 10: failed to locate my scanned documents")
        print(data)
        ws['I152'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera1(queue):  # STATUS: TESTED
    # Can you capture an image, save and then view it?
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(5)
    pag.press("c", interval=1)
    data = readFromQueue(queue)
    if 'Camera' in data:
        pag.press("enter", interval=1)

        while 'GMENU_SHOW_CAMERA_FEED' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_SHOW_CAMERA_FEED' in data:
            pag.press("enter", interval=1)
            start = 0

            while 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' in data:
                pag.sleep(5)
                pag.press("f2", interval=1)
                pag.press("s", presses=2, interval=1)  # save image
                pag.press("enter", interval=1)
                pag.write("camera test", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                pag.press("esc", presses=2, interval=1)
                pag.press("m", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'My pictures' in data:
                    pag.press("enter", interval=1)
                    pag.press("c", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                        print('Camera 1: Can you capture an image, save and view it? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I153'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Camera 1: Failed to capture, save and view an image")
                        print('Camera 1: Can you capture an image, save and view it? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I153'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Camera 1: Failed to locate my pictures")
                    print("Camera 1: Failed to locate My pictures")
                    print(data)
                    ws['I153'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Camera 1: Failed to load captured image ")
                print("Camera 1: Failed to load GMENU_SCANNER_SCAN_IMAGE_COMPLETE")
                print("Current String:", data)
                ws['I153'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Camera 1: Failed to load camera feed")
            print("Camera 1: Camera failed to load GMENU_SHOW_CAMERA_FEED")
            print("Current String:", data)
            ws['I153'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Camera 1: Failed to locate camera")
        print("Camera 1: failed to locate camera")
        print(data)
        ws['I153'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera2(queue):  # STATUS: TESTED
    # Can you rename a saved image?
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("m", interval=1, presses=2)
    data = readFromQueue(queue)
    if 'My pictures' in data:
        pag.press("enter", interval=1)
        pag.press("c", interval=1)
        data = readFromQueue(queue)
        if 'camera test' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'renamed camera test' in data:
                print('Camera 2: Can you rename a saved camera picture? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I154'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Camera 2: Failed to rename a saved captured image")
                print('Camera 2: Can you rename a saved camera picture? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I154'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.exception("Camera 2: Failed to locate saved picture")
            print("Camera 2: failed to locate saved picture")
            print("Current String:", data)
            ws['I154'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Camera 2: Failed to locate my pictures")
        print("Camera 2: failed to locate my pictures")
        print(data)
        ws['I154'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera3(queue):     # STATUS: TESTED
    # Deletes a saved captured image
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("m", interval=1, presses=2)
    data = readFromQueue(queue)
    if 'My pictures' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'renamed camera test' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'Go to folder - GuideConnect' in data:
                print('Camera 3: Can you delete a saved camera picture? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I155'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Camera 3: Failed to delete a saved camera picture")
                print('Camera 3: Can you delete a saved camera picture? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I155'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Camera 3: Failed to locate saved picture")
            print("Camera 3: failed to locate saved picture")
            print("Current String:", data)
            ws['I155'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Camera 3: Failed to locate My Pictures folder")
        print("Camera 3: failed to locate my pictures")
        print(data)
        ws['I155'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

                                      # MODULE: SCANNER AND CAMERA END
#######################################################################################################################

                                      # MODULE: BOOKSHELF START


def books1(queue):  # STATUS: TESTED
    # check all UK providers are present
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Find a new book' in data:
        pag.press("enter", interval=1)

        books_present = []

        while 'Calibre Audio' not in data:
            data = readFromQueue(queue)
            books_present.append(data)
            pag.sleep(1)
            pag.press("right", interval=1)
            if 'Calibre Audio' in data:
                break

        if books_present == books_expected:
            print('Books 1: Compares active vs expected book providers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I127'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 1: Providers list was different to expected")
            print('Books 1: Compares active vs expected book providers \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I127'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 1: Failed to locate find a new book")
        print("Books 1: failed to locate find a new book")
        print(data)
        ws['I127'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books2(queue):  # STATUS: TESTED
    # downloads book from Project Gutenburg
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'Project Gutenberg' in data:
        pag.press("enter", presses=3, interval=1)
        pag.sleep(5)
        data = readFromQueue(queue)
        global book
        book = data
        logging.info(f"Books 2: Book downloaded: {book}")
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Queued' in data:
            data = readFromQueue(queue)
            pag.press("right", interval=1)
            pag.press("left", interval=1)
            if 'Read' in data:
                pag.press("enter", interval=1)
                break

        while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
            print('Books 2: download a book from Project Gutenberg \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I128'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 2: Failed to download a book from Project Gutenberg")
            print('Books 2: download a book from Project Gutenberg \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I128'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 2: Failed to locate Project Gutenberg")
        print("failed to find project gutenberg")
        print(data)
        ws['I128'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books3(queue):  # STATUS: TESTED
    # Continue reading a book
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
        print('Books 3: Continue listening to a book \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I129'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Books 3: Failed to continue reading a book")
        print('Books 3: Continue listening to a book \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I129'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books4(queue):  # STATUS: TESTED
    # Is book present in your 'My Books'
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)
    if data == book:
        print('Books 4: is previously read book present in My Books \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I130'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Book 4: Failed to locate book in My Books")
        print('Books 4: is previously read book present in My Books \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I130'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books5(queue):  # STATUS: TESTED
    # View book information from My Books
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.sleep(5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKINFO' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKINFO' in data:
        print('Books 5: View details of a book via My Books \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I131'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Books 5: Failed to view details of a book via My Books")
        print('Books 5: View details of a book via My Books \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I131'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books6(queue):  # STATUS: TESTED
    # Copies a book from My Books to USB device
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)
    if 'Copy To device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if f'{found[0]}' in data:
            pag.press("enter", presses=2, interval=1)
            print('Books 6: Copy a book from My Books to USB device \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I132'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Book 6: Failed to locate and transfer a book to a USB device")
            print('Books 6: Copy a book from My Books to USB device \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I132'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 6: Failed to locate copy to device")
        print("Books 6: failed to locate copy to device")
        print(data)
        ws['I132'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books7(queue):  # STATUS: TESTED
    # Delete a book via my books
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Delete Book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Delete Book' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'No Books Found' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'No Books Found' in data:
            print('Books 7: Delete a book via My Books \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I133'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 7: Failed to delete a book via My Books")
            print('Books 7: Delete a book via My Books \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I133'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 7: Failed to locate Delete book")
        print("Books 7: Failed to locate delete book")
        print(data)
        # ws['I133'].value = 'FAIL'
        # wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books8(queue):  # STATUS: TESTED
    # View details of a library
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Find a new book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Find a new book' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' in data:
            print('Books 8: View details of a provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I134'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 8: Failed to view details of a content provider")
            print('Books 8: View details of a provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I134'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 8: Failed to locate find a new book")
        print("Books 8: failed to locate find a new book")
        print(data)
        ws['I134'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books9(queue):  # STATUS: NOT TESTED
    # View details of a book via find a new book
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Find a new book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Find a new book' in data:
        pag.press("enter", presses=5, interval=1)
        pag.press("b", interval=1)
        data = readFromQueue(queue)
        if 'Book information' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_BOOKSHELFMENU_BOOKINFO' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_BOOKSHELFMENU_BOOKINFO' in data:
                print('Books 9: View details of a book via find a new book \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I135'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 9: Failed to view details of a book via find a new book")
                print('Books 9: View details of a book via find a new book \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I135'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 9: Failed to locate book information")
            print("Books 9: failed to locate book information")
            print(data)
            ws['I135'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 8: Failed to locate find a new book")
        print("Books 8: failed to locate find a new book")
        print(data)
        ws['I135'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books10(queue):  # STATUS: TESTED
    # Open and views a book from USB
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("r", interval=1)
    data = readFromQueue(queue)
    if 'Read from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f'{found[0]}' in data:
            pag.press("enter", presses=2, interval=1)

            while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
                print('Books 10: Read a book from device \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I136'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 10: Failed to read a book from USB device")
                print('Books 10: Read a book from device \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I136'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 10: Failed to locate correct USB")
            print("Books 10: failed to locate correct regression usb")
            print(data)
            ws['I136'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 10: Failed to locate read from device")
        print("Books 10: failed to locate read from device")
        print(data)
        ws['I136'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers1(queue):  # STATUS: TESTED
    # Log in to provider and download a newspaper
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'RNIB NTNM' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'RNIB NTNM' in data:
            pag.press("enter", interval=1)
            pag.write("i.rollason", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("tna751", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'Magazines by category' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Magazines by category' in data:
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                newspaper = data
                logging.info(f"Newspaper downloaded: {newspaper}")
                pag.press("enter", presses=3, interval=1)
                pag.sleep(15)
                pag.press("enter", interval=1)
                pag.sleep(10)
                data = readFromQueue(queue)

                if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                elif '' in data:
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("News 1: Failed to log in, download and play an edition")
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("News 1: Failed to log into RNIB NTNM")
                print("News 1: failed to log into RNIB NTNM")
                print(data)
                ws['I137'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("News 1: Failed to locate RNIB NTNM")
            print("News 1: failed to locate RNIB NTNM")
            print(data)
            ws['I137'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 1: Failed to locate newspapers and magazines")
        print("News 1: failed to locate newspapers and magazines")
        print(data)
        ws['I137'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers2(queue):  # STATUS: TESTED
    # continue playing a newspaper
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.sleep(7)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
        print('News 2: Continue listening to an edition \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I138'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    elif '' in data:
        print('News 2: Continue listening to an edition \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I138'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("News 2: Failed to continue playing an edition")
        print('News 2: Continue listening to an edition \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I138'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers3(queue):  # STATUS: TESTED
    # Play an edition from My newspapers
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", presses=4, interval=1)
        pag.sleep(5)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        elif '' in data:
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 3: Failed to play edition from my newspapers")
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 3: Failed to locate newspapers and magazines")
        print("News 3: failed to locate newspapers and magazines")
        print(data)
        ws['I139'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers4(queue):  # STATUS: TESTED
    # unsubscribe from edition from my newspapers
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)

        while '> Unsubscribe' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if '> Unsubscribe' in data:
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'No Newspapers Or Magazines Found' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No Newspapers Or Magazines Found' in data:
                print('News 4: Unsubscribe from an edition from My Newspapers \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I140'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("News 4: Failed to unsubscribe from edition via my newspapers")
                print('News 4: Unsubscribe from an edition from My Newspapers \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I140'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("News 4: Failed to unsubscribe")
            print("News 4: failed to unsubscribe")
            print(data)
            ws['I140'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 4: Failed to locate newspapers and magazines")
        print("News 4: failed to locate newspapers and magazines")
        print(data)
        ws['I140'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers5(queue):  # STATUS: TESTED
    # View details of newspaper provider
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' in data:
            print('News 5: View details of a newspaper content provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I141'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 5: Failed to view details of a newspaper library provider")
            print('News 5: View details of a newspaper content provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I141'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 5: Failed to locate newspapers and magazines")
        print("News 5: failed to locate newspapers and magazines")
        print(data)
        ws['I141'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers6(queue):  # STATUS: TESTED
    # log out of newspaper provider
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_USERNAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_USERNAME' in data:
            print('News 6: Log out of newspaper provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I142'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 6: Failed to log out of newspaper provider")
            print('News 6: Log out of newspaper provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I142'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 6: Failed to locate newspapers and magazines")
        print("News 6: failed to locate newspapers and magazines")
        print(data)
        ws['I142'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books11(queue):
    # Copies a book from My Books to USB device
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("r", interval=1)
    data = readFromQueue(queue)
    if 'Read from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f'{found[0]}' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)

            while 'XDOC.docx' not in data:
                data = readFromQueue(queue)
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'XDOC.docx' in data:
                print('Books 11: Delete a book from device \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I191'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 11: Failed to delete a book from USB device")
                print('Books 11: Delete a book from device \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I191'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 11: Failed to locate regression USB drive")
            print("Books 11: failed to locate Regression USB drive")
            print(data)
            ws['I191'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 11: Failed to locate read from device")
        print("Books 11: failed to locate Read from device")
        print(data)
        ws['I191'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


                                      # MODULE: BOOKSHELF END
#######################################################################################################################

                                     # MODULE: ADDRESS BOOK AND CALENDAR START

def calendar1(queue):  # STATUS: TESTED
    # Are you able to add a new event to your calendar?
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)  # address book
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("enter", presses=3, interval=1)
    data = readFromQueue(queue)
    pag.sleep(2)

    while "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" in data:
        # TEST CASE 1: - TODAY'S EVENTS
        pag.write("tc1", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)  # DAILY REPEAT
        pag.press("enter", presses=4, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(2)
        data = readFromQueue(queue)
        pag.sleep(5)

        if "Event name" in data:
            print('Calendar 1 - TC1: Calendar - Todays Events \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I58'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')

            # TEST CASE 2: Upcoming Events
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            pag.sleep(1)
            if "Upcoming events" in data:
                pag.press("enter")
                data = readFromQueue(queue)
                pag.sleep(5)
                if 'tc1' in data:
                    pag.sleep(2)
                    print('Calendar 1 - TC2: Calendar - Upcoming Events \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I58'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')

                    # TEST CASE 3: Calendar - View Calendar
                    pag.press("esc", interval=1)
                    pag.press("right", presses=2, interval=1)
                    pag.press("enter", presses=3, interval=1)
                    data = readFromQueue(queue)
                    pag.sleep(5)
                    if 'Event name' in data:
                        pag.sleep(2)
                        print('Calendar 1 - TC3: Calendar - View Calendar \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I58'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Calendar 1 (TC3) - Failed to find calendar event")
                        print('Calendar 1 - TC3: Failed to find calendar event \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I58'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Calendar 1 (TC2) - Failed to find upcoming event")
                    print('Calendar 1 - TC2: Failed to find the upcoming event \n'
                          '>>> Current String:', data)
                    ws['I58'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Calendar 1 (TC2) - Failed to enter upcoming events")
                print('Calendar 1 - TC2: Failed to enter upcoming events \n'
                      '>>> Current String:', data)
                ws['I58'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 1 (TC1) - Failed to locate event name")
            print('Calendar 1 - TC1: Failed to locate event name \n'
                  '>>> Current String:', data)
            ws['I58'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 1: Failed to add a calendar event")
        print('Calendar 1: Failed to add a calendar event \n'
              '>>> Current String:', data)
        ws['I58'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar2(queue):   # STATUS: TESTED
    # edit details of an event
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", interval=1, presses=4)
        data = readFromQueue(queue)

        if 'GMENU_CALENDAR_EDIT_EVENT_UPDATE_DETAILS' in data:
            pag.write("update", interval=0.5)
            pag.press("enter", interval=1)

            print('Calendar 2: View and edit a calendar entry \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I59'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Calendar 2: Failed to view and edit calendar entry")
            print('Calendar 2: View and edit a calendar entry \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I59'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 2: Failed to navigate to calendar")
        print("Calendar 2: Failed to navigate to Calendar")
        print(data)
        ws['I59'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar3(queue):   # STATUS: TESTED
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)
        if 'Send details as attachment' in data:
            pag.press("enter", presses=3, interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("sent from calendar", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                returnhome()
                delay()
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)
                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter", interval=1)
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("f2", interval=1)
                    pag.press("v", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'calendar.ics' in data:
                        print('Calendar 3: Send event via email \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I60'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Calendar 3: Failed to send event via email")
                        print('Calendar 3: Send event via email \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I60'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Calendar 3: Failed to find Email acccount")
                    print("Calendar 3: failed to find 'guideautomation2@outlook.com'")
                    print(data)
                    ws['I60'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Calendar 3: Failed to send event as email")
                print("Calendar 3: Failed to send event as email")
                print(data)
                ws['I60'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 3: Failed to locate send event as attachment")
            print("Calendar 3: Failed to locate send event as attachment")
            print(data)
            ws['I60'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 3: Failed to navigate to Calendar")
        print("Calendar 3: Failed to navigate to Calendar")
        print(data)
        ws['I60'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar4(queue):   # STATUS: TESTED
    # Can you delete a calendar event?
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)
        if 'Event name' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'No events on this day' in data:
                print('Calendar 4: Can you delete a calendar entry? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I61'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Calendar 4: Failed to delete calendar entry")
                print('Calendar 4: Can you delete a calendar entry? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I61'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 4: Failed to find pre existing event")
            print("Calendar 4: Failed to find pre existing event")
            print(data)
            ws['I61'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 4: Failed to navigate to calendar")
        print("Calendar 4: Failed to navigate to Calendar")
        print(data)
        ws['I61'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address1(queue):    # STATUS: TESTED
    # Can you add a contact to your address book?
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'GMENU_ADDRESSBOOK_ADD' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_ADDRESSBOOK_ADD' in data:
        pag.write("Joe", interval=0.5)  # first name
        pag.press("enter", interval=1)
        pag.write("bloggs", interval=0.5)  # surname
        pag.press("enter", interval=1)
        pag.write("0123456789101", interval=0.5)  # telephone number
        pag.press("enter", interval=1)
        pag.write("22", interval=0.5)  # house number / number
        pag.press("enter", interval=1)
        pag.write("Dolphin street", interval=0.5)  # street name
        pag.press("enter", interval=1)
        pag.write("dolphin town", interval=0.5)  # town name
        pag.press("enter", interval=1)
        pag.write("dolphin county", interval=0.5)  # county name
        pag.press("enter", interval=1)
        pag.write("B45 9XE", interval=0.5)  # postcode
        pag.press("enter", interval=1)
        pag.write("United kingdom", interval=0.5)  # country
        pag.press("enter", interval=1)
        pag.write("joebloggs@dolphin.co.uk", interval=0.5)  # email
        pag.press("enter", interval=1)
        pag.write("Dolphin computer access", interval=0.5)  # company name
        pag.press("enter", interval=1)
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)
        if "Joe bloggs" in data:
            print('Address 1: Can you add a new contact? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I49'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 1: failed to add new contact")
            print('Address 1: Can you add a new contact? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I49'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 1: Failed to enter address book")
        print("Address 1: Failed to enter address book")
        print(data)
        ws['I49'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address2(queue):    # STATUS: TESTED
    # Edit an address book contact
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("j", interval=1)
    data = readFromQueue(queue)

    while 'Joe bloggs' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Joe bloggs' in data:
        pag.press("enter", presses=2, interval=1)
        pag.write("edited", interval=0.5)  # first name
        pag.press("enter", interval=1)
        pag.press("down", interval=1)
        pag.press("enter", interval=1)
        pag.write("user", interval=0.5)  # surname
        pag.press("enter", interval=1)
        pag.press("down", presses=2, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit phone number", interval=0.5)  # telephone number
        pag.press("enter", interval=1)
        pag.press("down", presses=3, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit house number", interval=0.5)  # house name / number
        pag.press("enter", interval=1)
        pag.press("down", presses=4, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit street name", interval=0.5)  # street name
        pag.press("enter", interval=1)
        pag.press("down", presses=5, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit town", interval=0.5)  # town
        pag.press("enter", interval=1)
        pag.press("down", presses=6, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit county", interval=0.5)  # county
        pag.press("enter", interval=1)
        pag.press("down", presses=7, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit postcode", interval=0.5)  # postcode
        pag.press("enter", interval=1)
        pag.press("down", presses=8, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit country", interval=0.5)  # country
        pag.press("enter", interval=1)
        pag.press("down", presses=9, interval=1)
        pag.press("enter", interval=1)
        pag.write("guideautomation2@outlook.com", interval=0.5)  # email
        pag.press("enter", interval=1)
        pag.press("down", presses=10, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit company", interval=0.5)  # company
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("e", interval=1)

        data = readFromQueue(queue)
        if "edited user" in data:
            print('Address 2: Edit an existing contact \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I50'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 2: Failed to edit existing contact")
            print('Address 2: Edit an existing contact \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I50'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 2: Failed to locate user 'Joe Bloggs")
        print("Address 2: Failed to locate user 'Joe Bloggs'")
        print(data)
        ws['I50'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address3(queue):    # STATUS: TESTED
    # Send an email to address book contact
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'GMENU_EMAIL_NEW_SUBJECT' in data:
            pag.write("Sent from Address book", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("this is a test email", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                pag.press("enter", interval=1)
                print('Address 3: Send an email to address contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I51'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 3: Failed to send an email to address contact")
                print('Address 3: Send an email to address contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I51'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 3: Failed to enter send email")
            print("Address 3: failed to enter send email")
            print(data)
            ws['I51'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 3: Failed to locate user 'Edited user'")
        print("Address 3: Failed to locate user 'Edited user'")
        print(data)
        ws['I51'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address4(queue):    # STATUS: TESTED
    # Invite to video calling
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'GMENU_EMAIL_INVITE_MESSAGE' in data:
            pag.press("enter", presses=4, interval=1)
            data = readFromQueue(queue)
            if 'Dolphin Computer Access Ltd.' in data:
                print('Address 4: Send an email to address contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I52'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 4: Failed to send email to address contact")
                print('Address 4: Send an email to address contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I52'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 4: Failed to enter invite to video calling")
            print("Address 4: failed to enter invite to video calling")
            print(data)
            ws['I52'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 4: Failed to locate 'Edited user'")
        print("Address 4: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I52'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address5(queue):    # STATUS: TESTED
    # Compose a letter from address book
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_DOCUMENTS_VIEW' in data:
            print('Address 5: Compose a letter \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I53'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 5: Failed to compose a letter")
            print('Address 5: Compose a letter \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I53'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 5: Failed to locate user 'Edited user'")
        print("Address 5: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I53'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address6(queue):    # STATUS: TESTED
    # Sends a VCF as attachment, deletes existing entry and then imports it to address book
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=3, interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("sent from address", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                pag.press("enter", interval=1)
                pag.press("e", interval=1)
                data = readFromQueue(queue)

                if 'edited user' in data:
                    pag.press("f2", interval=1)
                    pag.press("d", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    returnhome()
                    delay()
                    pag.press("enter", interval=1)
                    pag.press("g", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation2@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation2@outlook.com' in data:
                        pag.press("enter", interval=1)
                        pag.press("i", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("f2", interval=1)
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'contacts.vcf' in data:
                            pag.press("enter", interval=1)
                            pag.press("f2", interval=1)
                            pag.press("enter", presses=3, interval=1)
                            returnhome()
                            pag.press("a", interval=1)
                            pag.press("enter", presses=2, interval=1)
                            pag.press("e", interval=1)
                            data = readFromQueue(queue)
                            if 'edited user' in data:
                                print('Address 6: Send VCF via email, delete existing instance and then import \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I54'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Address 6: Failed to send VCF via email")
                                print('Address 6: Send VCF via email, delete existing instance and then import \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I54'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Address 6: Contact VCF not attached to email")
                            print("Address 6: contact VCF not attached to email")
                            print(data)
                            ws['I54'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Address 6: Failed to find email address")
                        print("Address 6: Failed to find guideautomation2@outlook.com")
                        print(data)
                        ws['I54'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Address 6: Failed to delete edited user")
                    print("Address 6: Failed to delete edited user")
                    print(data)
                    ws['I54'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Address 6: Failed to send a new email")
                print("Address 6: Failed to send a new email")
                print(data)
                ws['I54'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Address 6: Failed to enter new email")
            print("Address 6: Failed to enter new email")
            print(data)
            ws['I54'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 6: Failed to locate user 'Edited user'")
        print("Address 6: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I54'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address7(queue):    # STATUS: TESTED
    # Edit an imported contact details
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)
    if 'edited user' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ADDRESSBOOK_EDIT' in data:
            pag.write("imported", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            data = readFromQueue(queue)
            if 'imported user' in data:
                print('Address 7: Edit details of imported user \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I55'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 7: Failed to edit details of imported user")
                print('Address 7: Edit details of imported user \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I55'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 7: Failed edit imported contact")
            print("Address 7: Failed to edit imported contact")
            print(data)
            ws['I55'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 7: Failed to find imported contact")
        print("Address 7: failed to find imported contact")
        print(data)
        ws['I55'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address8(queue):    # STATUS: TESTED
    # sort contacts by last name
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    data = readFromQueue(queue)
    if 'imported user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=4, interval=1)
        pag.press("enter", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", interval=1)
        pag.press("u", interval=1)
        data = readFromQueue(queue)
        if 'user, imported' in data:
            print('Address 8: sort contacts by Last name \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I56'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 8: Failed to sort contacts by last name")
            print('Address 8: sort contacts by Last name \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I56'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 8: Failed to find imported user")
        print("Address 8: failed to find imported user")
        print(data)
        ws['I56'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address9(queue):    # STATUS: TESTED
    # Delete multiple address book contacts
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press(" ")
    pag.press("right", interval=1)
    pag.press(" ")
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)
    if 'Delete' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'No contacts' in data:
            print('Address 9: Delete multiple contacts \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I57'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 9: Failed to delete multiple contacts")
            print('Address 9: Delete multiple contacts \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I57'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 9: Failed to locate delete")
        print("Address 9: failed to locate delete")
        print(data)
        ws['I57'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


                                        # MODULE: ADDRESS BOOK AND CALENDAR END
#######################################################################################################################

                                        # MODULE: ENTERTAINMENT START


def radio1(queue):   # STATUS: TESTED
    # Play a radio station from favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=4, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
        pag.press("enter", interval=1)
        print('Radio 1: Play a station from favourites \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I115'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        pag.press("enter", interval=1)
        logging.critical("Radio 1: Failed to play a station from favourites")
        print('Radio 1: Play a station from favourites \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I115'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio2(queue):  # STATUS: TESTED
    # Continue listening to a previous station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
        pag.press("enter", interval=1)
        print('Radio 2: Continue Playing a station \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I116'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        pag.press("enter", interval=1)
        logging.critical("Radio 2: Failed to continue playing a radio station")
        print('Radio 2: Continue Playing a station \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I116'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio3(queue):  # STATUS: TESTED
    # Add a station to favourites via favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.press("a", interval=1)
        data = readFromQueue(queue)
        if "Absolute Radio 80's" in data:
            print('Radio 3: Add a radio station to favourites via favourites \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I117'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            print('Radio 3: Add a radio station to favourites via favourites \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I117'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 3: Failed to add a radio station to favourites via favourites")
        print("Radio 3: failed to navigate to favourites")
        print(data)
        ws['I117'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio4(queue):  # STATUS: TESTED
    # Add a custom radio station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("a", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_RADIO_CUSTOM' in data:
            pag.write("test", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("http://stream.live.vc.bbcmedia.co.uk/bbc_radio_cornwall", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)
            if 'test' in data:
                print('Radio 4: Add a custom radio station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I118'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 4: Failed to add a custom radio station")
                print('Radio 4: Add a custom radio station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I118'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 4: Failed to enter add custom station")
            print("Radio 4: Failed to enter add custom station")
            print(data)
            ws['I118'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 4: Failed to navigate to favourites")
        print("Radio 4: Failed to enter favourites")
        print(data)
        ws['I118'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio5(queue):  # STATUS: TESTED
    # Remove a radio station from favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'RNIB Connect radio' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'Remove selected from favourites' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'BBC Radio 4' in data:
                    print('Radio 5: Remove a station from favourites \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I119'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Radio 5: Failed to remove a radio station from favourites")
                    print('Radio 5: Remove a station from favourites \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I119'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Radio 5: Failed to locate remove from favourites")
                print("Radio 5: Failed to locate remove from favourites")
                print(data)
                ws['I119'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 5: Failed to locate RNIB Radio ")
            print("Radio 5: Failed to locate RNIB Radio")
            print(data)
            ws['I119'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 5: Failed to locate favourites")
        print("Radio 5: Failed to locate favourites")
        print(data)
        ws['I119'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio6(queue):  # STATUS: TESTED
    # Views details of a radio favourite
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Name' in data:
            print('Radio 6: View details of a radio favourite \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I120'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Radio 6: Failed to view details of a radio favourite")
            print('Radio 6: View details of a radio favourite \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I120'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 6: Failed to locate favourites")
        print("Radio 6: Failed to locate favourites")
        print(data)
        ws['I120'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio7(queue):  # STATUS: TESTED
    # Remove a favourite podcast from play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("right", interval=1)
        pag.press("left", interval=1)
        data = readFromQueue(queue)
        if "Remove From favourites" in data:
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("esc", interval=1)
            pag.press("f", interval=1)
            pag.press("enter", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)
            if 'BBC Radio 4' in data:
                print('Radio 7: Removes a radio favourite via Play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I121'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 7: Failed to remove a radio favourite via play a new station")
                print('Radio 7: Removes a radio favourite via Play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I121'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 7: Failed to locate remove from favourites via play a new station")
            print("Radio 7: Failed to locate remove from favourites")
            print(data)
            ws['I121'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 7: Failed to locate play a new station")
        print("Radio 7: Failed to locate play a new station")
        print(data)
        ws['I121'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio8(queue):  # STATUS: TESTED
    # Adds a radio to favourites via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("c", interval=1)
        data = readFromQueue(queue)
        if 'Classic Fm' in data:
            pag.press("f2", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("esc", interval=1)
            pag.press("f", interval=1)
            pag.press("enter", interval=1)
            pag.press("c", interval=1)
            data = readFromQueue(queue)
            if 'Classic Fm' in data:
                print('Radio 8: Adds a station to favourites via play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I122'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 8: Failed to addd a station to favourites via play a new station")
                print('Radio 8: Adds a station to favourites via play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I122'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 8: Failed to locate Classic FM")
            print("Radio 8: Failed to locate Classic FM")
            print(data)
            ws['I122'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 8: Failed to locate play a new station")
        print("Radio 8: Failed to locate play a new station")
        print(data)
        ws['I122'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio9(queue):  # STATUS: TESTED
    # Searches for a specific station via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_RADIO_STATIONS_SEARCH_EDIT' in data:
            pag.write("talk", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Talksport' in data:
                print('Radio 9: Searches for a specific station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I123'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 9: Failed to search for a specific station")
                print('Radio 9: Searches for a specific station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I123'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 9: Failed to enter search radio list")
            print("Radio 9: Failed to enter search radio list")
            print(data)
            ws['I123'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 9: Failed to locate play a new station")
        print("Radio 9: Failed to locate play a new station")
        print(data)
        ws['I123'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio10(queue):  # STATUS: TESTED
    # Changes the Language list to Czech
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'Switch language' in data:
            pag.press("enter", interval=1)
            pag.press("c", presses=2, interval=1)
            data = readFromQueue(queue)
            if 'Czech' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'ABradio.cz\xa0Radio\xa0Folk' in data:
                    print('Radio 10: Changes radio language list to Czech \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I124'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Radio 10: Failed to change language list to Czech")
                    print('Radio 10: Changes radio language list to Czech \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I124'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Radio 10: Failed to locate Czech in language list")
                print("Radio 10: Failed to locate Czech in language list")
                print(data)
                ws['I124'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 10: Failed to locate switch language")
            print("Radio 10: Failed to locate switch language")
            print(data)
            ws['I124'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 10: Failed to locate play a new station")
        print("Radio 10: Failed to locate play a new station")
        print(data)
        ws['I124'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio11(queue):  # STATUS: TESTED
    # Views details of station via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)
        if 'Details' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Name' in data:
                print('Radio 11: Views details of station via play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I125'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 11: Failed to view details of station via play a new station")
                print('Radio 11: Views details of station via play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I125'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 11: Failed to locate details via play a new station")
            print("Radio 11: Failed to locate details")
            print(data)
            ws['I125'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 11: Failed to locate play a new station")
        print("Radio 11: Failed to locate play a new station")
        print(data)
        ws['I125'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio12(queue):  # STATUS: TESTED
    # iterates through active radio list and compares vs expected
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)

        radio_present = []

        while 'UCB UK' not in data:
            data = readFromQueue(queue)
            radio_present.append(data)
            pag.sleep(1)
            pag.press("right", interval=1)
            if 'UCB UK' in data:
                break

        if radio_present == radio_expected:
            print('Radio 12: Compares active vs expected radio stations \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I126'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Radio 12: Failed to match active vs expected radio list")
            print('Radio 12: Compares active vs expected radio stations \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I126'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 12: Failed to locate play a new station")
        print("Radio 12: Failed to locate play a new station")
        print(data)
        ws['I126'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def Podcast1(queue):    # STATUS: TESTED
    # Play a new podcast from find a new podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p")
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.sleep(10)
            pag.press("enter")
            print('Podcast 1: Can you play a new podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I10'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Podcast 1: Failed to play a new podcast")
            print('Podcast 1: Can you play a new podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I10'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 1: Failed to open podcasts")
        print("Podcast 1: Failed to open Podcasts")
        print(data)
        ws['I10'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast2(queue):   # STATUS: TESTED
    # Continue playing a recently played podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.sleep(10)
            pag.press("enter")
            print('Podcast 2: Can you continue playing a podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I11'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            pag.sleep(10)
            pag.press("enter")
            logging.critical("Podcast 2: Failed to continue listening to a podcast")
            print('Podcast 2: Can you continue playing a podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I11'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 2: Failed to open podcasts")
        print("Podcast 2: Failed to open Podcasts")
        print(data)
        ws['I11'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast3(queue):  # STATUS: TESTED
    # can you add a podcast to your favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2")
            pag.sleep(3)
            pag.press("enter", interval=1)
            pag.press("w", interval=1)
            pag.sleep(3)
            data = readFromQueue(queue)

            while "Woman's Hour: News, Politics, Culture" not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if "Woman's Hour: News, Politics, Culture" in data:
                pag.press("enter", interval=1, presses=2)
                pag.press("w")
                pag.sleep(3)
                data = readFromQueue(queue)

                while "Woman's Hour: News, Politics, Culture" not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if "Woman's Hour: News, Politics, Culture" in data:
                    print('Podcast 3: Can you add a podcast to favourites? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I12'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 3: Failed to add a podcast favourite")
                    print('Podcast 3: Can you add a podcast to favourites? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I12'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 3: Failed to add Women's Hour to favourites")
                print("Podcast 3: Failed to add Women's Hour to favourites")
                print(data)
                ws['I12'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 3: Failed to open podcast favourites")
            print("Podcast 3: Failed to open podcast favourites")
            print(data)
            ws['I12'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 3: Failed to open podcasts")
        print("Podcast 3: Failed to open Podcasts")
        print(data)
        ws['I12'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast4(queue):  # STATUS: TESTED
    # can you remove a podcast from favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("w", interval=1)
            pag.sleep(3)
            data = readFromQueue(queue)

            while "Woman's Hour: News, Politics, Culture" not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if "Woman's Hour: News, Politics, Culture" in data:
                pag.press("f2", interval=1)
                pag.press("r")
                data = readFromQueue(queue)

                while 'Remove selected from favourites' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Remove selected from favourites' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.sleep(5)
                    pag.press("w")
                    data = readFromQueue(queue)

                    while 'In Touch' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'In Touch' in data:
                        print('Podcast 4: Can you add remove a podcast from favourites? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I13'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Podcast 4: Failed to remove a podcast")
                        print('Podcast 4: Can you add remove a podcast from favourites? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I13'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 4: Failed to find remove inside favourites")
                    print("Podcast 4: Failed to find remove from favourites")
                    print(data)
                    ws['I13'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 4: Failed to find Women's Hour in favourites")
                print("Podcast 4: Failed to find Women's Hour in favourites")
                print(data)
                ws['I13'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 4: Failed to open podcast favourites")
            print("Podcast 4: Failed to open podcast favourites")
            print(data)
            ws['I13'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 4: Failed to open podcasts")
        print("Podcast 4: Failed to open Podcasts")
        print(data)
        ws['I13'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast5(queue):  # STATUS: TESTED
    # can you view details of a favourite podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 5: Can you view details of a favourite podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I14'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Podcast 5: Failed to view details of a favourite podcast")
                print('Podcast 5: Can you view details of a favourite podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I14'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 5: Failed to open podcast favourites")
            print("Podcast 5: Failed to open podcast favourites")
            print(data)
            ws['I14'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 5: Failed to open podcasts")
        print("Podcast 5: Failed to open Podcasts")
        print(data)
        ws['I14'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast6(queue):  # STATUS: TESTED
    # Can you add a custom podcast from favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add custom podcast' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add custom podcast' in data:
                pag.press("enter", interval=1)
                pag.write("test", interval=0.5)
                pag.press("enter", interval=1)
                pag.write("https://podcasts.files.bbci.co.uk/p089sfrz.rss", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                pag.press("t", presses=2, interval=1)
                data = readFromQueue(queue)
                while 'test' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'test' in data:
                    print('Podcast 6: Can you add a custom podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I15'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 6: Failed to add custom podcast")
                    print('Podcast 6: Can you add a custom podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I15'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 6: Failed to add custom podcast via favourites")
                print("Podcast 6: Failed to add custom podcast via favourites")
                print(data)
                ws['I15'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 6: Failed to open podcast favourites")
            print("Podcast 6: Failed to open podcast favourites")
            print(data)
            ws['I15'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 6: Failed to open podcasts")
        print("Podcast 6: Failed to open Podcasts")
        print(data)
        ws['I15'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast7(queue):    # STATUS: TESTED
    # Can you search a podcast list and play via favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1, presses=2)
            pag.sleep(3)
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_SEARCH_EDIT' in data:
                pag.write("central", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'Central heating, Mike Lambert column' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)
                    data = readFromQueue(queue)
                    if '' in data:
                        print('Podcast 7: Can you search and play a podcast via favourites? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I16'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("enter")
                        returnhome()

                    else:
                        logging.critical("Podcast 7: Failed to search and play a podcast via favourites")
                        print('Podcast 7: Can you search and play a podcast via favourites? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I16'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("enter")
                        returnhome()

                else:
                    logging.critical("Podcast 7: Failed to find podcast - Central heating, Mike lambert column")
                    print("Podcast 7: Failed to find Podcast: Central heating, Mike Lambert column")
                    print(data)
                    ws['I16'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 7: Failed to open podcast search via favourites")
                print("Podcast 7: Failed to open podcast search via favourites")
                print(data)
                ws['I16'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 7: Failed to open podcast favourites")
            print("Podcast 7: Failed to open podcast favourites")
            print(data)
            ws['I16'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Podcast 7: Failed to open podcasts")
        print("Podcast 7: Failed to open Podcasts")
        print(data)
        ws['I16'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast8(queue):   # STATUS: TESTED
    # Can you search the play a new podcast list
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_SEARCH_EDIT' in data:
                pag.write("desert", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                while 'Desert Island Discs with Kirsty Young' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Desert Island Discs with Kirsty Young' in data:
                    print('Podcast 8: Can you search for a podcast via find new podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I17'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 8: Failed to search a new podcast via find new podcast")
                    print('Podcast 8: Can you search for a podcast via find new podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I17'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 8: Failed to search for a new podcast")
                print("Podcast 8: Failed to search for new podcast")
                print(data)
                ws['I17'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 8: Failed to open play a new podcast")
            print("Podcast 8: Failed to open Play a new Podcast")
            print(data)
            ws['I17'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 8: Failed to open podcasts")
        print("Podcast 8: Failed to open Podcasts")
        print(data)
        ws['I17'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast9(queue):    # STATUS: TESTED
    # Can you view details of a podcast via find new podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 9: Can you view details for a podcast via find new podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I18'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

            else:
                logging.critical("Podcast 9: Failed to view details of a podcast via new podcast")
                print('Podcast 9: Can you view details for a podcast via find new podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I18'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

        else:
            logging.critical("Podcast 9: Failed to open play a new podcast")
            print("Podcast 9: Failed to open Play a new Podcast")
            print(data)
            ws['I18'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 9: Failed to open podcasts")
        print("Podcast 9: Failed to open Podcasts")
        print(data)
        ws['I18'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast10(queue):   # STATUS: TESTED
    # Can you change the language of the podcast list via find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", presses=2,  interval=1)
            pag.press("enter", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'Fire Patroner Pa Lommen' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Fire Patroner Pa Lommen' in data:
                print('Podcast 10: Can you change the list language via find a podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I19'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

            else:
                logging.critical("Podcast 10: Failed to change language list via find a podcast")
                print('Podcast 10: Can you change the list language via find a podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I19'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

        else:
            logging.critical("Podcast 10: Failed to open play a new podcast")
            print("Podcast 10: Failed to open Play a new Podcast")
            print(data)
            ws['I19'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 10: Failed to open podcasts")
        print("Podcast 10: Failed to open Podcasts")
        print(data)
        ws['I19'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast11(queue):   # STATUS: TESTED
    # Search a podcast and then proceed to play it via find a podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            pag.write("life after", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'Life After Blindness' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Life After Blindness' in data:
                pag.press("enter", interval=1, presses=2)
                pag.sleep(15)
                data = readFromQueue(queue)
                if '' in data:
                    print('Podcast 11: Can you search and then play a podcast via find a podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I20'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 11: Failed to search and play a podcast via find a podcast")
                    print('Podcast 11: Can you search and then play a podcast via find a podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I20'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 11: Failed to search for Life After Blindness")
                print("Podcast 11: Failed to search for Life After Blindness")
                print(data)
                ws['I20'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 11: Failed to open play a new podcast")
            print("Podcast 11: Failed to open Play a new Podcast")
            print(data)
            ws['I20'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 11: Failed to open podcasts")
        print("Podcast 11: Failed to open Podcasts")
        print(data)
        ws['I20'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast12(queue):   # STATUS: TESTED
    # Can you view details of an episode via find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("d")
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 12: Can you view details of a podcast episode? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I21'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.sleep(5)
                returnhome()

            else:
                logging.critical("Podcast 12: Failed to view details of a podcast episode")
                print('Podcast 12: Can you view details of a podcast episode? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I21'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.sleep(5)
                returnhome()

        else:
            logging.critical("Podcast 12: Failed to open play a new podcast")
            print("Podcast 12: Failed to open Play a new Podcast")
            print(data)
            ws['I21'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 12: Failed to open podcasts")
        print("Podcast 12: Failed to open Podcasts")
        print(data)
        ws['I21'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast13(queue):  # STATUS: TESTED
    # Can you add a podcast to favourites from find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("f", interval=1)
            data = readFromQueue(queue)
            while 'Farming Today' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Farming Today' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1, presses=2)
                pag.sleep(3)
                pag.press("esc", interval=1)
                pag.press("f")
                pag.press("enter", interval=1)
                pag.press("f")
                pag.sleep(3)
                data = readFromQueue(queue)
                while 'Farming Today' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Farming Today' in data:
                    print('Podcast 13: Can you add a podcast to favourites from find a podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I22'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 13: Failed to add a podcast to favourites from find a podcast")
                    print('Podcast 13: Can you add a podcast to favourites from find a podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I22'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 13: Failed to find podcast Farming Today")
                print("Podcast 13: Failed to find podcast: Farming Today")
                print(data)
                ws['I22'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 13: Failed to open play a new podcast")
            print("Podcast 13: Failed to open Play a new Podcast")
            print(data)
            ws['I22'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 13: Failed to open podcasts")
        print("Podcast 13: Failed to open Podcasts")
        print(data)
        ws['I22'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast14(queue):   # STATUS: TESTED
    # Can you download a podcast and play it from my downloads?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)

    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'A History of the World in 100 Objects' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'A History of the World in 100 Objects' in data:
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("d", presses=2, interval=1)
                pag.press("enter", interval=1, presses=2)
                pag.press("esc", presses=2, interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'My downloaded podcast episodes' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'My downloaded podcast episodes' in data:
                    pag.sleep(20)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'A History of the World in 100 Objects - Object 101' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'A History of the World in 100 Objects - Object 101' in data:
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        if '' in data:
                            pag.sleep(10)
                            pag.press("enter", interval=1)
                            print('Podcast 14: Can you download and play a podcast? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I23'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Podcast 14: Failed to download and play a podcast")
                            print('Podcast 14: Can you download and play a podcast? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I23'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Podcast 14: A History of the World in 100 Objects not in downloads ")
                        print("Podcast 14: A History of the World in 100 Objects - Object 101 not present in downloads")
                        print(data)
                        ws['I23'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 14: Failed to navigate to my downloaded podcasts")
                    print("Podcast 14: Failed to navigate to My downloaded podcasts")
                    print(data)
                    ws['I23'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 14: Failed to find podcast A history of the world")
                print("Podcast 14: Failed to find podcast: A History of the world in 100 Objects")
                print(data)
                ws['I23'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 14: Failed to open play a new podcast")
            print("Podcast 14: Failed to open Play a new Podcast")
            print(data)
            ws['I23'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 14: Failed to open podcasts")
        print("Podcast 14: Failed to open Podcasts")
        print(data)
        ws['I23'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast15(queue):  # STATUS: TESTED
    # views details of a downloaded podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'A History of the World in 100 Objects - Object 101' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'A History of the World in 100 Objects - Object 101' in data:
            pag.press("f2", interval=1)
            pag.press("d", presses=2, interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_DETAILS' in data:
                pag.sleep(5)
                print('Podcast 15: Can you view details of a downloaded podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I24'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                pag.sleep(5)
                logging.critical("Podcast 15: Failed to view details of a downloaded podcast")
                print('Podcast 15: Can you view details of a downloaded podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I24'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 15: Failed to download podcast")
            print("Podcast 15: Failed to locate downloaded podcast: "
                  "A History of the World in 100 Objects - Object 101")
            print(data)
            ws['I24'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 15: Failed to open podcasts")
        print("Podcast 15: Failed to open Podcasts")
        print(data)
        ws['I24'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast16(queue):   # STATUS: TESTED
    # Copy a downloaded podcast to a USB
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Radio' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Radio' in data:
        pag.press("p")
        data = readFromQueue(queue)
        if 'Podcasts' in data:
            pag.press("enter", interval=1)
            pag.press("m", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'A History of the World in 100 Objects - Object 101' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'A History of the World in 100 Objects - Object 101' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                if f"{found[0]}" in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_COPYTODEVICE_PROGRESS' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)
                        if 'OK' in data:
                            break

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        print('Podcast 16: Can you copy a podcast to a device? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I25'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Podcast 16: Failed to copy a podcast to a USB")
                        pag.press("enter", interval=1)
                        print('Podcast 16: Can you copy a podcast to a device? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I25'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 16: Failed to open copy to device")
                    print("Podcast 16: Failed to open Copy to device")
                    print(data)
                    ws['I25'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 16: Failed to locate downloaded podcast")
                print("Podcast 16: Failed to locate downloaded podcast:"
                      " A History of the World in 100 Objects - Object 101")
                print(data)
                ws['I25'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 16: Failed to open podcasts")
            print("Podcast 16: Failed to open Podcasts")
            print(data)
            ws['I25'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 16: Failed to enter entertainment")
        print("Podcast 16: Failed to enter Entertainment")
        print(data)
        ws['I25'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast17(queue):   # STATUS: TESTED
    # deletes a downloaded podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'A History of the World in 100 Objects - Object 101' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'A History of the World in 100 Objects - Object 101' in data:
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            data = readFromQueue(queue)

            if 'No downloaded podcasts' in data:
                print('Podcast 17: Can you delete a podcast from my downloaded podcasts? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I26'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Podcast 17: Failed to delete podcasts from my downloaded podcasts")
                print('Podcast 17: Can you delete a podcast from my downloaded podcasts? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I26'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 17: Failed to find downloaded podcast")
            print("Podcast 17: Failed to find downloaded podcast")
            print(data)
            ws['I26'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 17: Failed to open podcasts")
        print("Podcast 17: Failed to open Podcasts")
        print(data)
        ws['I26'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast18(queue):
    # Delete a podcast from a USB
    start = 0
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.sleep(5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1)
            pag.press("o", interval=1)
            data = readFromQueue(queue)
            if 'Object 101' in data:
                pag.press("del", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("o", interval=1)
                data = readFromQueue(queue)
                if 'testdata' in data:
                    print('Podcast 18: Delete a podcast from USB \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I192'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 18: Failed to delete a podcast from USB")
                    print('Podcast 18: Delete a podcast from USB \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I192'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Podcast 18: Failed to locate Object 101 on USB device")
                print("Podcast 18: Failed to locate Object 101 On USB device")
                print(data)
                ws['I192'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 18: Failed to locate regression USB")
            print("Podcast 18: failed to locate regression USB")
            print(data)
            ws['I192'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 18: Failed to locate file explorer")
        print("Podcast 18: failed to locate file explorer")
        print(data)
        ws['I192'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

def music_1(queue):  # STATUS: TESTED
    start = 0
    delay()
    # Can you import music?
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)
    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
        data = readFromQueue(queue)

    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if '9TAILS & guccihighwaters - NOVEMBER' in data:
        print('Music 1: Can you import Music? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I27'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 1: Failed to import music")
        print('Music 1: Can you import Music? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I27'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_2(queue):  # STATUS: TESTED
    # ALBUM VIEW: Can you play a track from album?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1, presses=4)
    pag.sleep(10)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 2: ALBUM VIEW: Can you play a track? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I28'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 2: Failed to play a track")
        pag.press("enter", interval=1)
        print('Music 2: ALBUM VIEW: Can you play a track? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I28'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_3(queue):     # STATUS: TESTED
    # Can you continue listening to a song?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(10)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 3: ALBUM VIEW: Can you continue playing a track? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I29'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 3: Failed to continue playing a track")
        pag.press("enter", interval=1)
        print('Music 3: ALBUM VIEW: Can you continue playing a track? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I29'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_4(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you view and change album details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while '9TAILS & guccihighwaters - NOVEMBER' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '9TAILS & guccihighwaters - NOVEMBER' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'Details' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Details' in data:
            pag.press("enter", presses=2, interval=1)
            pag.press("u", interval=1)
            data = readFromQueue(queue)

            while 'Unknown Album [55CC5344]' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Unknown Album [55CC5344]' in data:
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'Unknown Album [55CC5344]' in data:
                    print('Music 4: ALBUM VIEW: Can you view and change the album details? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I30'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Music 4: Failed to change album details")
                    print('Music 4: ALBUM VIEW: Can you view and change the album details? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I30'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 4: Failed to switch albums")
                print('Music 4: ALBUM VIEW: Failed to switch albums \n'
                      '>>> Current String:', data)
                ws['I30'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 4: Failed to open details")
            print('Music 4: ALBUM VIEW: Failed to open details \n'
                  '>>> Current String:', data)
            ws['I30'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 4: Failed to locate 9Tails Album")
        print('Music 4: ALBUM VIEW: Failed to locate 9Tails album \n'
              '>>> Current String:', data)
        ws['I30'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_5(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you rename an album?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Unknown Album [55CC5344]' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Unknown Album [55CC5344]' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'Rename' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Rename' in data:
            pag.press("enter", interval=1)
            pag.write("A", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'A' in data:
                print('Music 5: ALBUM VIEW: Can you rename an existing album? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I31'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 5: Failed to rename existing album")
                print('Music 5: ALBUM VIEW: Can you rename an existing album? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I31'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 5: Failed to locate rename album")
            print('Music 5: ALBUM VIEW: Failed to locate rename album \n'
                  '>>> Current String:', data)
            ws['I31'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 5: Failed to find album")
        print('Music 5: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I31'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_6(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you add a new album in details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'A' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'A' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ALBUM_ACTIONS_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ALBUM_ACTIONS_ADD' in data:
            pag.write("Added", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Added' in data:
                print('Music 6: ALBUM VIEW: Can you add a new album via details \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I32'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 6: Failed to add new album details")
                print('Music 6: ALBUM VIEW: Can you add a new album via details \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I32'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 6: Failed to add new album")
            print('Music 6: ALBUM VIEW: Failed to add new album \n'
                  '>>> Current String:', data)
            ws['I32'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 6: Failed to find album")
        print('Music 6: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I32'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_7(queue):     # STATUS: TESTED
    # ALBUM VIEW: Perform an album search
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'A' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'A' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' in data:
            pag.write("Added", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Added' in data:
                print('Music 7: ALBUM VIEW: Can you search for a specific album? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I33'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 7: Failed to search for specific album")
                print('Music 7: ALBUM VIEW: Can you search for a specific album? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I33'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 7: Failed to open search album")
            print('Music 7: ALBUM VIEW: Failed to open search album \n'
                  '>>> Current String:', data)
            ws['I33'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 7: Failed to find album")
        print('Music 7: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I33'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_8(queue):     # STATUS: TESTED
    # ALBUM VIEW (Track list) - Can you view and change all the details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while '9TAILS & guccihighwaters - for once (prod. notmorgn)' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '9TAILS & guccihighwaters - for once (prod. notmorgn)' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TITLE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TITLE' in data:
            pag.write("changed", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("a", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while '9TAILS Archive' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if '9TAILS Archive' in data:
                pag.press("enter", interval=1)
                pag.press("a", presses=2, interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'A' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'A' in data:
                    pag.press("a", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("t", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TRACK' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TRACK' in data:
                        pag.write("7", interval=0.5)
                        pag.press("enter", interval=1)
                        pag.press("g", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_GENRE' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_GENRE' in data:
                            pag.write("changed", interval=0.5)
                            pag.press("enter", interval=1)
                            pag.press("y", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_YEAR' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_YEAR' in data:
                                pag.write("2021", interval=0.5)
                                pag.press("enter", interval=1)
                                pag.press("esc", presses=2, interval=1)
                                pag.press("a", interval=1)
                                pag.press("enter", interval=1)
                                data = readFromQueue(queue)
                                if 'changed' in data:
                                    print('Music 8: ALBUM VIEW: Can you edit the track list details? \n'
                                          '>>> Result: PASS \n'
                                          '>>> Current String:', data)
                                    ws['I34'].value = 'PASS'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                                else:
                                    logging.critical("Music 8: Failed to edit track list details")
                                    print('Music 8: ALBUM VIEW: Can you edit the track list details? \n'
                                          '>>> Result: FAIL \n'
                                          '>>> Current String:', data)
                                    ws['I34'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                            else:
                                logging.critical("Music 8: Failed to open edit year")
                                print('Music 8: ALBUM VIEW: Failed to open edit year \n'
                                      '>>> Current String:', data)
                                ws['I34'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Music 8: Failed to open change genre")
                            print('Music 8: ALBUM VIEW: Failed to open change genre \n'
                                  '>>> Current String:', data)
                            ws['I34'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Music 8: Failed to open edit track number")
                        print('Music 8: ALBUM VIEW: Failed to open edit track number \n'
                              '>>> Current String:', data)
                        ws['I34'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Music 8: Failed to open change album")
                    print('Music 8: ALBUM VIEW: Failed to open change album \n'
                          '>>> Current String:', data)
                    ws['I34'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 8: Failed to open change artist")
                print('Music 8: ALBUM VIEW: Failed to open change artist \n'
                      '>>> Current String:', data)
                ws['I34'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 8: Failed to open edit song title")
            print('Music 8: ALBUM VIEW: Failed to open edit song title \n'
                  '>>> Current String:', data)
            ws['I34'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 8: Failed to find album")
        print('Music 8: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I34'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_9(queue):     # STATUS: TESTED
    # ALBUM VIEW (Track list) - Can you delete a track?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("a", interval=1)
    data = readFromQueue(queue)

    while 'Added' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Added' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'No items found' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'No items found' in data:
            print('Music 9: ALBUM VIEW: Can you delete a track?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I35'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 9: Failed to delete a track")
            print('Music 9: ALBUM VIEW: Can you delete a track?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I35'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 9: Failed to locate album 'Added'")
        print('Music 9: ALBUM VIEW: Failed to locate Album "Added" \n'
              '>>> Current String:', data)
        ws['I35'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_10(queue):        # STATUS: TESTED
    # ALBUM VIEW: Search for a specific track in an album
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("s", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_NEXT_EDIT' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_NEXT_EDIT' in data:
        pag.write("scars", interval=0.5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9TAILS & guccihighwaters - scars (prod. notmorgn)' in data:
            print('Music 10: ALBUM VIEW: Can you search for a specific track?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I36'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 10: Failed to search for specific track")
            print('Music 10: ALBUM VIEW: Can you search for a specific track?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I36'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 10: Failed to open search for song")
        print('Music 10: ALBUM VIEW: Failed to open search for song \n'
              '>>> Current String:', data)
        ws['I36'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_11(queue):    # STATUS: TESTED
    # ALBUM VIEW: Can you remove albums then import via the F2 feature
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    for _ in range(4):
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'Remove' in data:
            pag.press("enter", presses=3, interval=1)

        else:
            print("failed to locate remove")
            break

    data = readFromQueue(queue)
    if 'No items found' in data:
        pag.press("f2", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
            data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if '9TAILS & guccihighwaters - NOVEMBER' in data:
                print('Music 11: ALBUM VIEW: Can you remove all albums and then import via the f2 menu  \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I37'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 11: Failed to remove all albums and re import via f2 menu")
                print('Music 11: ALBUM VIEW: Can you remove all albums and then import via the f2 menu  \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I37'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 11: Failed to import music")
            print('Music 11: ALBUM VIEW: Failed to import music \n'
                  '>>> Current String:', data)
            ws['I37'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 11: Failed to delete all albums")
        print('Music 11: ALBUM VIEW: Failed to delete all albums \n'
              '>>> Current String:', data)
        ws['I37'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_12(queue):    # STATUS: TESTED
    # ALBUM VIEW: Can you change your view?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("a", interval=1)
    data = readFromQueue(queue)

    while 'Artist view' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Artist view' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9TAILS Archive' in data:
            print('Music 12: ALBUM VIEW: Can you change to artist view?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I38'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 12: Failed to change to artist view")
            print('Music 12: ALBUM VIEW: Can you change to artist view?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I38'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 12: Failed to switch artist view")
        print('Music 12: ALBUM VIEW: Failed to switch to Artist View \n'
              '>>> Current String:', data)
        ws['I38'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_13(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you play a track?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(10)
    data = readFromQueue(queue)
    if '' in data:
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I39'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    elif 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I39'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 13: Failed to play a track")
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I39'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_14(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you add and change the artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_ADD' in data:
            pag.write("tester", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("enter", presses=2, interval = 1)
            data = readFromQueue(queue)
            if 'tester' in data:
                print('Music 14: ARTIST VIEW: Can you add and change the artist details? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I40'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 14: Failed to change artist details")
                print('Music 14: ARTIST VIEW: Can you add and change the artist details? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I40'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 14: Failed to add a new artist")
            print('Music 14 ARTIST VIEW: Failed to add a new artist \n'
                  '>>> Current String:', data)
            ws['I40'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 14: Failed to locate view details")
        print('Music 14 ARTIST VIEW: Failed to locate view details \n'
              '>>> Current String:', data)
        ws['I40'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_15(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you rename an artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_RENAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_RENAME' in data:
            pag.write("renamed", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'renamed' in data:
                print('Music 15: ARTIST VIEW: Can you rename the artist details? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I41'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 15: Failed to rename artist details")
                print('Music 15: ARTIST VIEW: Can you rename the artist details? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I41'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 15: Failed to rename artist")
            print('Music 15 ARTIST VIEW: Failed to rename artist \n'
                  '>>> Current String:', data)
            ws['I41'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 15: Failed to locate view details")
        print('Music 15 ARTIST VIEW: Failed to locate view details \n'
              '>>> Current String:', data)
        ws['I41'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_16(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you remove an artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    for _ in range(2):
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

    if 'No items found' in data:
        print('Music 16: ARTIST VIEW: Can you remove an artist? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I42'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 16: Failed to remove an artist")
        print('Music 16: ARTIST VIEW: Can you remove an artist? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I42'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_17(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you perform a search?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
        data = readFromQueue(queue)

    while 'OK' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'OK' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if '9TAILS Archive' in data:
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' in data:
                pag.write("9", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                pag.press("right", interval=1)
                if '9TAILS Archive' in data:
                    print('Music 17: ARTIST VIEW: Can you import music and then search for an artist? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I43'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Music 17: Failed to import music and search for an artist")
                    print('Music 17: ARTIST VIEW: Can you import music and then search for an artist? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I43'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 17: Failed to search for an artist")
                print('Music 17 ARTIST VIEW: Failed to search for an artist \n'
                      '>>> Current String:', data)
                ws['I43'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 17: Failed to locate album")
            print('Music 17 ARTIST VIEW: Failed to locate album \n'
                  '>>> Current String:', data)
            ws['I43'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 17: Failed to import music")
        print('Music 17 ARTIST VIEW: Failed to import music \n'
              '>>> Current String:', data)
        ws['I43'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_18(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you change to folder view?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Folder view' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'C:' in data:
            print('Music 18: ARTIST VIEW: Can you change to folder view?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I44'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 18: Failed to change to folder view")
            print('Music 18: ARTIST VIEW: Can you change to folder view?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I44'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 18: Failed to switch to folder view")
        print('Music 18: ALBUM VIEW: Failed to switch to folder View \n'
              '>>> Current String:', data)
        ws['I44'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_19(queue):    # STATUS: TESTED
    # FOLDER VIEW: Search through folders to play a track
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'C:' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9tails' in data:
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            data = readFromQueue(queue)
            if '' in data:
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            elif 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 19: Failed to play a track")
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 19: Failed to locate 9Tails folder")
            print('Music 19: FOLDER VIEW: Failed to locate 9Tails folder \n'
                  '>>> Current String:', data)
            ws['I45'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 19: Failed to be in the folder view")
        print('Music 19: FOLDER VIEW: Failed to be in folder View \n'
              '>>> Current String:', data)
        ws['I45'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

def Hangman(queue):
    # Launches Hangman
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", presses=4, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_HANGMAN' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_HANGMAN' in data:
            pag.sleep(5)
            print('Games 1 - Hangman: Can you launch hangman? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical("Games 1 - Failed to launch Hangman")
            print('Games 1 - Hangman: Can you launch hangman? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 1 - Failed to open games")
        print("Games 1: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def Sudoku(queue):
    # TEST CASE 2: LAUNCH SUDOKU
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_SUDOKU' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_SUDOKU' in data:
            pag.sleep(5)
            print('Games 2 - Sudoku: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical("Games 2 - Failed to launch Sudoku")
            print('Games 2 - Sudoku: Can you launch Sudoku? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 2 - Failed to open games")
        print("Games 2 - Sudoku: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def Blackjack(queue):
    # TEST CASE 3: LAUNCH BLACKJACK
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", presses=2, interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)
        pag.sleep(5)

        while 'GMENU_GAMES_BLACKJACK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_BLACKJACK' in data:
            print('Games 3 - Blackjack: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical('Games 3 - Failed to launch Blackjack')
            print('Games 3 - Blackjack: Can you launch Blackjack? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 3 - Blackjack: Failed to open games")
        print("Games 3 - Blackjack: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()

                                            # MODULE: ENTERTAINMENT END
#######################################################################################################################

                                                 # MODULE: NOTES

def tnote1(queue):  # STATUS: TESTED
    # with no notes present, does the prompt say no notes available?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    pag.sleep(3)
    data = readFromQueue(queue)
    if 'No notes available' in data:
        print('Notes 1: with no notes present, does the prompt say no notes available? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I62'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Notes 1: No available notes prompt missing")
        print('Notes 1: with no notes present, does the prompt say no notes available? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I62'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote2(queue): # STATUS: TESTED
    # Can you create a new text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Add new text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Add new text note' in data:
        pag.press("enter", interval=1)
        pag.write("text note", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("this is a text note", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'text note' in data:
            print('Notes 2: Can you create and view a text note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I63'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 2: Failed to create and view a text note")
            print('Notes 2: Can you create and view a text note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I63'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 2: Failed to add a new text note")
        print("Notes 2: failed to add a new text note")
        print(data)
        ws['I63'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote3(queue):  # STATUS: TESTED
    # can you rename an existing text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'text note' in data:
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_NOTES_MAIN_RENAME_TEXT_TITLE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_NOTES_MAIN_RENAME_TEXT_TITLE' in data:
            pag.write("renamed", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.sleep(5)
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'renamed' in data:
                print('Notes 3: Can you rename an existing text note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I64'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Notes 3: Failed to rename existing text note")
                print('Notes 3: Can you rename an existing text note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I64'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 3: Failed to rename text note")
            print("Notes 3: failed to Rename text note")
            print(data)
            ws['I64'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 3: Text note not found")
        print("Notes 3: text note not found in notes")
        print(data)
        ws['I64'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote4(queue):  # STATUS: TESTED
    # can you edit a text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("edited ", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(3)
        data = readFromQueue(queue)
        if 'OK' in data:
            pag.press("enter", interval=1)
            print('Notes 4: Can you edit a text note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I65'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 4: Failed to edit a text note")
            print('Notes 4: Can you edit a text note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I65'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 4: Failed to enter text note")
        print("Notes 4: Failed to enter text note")
        print(data)
        ws['I65'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote5(queue):  # STATUS: TESTED
    # Can you send a note via an email?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)

        while 'Send as email' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Send as email' in data:
            pag.press("enter", interval=1, presses=3)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Text note", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("text note attached", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_EMAIL_SENDING' in data:
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'OK' in data:
                pag.press("enter", interval=1)
                returnhome()
                pag.sleep(3)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)

                while 'guideautomation2@outlook.com'not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter", interval=1)
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.sleep(3)
                        pag.press("f2", interval=1)
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'renamed.txt' in data:
                            print('Notes 5: Can you send a note via email? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I66'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Notes 5: Failed to send text note via attachment")
                            print('Notes 5: Can you send a note via email? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I66'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Notes 5: Failed to locate email with attachment")
                        print("Notes 5: failed to locate email with attachment")
                        print(data)
                        ws['I66'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Notes 5: Failed to enter guide automation 2")
                    print("Notes 5: Failed to enter guide automation 2")
                    print(data)
                    ws['I66'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Notes 5: Failed to send email with attachment")
                print("Notes 5: Failed to send email with attachment")
                print(data)
                ws['I66'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 5: Failed to navigate to send as email")
            print("Notes 5: failed to navigate to send as email")
            print(data)
            ws['I66'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 5: Unable to find text note")
        print("Notes 5: unable to find text note")
        print(data)
        ws['I66'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote6(queue):  # STATUS: TESTED
    # Can you view details and delete a note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.sleep(3)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_NOTES_MAIN_ACTIONS_DETAILS' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_NOTES_MAIN_ACTIONS_DETAILS' in data:
            pag.sleep(3)
            pag.press("esc", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", presses=2, interval=1)
            pag.press("enter", interval=1, presses=3)
            data = readFromQueue(queue)
            if 'No notes available' in data:
                print('Notes 6: Can you view details and then delete a note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I67'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Notes 6: Failed to view details and delete a note")
                print('Notes 6: Can you view details and then delete a note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I67'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 6: Unable to view details of a note")
            print("Notes 6: unable to view details of note")
            print(data)
            ws['I67'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 6: Unable to find text note")
        print("Notes 6: unable to find text note")
        print(data)
        ws['I67'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote7(queue):  # STATUS: TESTED
    # Can you create a note without saving?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Add new text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Add new text note' in data:
        pag.press("enter", interval=1)
        pag.write("text note", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("this is a text note", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1, presses=2)
        data = readFromQueue(queue)
        if 'No notes available' in data:
            print('Notes 7: Can you create a text note without saving? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I68'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 7: Failed to ESC a text note without saving ")
            print('Notes 7: Can you create a text note without saving? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I68'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 7: Failed to add a new text note")
        print("Notes 7: Failed to add a new text note")
        print(data)
        ws['I68'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def anote1(queue):  # STATUS: TESTED
    # Can you create a new audio note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_NOTES_MAIN_VOICE_TITLE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_NOTES_MAIN_VOICE_TITLE' in data:
        pag.write("voice note", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(5)
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("v")
        data = readFromQueue(queue)
        if 'voice note' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)
            if '' in data:
                print('Audio Notes 1: Can you create and view an audio note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I69'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Audio Notes 1: Failed to create a new audio note")
                print('Audio Notes 1: Can you create and view an audio note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I69'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Audio notes 1: Failed to find voice note")
            print("Audio Notes 1: failed to find voice note")
            print(data)
            ws['I69'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Audio notes 1: Failed to add new voice note")
        print("Audio Notes 1: failed to add new voice note")
        print(data)
        ws['I69'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def anote2(queue):  # STATUS: TESTED
    # can you send a audio note via email?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    pag.sleep(3)
    data = readFromQueue(queue)

    while 'voice note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'voice note' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)

        while 'Send as email' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Send as email' in data:
            pag.press("enter", interval=1, presses=3)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Audio note", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("audio note attached", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_EMAIL_SENDING' in data:
                print("Email Sending...")
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'OK' in data:
                pag.press("enter", interval=1)
                returnhome()
                pag.sleep(3)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)

                while 'guideautomation2@outlook.com' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter")
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.sleep(3)
                        pag.press("f2")
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'voice note.mp3' in data:
                            print('Audio note 2: Can you send a note via email? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I70'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Audio note 2: Failed to send audio note via email")
                            print('Audio note 2: Can you send a note via email? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I70'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Audio note 2: Failed to locate email with attachement")
                        print("Audio note 2: failed to locate email with attachment")
                        print(data)
                        ws['I70'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Audio note 2: Failed to enter guide automation")
                    print("Audio note 2: Failed to enter guide automation 2")
                    print(data)
                    ws['I70'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Audio note 2: Failed to send email with attachment")
                print("Audio note 2: Failed to send email with attachment")
                print(data)
                ws['I70'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Audio note 2: Failed to navigate to send as email")
            print("Audio note 2: failed to navigate to send as email")
            print(data)
            ws['I70'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Audio note 2: Unable to find audio note")
        print("Audio note 2: unable to find audio note")
        print(data)
        ws['I70'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

                                                # MODULE: NOTES END
#######################################################################################################################

                                                  # MODULE: TOOLS

def about(queue):   # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Check my Premium Plan' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Check my Premium Plan' in data:
        pag.press("enter", interval=1)
        pag.sleep(2)
        print("About 1: Check Premium plan status \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['I104'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Tools 1: Failed to check premium plan status")
        print("About 1: Check Premium plan status \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I104'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calc(queue):    # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_CALCULATOR_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_CALCULATOR_MAIN' in data:
        # TEST CASE 1: 50,000 + 90,000 = 140,000
        pag.press("5", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("+")
        pag.press("9", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 2: 666 / 22 = 30.272
        pag.press("6", presses=3, interval=0.5)
        pag.press("/")
        pag.press("2", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 3: 200 - 11 = 189
        pag.press("2", interval=0.5)
        pag.press("0", presses=2, interval=0.5)
        pag.press("-")
        pag.press("1", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 4: 497 x 15 = 7,455
        pag.press("4", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("*")
        pag.press("1", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)

        print("Calc 1: Use the calculator to add, multiply, divide and minus \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['I105'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Tools 2: Failed to use the calculator")
        print("Calc 1: Use the calculator to add, multiply, divide and minus \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I105'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def dict_valid(queue):  # STATUS: TESTED
    delay()
    start = 0
    # TEST CASE 1: Search a valid valid dictionary entry = test
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("test", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' in data:  # change back too
            print("Dictionary Valid 1: Can you search for a valid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I106'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 3: Failed to search a valid dictionary term")
            print("Dictionary Valid 1: Can you search for a valid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I106'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 3: Failed to open dictionary")
        print("Dictionary Valid 1: Failed to open dictionary  \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I106'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def dict_invalid(queue):    # STATUS: TESTED
    # TEST CASE 2: Can you search an invalid search term in dictionary
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("twest", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            print("Dictionary Invalid 1: Can you search for an invalid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I107'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 4: Failed to search an invalid dictionary term")
            print("Dictionary Invalid 1: Can you search for an invalid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I107'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 4: Failed to open dictionary")
        print("Dictionary Invalid 1: Failed to open dictionary \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I107'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def gstart_vids(queue):  # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'Keyboard and mouse video' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Keyboard and mouse video' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("esc", interval=1)
        pag.press("R", interval=1)
        data = readFromQueue(queue)
        if 'Remote control video' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            pag.press("esc", interval=1)
            pag.press("T", interval=1)
            data = readFromQueue(queue)
            if 'Touch screen video' in data:
                pag.press("enter", interval=1)
                pag.sleep(5)
                pag.press("esc", interval=1)
                pag.press("V", interval=1)
                data = readFromQueue(queue)
                if 'Voice input video' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)

                    print("Getting Started Videos: Can you play the Getting Started Videos \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I108'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 5: Failed to find Voice Video")
                    print("Getting Started Videos: Voice video missing")
                    print("Current String:", data)
                    ws['I108'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 5: Failed to find Touch Video")
                print("Getting Started Videos: Touch video missing")
                print("Current String:", data)
                ws['I108'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 5: Failed to find Remote video")
            print("Getting Started Videos: Remote video missing")
            print("Current String:", data)
            ws['I108'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 5: Failed to find M&K Video")
        print("Getting Started Videos: Keyboard and mouse video missing \n"
              "Result: FAIL -  Keyboard video not present \n"
              "Current String:", data)
        ws['I108'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def learn_keyboard(queue):  # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_TYPING_TUTOR_DIALOGUE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_TYPING_TUTOR_DIALOGUE' in data:
        pag.press("1", interval=0.5)
        pag.press("2", interval=0.5)
        pag.press("3", interval=0.5)
        pag.press("4", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("6", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("8", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("q", interval=0.5)
        pag.press("w", interval=0.5)
        pag.press("e", interval=0.5)
        pag.press("r", interval=0.5)
        pag.press("t", interval=0.5)
        pag.press("y", interval=0.5)
        pag.press("u", interval=0.5)
        pag.press("i", interval=0.5)
        pag.press("o", interval=0.5)
        pag.press("p", interval=0.5)
        pag.press("a", interval=0.5)
        pag.press("s", interval=0.5)
        pag.press("d", interval=0.5)
        pag.press("f", interval=0.5)
        pag.press("g", interval=0.5)
        pag.press("h", interval=0.5)
        pag.press("j", interval=0.5)
        pag.press("k", interval=0.5)
        pag.press("l", interval=0.5)
        pag.press("z", interval=0.5)
        pag.press("x", interval=0.5)
        pag.press("c", interval=0.5)
        pag.press("v", interval=0.5)
        pag.press("b", interval=0.5)
        pag.press("n", interval=0.5)
        pag.press("m", interval=0.5)
        pag.press("enter", interval=0.5)
        data = readFromQueue(queue)
        if 'Learn the Keyboard' in data:
            print("LTK 1: Can you use the typing tutor? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I109'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 6: Failed to use typing tutor")
            print("LTK 1: Can you use the typing tutor? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I109'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 6: Failed to enter Learn the keyboard")
        print("LTK 1: Could not enter learn the keyboard \n"
              "Current String:", data)
        ws['I109'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def shortcut(queue):    # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    data = readFromQueue(queue)
    while 'Training' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Training' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)
        while 'Shortcut keys' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Shortcut keys' in data:
            pag.press("enter", interval=1)
            while 'GMENU_COREMODULE_SHORTCUTKEYS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_COREMODULE_SHORTCUTKEYS' in data:
                print("Shortcuts 1: Can you view the shortcuts documentation? \n"
                      "Result: PASS \n"
                      "Current String:", data)
                ws['I110'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Tools 7: Failed to view shortcut keys")
                print("Shortcuts 1: Can you view the shortcuts documentation? \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I110'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 7: Failed to enter shortcut keys")
            print("Shortcuts 1: failed to enter shortcut keys")
            print(data)
            ws['I110'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 7: Failed to navigate to Training function")
        print("Shortcuts 1: failed to navigate to training")
        print("data")
        ws['I110'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def backup(queue):  # STATUS: TESTED
    # Backs up Guide Connect
    start = 0
    delay()

    # ADD A TEST NOTE
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("a", interval=1)
    pag.press("enter", interval=1)
    pag.write("restore", interval=0.5)
    pag.press("enter", interval=1)
    pag.press("esc", interval=1)
    pag.press("enter", presses=2, interval=1)
    returnhome()

    # BEGIN BACK UP TEST
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)
    if 'Backup and restore' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            returnhome()
            delay()
            pag.press("n", interval=1)
            pag.press("enter", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'restore' in data:
                pag.press("f2", interval=1)
                pag.press("d", presses=2, interval=1)
                pag.press("enter", presses=3, interval=1)
                data = readFromQueue(queue)
                if 'No notes available' in data:
                    print("Back up 1: Back up GC \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I111'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 8: Failed to back up GC")
                    print("Back up 1: Back up GC")
                    print(data)
                    ws['I111'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Tools 8: Failed to locate Restore Note")
                print("Back up 1: Failed to locate note 'Restore'")
                print(data)
                ws['I111'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 8: Failed to Back up GC")
            print("Back up 1: Failed to back up GC")
            print(data)
            ws['I111'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 8: Failed to locate back up and restore")
        print("failed to locate back and restore")
        print(data)
        ws['I111'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def restore(queue): # STATUS: TESTED
    # Backs up Guide Connect
    start = 0
    launch = 300
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)
    if 'Backup and restore' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'Restore GuideConnect' in data:
            pag.press("enter", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)

            while 'Emails' not in data:
                data = readFromQueue(queue)
                if start < launch:
                    pag.sleep(10)
                    start += 10
                    if start == launch:
                        print(data, ": Timed out after", launch, "Seconds")
                        returnhome()
                        break

            if 'Emails' in data:
                delay()
                pag.press("n", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'restore' in data:
                    print("Restore 1: restore GC \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I112'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 9: Failed to restore GC")
                    print("Restore 1: restore GC \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I112'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 9: Failed to find Emails in data on relaunch")
                print("Restore 1: restore GC \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I112'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 9: Failed to locate restore")
            print("Restore 1: failed to locate restore")
            print(data)
            ws['I112'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 9: Failed to locate back up and restore")
        print("failed to locate back and restore")
        print(data)
        ws['I112'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def reset_settings(queue):  # STATUS: TESTED
    # Resets settings only
    start = 0
    launch = 300
    delay()
    # CHANGE TILE VIEW TO OFF
    pag.press("s", presses=2, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("t", presses=2, interval=1)
    data = readFromQueue(queue)
    if 'Tile view mode' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("t", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Off' in data:
            returnhome()
            pag.press("t", interval=1)
            pag.press("enter", interval=1)
            pag.press("b", interval=1)
            data = readFromQueue(queue)
            if 'Backup and restore' in data:
                pag.press("enter", interval=1)
                pag.press("r", presses=2, interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'Reset settings only' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)

                    while 'Emails' not in data:
                        data = readFromQueue(queue)
                        if start < launch:
                            pag.sleep(10)
                            start += 10
                            if start == launch:
                                print(data, ": Timed out after", launch, "Seconds")
                                returnhome()
                                break

                    if 'Emails' in data:
                        pag.press("s", presses=2, interval=1)
                        pag.press("enter", presses=2, interval=1)
                        pag.press("t", presses=2, interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'On' in data:
                            print("Reset 1: Reset Settings Only \n"
                                  "Result: PASS \n"
                                  "Current String:", data)
                            ws['I113'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Tools 10: Failed to reset GC Settings only")
                            print("Reset 1: Reset Settings Only \n"
                                  "Result: FAIL \n"
                                  "Current String:", data)
                            ws['I113'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Tools 10: Failed to relaunch GC after reset settings")
                        print("Failed to relaunch GC after Re setting settings ")
                        print(data)
                        ws['I113'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Tools 10: Failed to locate reset settings only")
                    print("Failed to locate reset settings only")
                    print(data)
                    ws['I113'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 10: Failed to locate back up and restore")
                print("Failed to locate back up and restore")
                print(data)
                ws['I113'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 10: Failed to turn tile view off")
            print("Failed to turn tile view off")
            print(data)
            ws['I113'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 10: Failed to navigate to tile view mode in settings")
        print("Failed to navigate to tile view mode setting")
        print(data)
        ws['I113'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def factory(queue): # STATUS: TESTED
    # Factory resets all settings
    start = 0
    launch = 300
    init()
    pag.press("enter", presses=2, interval=1)
    pag.write("guideautomation1@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("test", interval=0.5)
    pag.press("enter", presses=3, interval=1)
    pag.write("guideconnect1", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)
    while 'OK' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'OK' in data:
        returnhome()
        delay()
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("b", interval=1)
        data = readFromQueue(queue)
        if 'Backup and restore' in data:
            pag.press("enter", interval=1)
            pag.press("r", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("f", interval=1)
            data = readFromQueue(queue)
            if 'Full factory reset' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

                while 'Emails' not in data:
                    data = readFromQueue(queue)
                    if start < launch:
                        pag.sleep(10)
                        start += 10
                        if start == launch:
                            print(data, ": Timed out after", launch, "Seconds")
                            returnhome()
                            break

                if 'Emails' in data:
                    delay()
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'Yes' in data:
                        print("Factory 1: Factory Reset Settings \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I114'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Tools 11: Failed to Factory Reset settings")
                        print("Factory 1: Factory Reset Settings \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I114'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

            else:
                logging.critical("Tools 11: Failed to locate factory reset")
                print("Factory 1: Failed to locate factory reset")
                print(data)
                ws['I114'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 11: Failed to locate back up and restore")
            print("Factory 1: Failed to locate back up and restore")
            print(data)
            ws['I114'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 11: Failed to log into email account")
        print("Factory 1: Failed to log into email account")
        print(data)
        ws['I114'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

def file_1(queue):  # STATUS: TESTED
    # Locate a DOCX file from Test data and opens it
    try:
        testdata()

    except OSError as test_data:
        print("Error: Files already Exist")

    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("x", interval=1)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_1: Can you navigate File explorer to find a file? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I90'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 1: Failed to find PDF file")
                    print('File_1: Can you navigate File explorer to find a file? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I90'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 1: Failed to find PDF")
                print("File_1: Failed to find PDF.pdf")
                print(data)
                ws['I90'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 1: Failed to locate test data folder")
            print("File_1: Failed to locate Test data folder")
            print(data)
            ws['I90'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 1: Failed to locate file explorer")
        print("File_1: Failed to locate File explorer")
        print(data)
        ws['I90'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_2(queue):   # STATUS: TESTED
    # Can you open an RTF
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'RTF' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_2: Can you open an RTF with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I91'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 2: Failed to open RTF with file explorer")
                    print('File_2: Can you open an RTF with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I91'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 2: Failed to find RTF")
                print("File_2: Failed to find RTF.rtf")
                print(data)
                ws['I91'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 2: Failed to locate test data folder")
            print("File_2: Failed to locate Test data folder")
            print(data)
            ws['I91'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 2: Failed to locate file explorer")
        print("File_2: Failed to locate File explorer")
        print(data)
        ws['I91'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_3(queue):   # STATUS: TESTED
    # Opens TXT file from file explorer
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)
        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)
            if 'TXT' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_3: Can you open an TXT with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I92'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 3: Failed to open a TXT with file explorer")
                    print('File_3: Can you open an TXT with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I92'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 3: Failed to find TXT")
                print("File_3: Failed to find TXT.txt")
                print(data)
                ws['I92'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 3: Failed to locate test data folder")
            print("File_3: Failed to locate Test data folder")
            print(data)
            ws['I92'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 3: Failed to locate file explorer")
        print("File_3: Failed to locate File explorer")
        print(data)
        ws['I92'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_4(queue):    # STATUS: TESTED
    # Opens JPG
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("j", interval=1)
            data = readFromQueue(queue)
            if 'JPEG' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                    print('File_4: Can you open an JPG with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I93'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 4: Failed to open a JPG with file explorer")
                    print('File_4: Can you open an JPG with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I93'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 4: Failed to find JPG")
                print("File_4: Failed to find JPG.jpg")
                print(data)
                ws['I93'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 4: Failed to locate test data folder")
            print("File_4: Failed to locate Test data folder")
            print(data)
            ws['I93'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 4: Failed to locate file explorer")
        print("File_4: Failed to locate File explorer")
        print(data)
        ws['I93'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_5(queue):   # STATUS: TESTED
    # Opens PNG
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("p", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'PNG' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                    print('File_5: Can you open an PNG with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I94'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 5: Failed to open PNG with file explorer")
                    print('File_5: Can you open an PNG with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I94'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 5: Failed to find PNG")
                print("File_5: Failed to find PNG.png")
                print(data)
                ws['I94'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 5: Failed to locate test data folder")
            print("File_5: Failed to locate Test data folder")
            print(data)
            ws['I94'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 5: Failed to locate file explorer")
        print("File_5: Failed to locate File explorer")
        print(data)
        ws['I94'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_6(queue):  # STATUS: TESTED
    # Opens DOC file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("d", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'DOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_6: Can you open an DOC with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I95'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 6: Failed to open DOC file with file explorer")
                    print('File_6: Can you open an DOC with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I95'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 6: Failed to find DOC ")
                print("File_6: Failed to find DOC.doc")
                print(data)
                ws['I95'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 6: Failed to locate test data folder")
            print("File_6: Failed to locate Test data folder")
            print(data)
            ws['I95'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 6: Failed to locate file explorer")
        print("File_6: Failed to locate File explorer")
        print(data)
        ws['I95'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_7(queue):  # STATUS: TESTED
    # Opens XDOC
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("x", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_7: Can you open an DOCX with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I96'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 7: Failed to open DOCX with file explorer")
                    print('File_7: Can you open an DOCX with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I96'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 7: Failed to find DOCX")
                print("File_7: Failed to find DOCX.docx")
                print(data)
                ws['I96'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 7: Failed to locate test data folder")
            print("File_7: Failed to locate Test data folder")
            print(data)
            ws['I96'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 7: Failed to locate file explorer")
        print("File_7: Failed to locate File explorer")
        print(data)
        ws['I96'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_8(queue):  # STATUS: TESTED
    # Move a selected file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("t", interval=1)
    pag.press("enter", interval=1)
    pag.press("d", interval=1)
    pag.press("f2", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'Move Selected File' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Move Selected File' in data:
        pag.press("enter", interval=1)
        pag.press("g", interval=1)
        pag.press("enter", interval=1)
        pag.press("g", interval=1)
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("esc", presses=2, interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)
        if 'DOC' in data:
            pag.sleep(2)

            print("File_8: Can you move a file with file explorer \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I97'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("File 8: Failed to move DOC file with file explorer")
            print("File_8: Can you move a file with file explorer \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I97'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 8: Failed to navigate to Move file")
        print("File_8: Navigation was not present on Move File \n"
              "Current String:", data)
        ws['I97'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_9(queue):  # STATUS: TESTED
    # can you rename a file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)

        if 'JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'Rename Selected File' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Rename Selected File' in data:
                pag.press("enter", interval=1)
                pag.write("rename", interval=0.5)
                pag.press("enter", interval=1, presses=2)
                pag.press("r", interval=1)
                data = readFromQueue(queue)
                if 'renameJPEG' in data:
                    print("File_9: Can you rename a file from file explorer? \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I98'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 9: Failed to rename a file ")
                    print("File_9: Can you rename a file from file explorer? \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I98'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 9: Failed to locate rename file")
                print("File_9: Failed to locate rename file \n"
                      "Current String:", data)
                ws['I98'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 9: Failed to find JPEG to rename")
            print("File_9: Failed to find JPEG to rename \n"
                  "Current String:", data)
            ws['I98'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 9: Failed to enter file explorer")
        print("File_9: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I98'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_10(queue):  # STATUS: TESTED
    # Copy a file with file explorer
    delay()
    start = 0
    x = os.environ['USERPROFILE'] + "\\Documents\\GuideConnect\\Music\\DOC.doc"
    y = os.environ['USERPROFILE'] + "\\Documents\\GuideConnect\\Documents\\testdata\\DOC.doc"
    try:
        shutil.copy(x, y)
    except:
        pass

    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        if 'DOC' in data:
            pag.press("f2", interval=1)
            pag.press("c", interval=1, presses=2)
            data = readFromQueue(queue)

            while 'Copy Selected File' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Copy Selected File' in data:
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                pag.press("enter", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'Music' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Music' in data:
                    pag.press("enter", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.press("esc", presses=2, interval=1)
                    pag.press("m", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("d", interval=1)
                    data = readFromQueue(queue)
                    if 'DOC' in data:
                        print("File_10: Can you copy a file from file explorer? \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I99'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 10: Failed to copy a file from file explorer")
                        print("File_10: Can you copy a file from file explorer? \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I99'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 10: Failed to find music folder")
                    print("File_10: Failed to find music folder \n"
                          "Current String:", data)
                    ws['I99'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
        else:
            logging.critical("File 10: Failed to find DOC to copy")
            print("File_10: Failed to find DOC to copy \n"
                  "Current String:", data)
            ws['I99'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("File 10: Failed to enter file explorer")
        print("File_10: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I99'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_11(queue):     # STATUS: TESTED
    # delete multiple files and locate them in deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        data = readFromQueue(queue)
        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            data = readFromQueue(queue)
            if 'Delete Selected Items' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("esc", presses=2, interval=1)
                pag.press("d", presses=2, interval=1)
                data = readFromQueue(queue)

                if 'Deleted items' in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'DOC.doc' in data:
                        pag.press("right", interval=1)
                        data = readFromQueue(queue)
                        if 'PDF.pdf' in data:
                            pag.press("right", interval=1)
                            data = readFromQueue(queue)
                            if 'PNG.png' in data:
                                pag.press("right", interval=1)
                                data = readFromQueue(queue)
                                if 'renameJPEG.jpg' in data:
                                    pag.press("right", interval=1)
                                    data = readFromQueue(queue)
                                    if 'RTF.rtf' in data:
                                        pag.press("right", interval=1)
                                        data = readFromQueue(queue)
                                        if 'TXT.txt' in data:
                                            pag.press("right", interval=1)
                                            data = readFromQueue(queue)
                                            if 'XDOC.docx' in data:
                                                print("File_11: Can you view multiple deleted items in deleted items \n"
                                                      "Result: PASS \n"
                                                      "Current String:", data)
                                                ws['I100'].value = 'PASS'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()

                                            else:
                                                logging.critical("File 11: Failed to locate all files in deleted items")
                                                print("File_11: Can you view multiple deleted items in deleted items \n"
                                                      "Result: FAIL \n"
                                                      "Current String:", data)
                                                ws['I100'].value = 'FAIL'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()

                                        else:
                                            logging.critical("File 11: Failed to find TXT in deleted items")
                                            print("File_11: failed to find TXT in deleted items \n"
                                                  "Current String:", data)
                                            ws['I100'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            returnhome()
                                    else:
                                        logging.critical("File 11: Failed to find RTF in deleted items")
                                        print("File_11: failed to find RTF in deleted items \n"
                                              "Current String:", data)
                                        ws['I100'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()
                                else:
                                    logging.critical("File 11: Failed to find JPG in deleted items")
                                    print("File_11: failed to find JPG in deleted items \n"
                                          "Current String:", data)
                                    ws['I100'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()
                            else:
                                logging.critical("File 11: Failed to find PNG in deleted items")
                                print("File_11: failed to find PNG in deleted items \n"
                                      "Current String:", data)
                                ws['I100'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()
                        else:
                            logging.critical("File 11: Failed to find PDF in deleted items")
                            print("File_11: failed to find PDF in deleted items \n"
                                  "Current String:", data)
                            ws['I100'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()
                    else:
                        logging.critical("File 11: Failed to find DOC in deleted items")
                        print("File_11: failed to find DOC in deleted items \n"
                              "Current String:", data)
                        ws['I100'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("File 11: Failed to find deleted items")
                    print("File_11: failed to find deleted items \n"
                          "Current String:", data)
                    ws['I100'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("File 11: Failed to locate deleted selected items")
                print("File_11: failed to locate deleted selected items \n"
                      "Current String:", data)
                ws['I100'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("File 11: Failed to find test data folder")
            print("File_11: failed to find test data folder \n"
                  "Current String:", data)
            ws['I100'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("File 11: Failed to enter file explorer")
        print("File_11: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I100'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_12(queue):     # STATUS: TESTED
    # restore multiple files from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'DOC.doc' in data:
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            if 'PDF.pdf' in data:
                pag.press(" ", interval=1)
                pag.press("f2", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)
                if 'Move selected items' in data:
                    pag.press("enter", interval=1)
                    pag.press("a", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)
                    if 'PNG.png' in data:
                        print("File_12: Restore multiple files from deleted items to attachments \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I101'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 12: Failed to restore DOC and PDF files from deleted items")
                        print("File_12: Restore multiple files from deleted items to attachments \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I101'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 12: Failed to locate move selected items")
                    print("File_12: Failed to locate move selected items \n"
                          "Current String:", data)
                    ws['I101'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 12: Failed to find PDF in deleted items")
                print("File_12: Failed to find PDF in deleted items \n"
                      "Current String:", data)
                ws['I101'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 12: Failed to find DOC in deleted items")
            print("File_12: Failed to find DOC in deleted items \n"
                  "Current String:", data)
            ws['I101'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 12: Failed to enter file explorer")
        print("File_12: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I101'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_13(queue):     # STATUS: TESTED
    # Delete multiple files from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'PNG.png' in data:
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            if 'renameJPEG.jpg' in data:
                pag.press(" ", interval=1)
                pag.press("f2", interval=1)
                pag.press("r", interval=1)
                data = readFromQueue(queue)

                if 'Remove selected items' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)
                    if 'RTF.rtf' in data:
                        print("File_13: Delete multipe files from deleted items \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I102'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 13: Failed to delete PNG and JPG from deleted items")
                        print("File_13: Delete multipe files from deleted items \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I102'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 13: Failed to locate remove selected items")
                    print("File_13: Failed to locate remove selected items \n"
                          "Current String:", data)
                    ws['I102'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 13: Failed to find JPG in deleted items")
                print("File_13: Failed to find JPG in deleted items \n"
                      "Current String:", data)
                ws['I102'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 13: Failed to find PNG in deleted items")
            print("File_13: Failed to find PNG in deleted items \n"
                  "Current String:", data)
            ws['I102'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 13: Failed to enter file explorer")
        print("File_13: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I102'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_14(queue):     # STATUS: TESTED
    # Remove all items from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'RTF.rtf' in data:
            pag.press("f2", interval=1)
            pag.press("r", presses=2, interval=1)
            data = readFromQueue(queue)
            if 'Remove all' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'No items found' in data:
                    print("File_14: Remove all items from deleted items \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I103'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 14: Failed to remove all from deleted items")
                    print("File_14: Remove all items from deleted items \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I103'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 14: Failed to locate remove all")
                print("File_14: Failed to locate remove all \n"
                      "Current String:", data)
                ws['I103'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 14: Failed to find RTF in deleted items")
            print("File_14: Failed to find RTF in deleted items \n"
                  "Current String:", data)
            ws['I103'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 14: Failed to enter file explorer")
        print("File_14: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I103'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

                                                # MODULE: TOOLS END
#######################################################################################################################

