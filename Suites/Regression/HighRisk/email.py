import logging
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120


def email_send1(queue):     # STATUS: TESTED
    # Logs In And Sends An Email
    delay()
    start = 0
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_SETTINGS_EMAILADDRESS' in data:
        pag.write("guideautomation1@outlook.com", interval=0.5)
        pag.press("enter", interval=1)
        pag.write("testing", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.write("guideconnect1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(15)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", presses=3, interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Automation", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.write("test email", interval=0.5)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SENDING' in data:
                    data = readFromQueue(queue)

                if 'OK' in data:
                    pag.press("enter", interval=1)
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                elif 'GMENU_EMAIL_SENT_SUCCESS' in data:
                    pag.press("enter", interval=1)
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Emails 1: Failed to compose and send an email")
                    print('Emails 1 : Sign in and send an email \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I156'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 1: Failed to compose a new email")
                print("Emails 1: Failed to compose new email")
                print(data)
                ws['I156'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 1: Failed to sign into email address")
            print("Emails 1: failed to sign into email address")
            print(data)
            ws['I156'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 1: Failed to add email address")
        print("Emails 1: failed to add email address")
        print(data)
        ws['I156'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send2(queue):     # STATUS: TESTED
    # Checks sent folder for 'Sent' email
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    pag.press("s", interval=1)
    data = readFromQueue(queue)

    while 'Sent' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Sent' in data:
        pag.press("enter", interval=1)

        while 'guideautomation2@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation2@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.sleep(5)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)

            print('Emails 2: Check for the sent email via the Sent folder \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I157'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails 2: Failed to find sent email in Sent folder")
            print('Emails 2: Check for the sent email via the Sent folder \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I157'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 2: Failed to enter the sent folder")
        print("Emails 2: failed to enter the sent folder")
        print(data)
        ws['I157'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send3(queue):     # STATUS: TESTED
    # Log into an addition email account
    delay()
    start = 0
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'Email' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Email' in data:
        pag.press("enter", interval=1)
        pag.press("a", interval=1)
        data = readFromQueue(queue)

        if 'Account manager' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("testing", interval=0.5)
            pag.press("enter", presses=3, interval=1)
            pag.write("guideconnect2", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            pag.press("enter", interval=1)

            while 'Write new email' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Write new email' in data:
                print('Emails 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I158'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Emails 3: Failed to add a secondary email account via account manager")
                print('Emails 3: Can you use account manager to log into multiple email accounts \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I158'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 3: Failed to enter account manager")
            print("Emails 3: failed to enter account manager")
            print(data)
            ws['I158'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 3: Failed to navigate to email settings")
        print("Emails 3: failed to navigate to email settings")
        print(data)
        ws['I158'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send4(queue):     # STATUS: TESTED
    # Can you open a received email?
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        pag.sleep(5)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("enter")
            pag.sleep(10)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.sleep(5)

            print('Emails 4: Can you view a received email? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I159'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails 4: Failed to view and delete received email")
            print('Emails 4: Can you view a received email? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I159'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 4: Failed to locate guideautomation2@outlook.com")
        print("Emails 4: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I159'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send5(queue):     # STATUS: TESTED
    # Send a new email via address book
    delay()
    start = 0
    pag.press("enter", presses=3, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Select from address book' not  in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Select from address book' in data:
        pag.press("enter", interval=1)
        pag.press("e", interval=1)
        data = readFromQueue(queue)

        while 'Email Testing' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Email Testing' in data:
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.write("test", interval=0.5)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SENDING' in data:
                    data = readFromQueue(queue)
                    pag.sleep(3)

                if 'OK' in data:
                    returnhome()
                    pag.press("enter", interval=1)
                    pag.press("g", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation2@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation2@outlook.com' in data:
                        pag.press("enter", interval=1)
                        pag.press("i", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'GMENU_EMAIL_VIEW_MESSAGE' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_EMAIL_VIEW_MESSAGE' in data:
                                pag.sleep(10)
                                print('Emails 5: Send and view an email via Address book contact \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I160'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Emails 5: Failed to send and view and email via address contact")
                                print('Emails 5: Send and view an email via Address book contact \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I160'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Emails 5: Failed to find email sent via address book")
                            print("Emails 5: failed to find email sent via address book")
                            print(data)
                            ws['I160'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Emails 5: Failed to locate guideautomation2@outlook.com")
                        print("Emails 5: failed to locate guideautomation2@outlook.com")
                        print(data)
                        ws['I160'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Emails 5: Failed to send email")
                    print("Emails 5: failed to send email")
                    print(data)
                    ws['I160'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 5: Failed to open body of a new email")
                print("Emails 5: failed to open new message body")
                print(data)
                ws['I160'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails 5: Failed to locate contact (Email Testing")
            print("Emails 5: failed to locate contact - Email Testing")
            print(data)
            ws['I160'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 5: Failed to locate select from address book")
        print("Emails 5: failed to locate select from address book")
        print(data)
        ws['I160'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send6(queue):     # STATUS: TESTED
    # add and remove recipients from both email address and contacts
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", presses=2, interval=2)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add or view recipients' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add or view recipients' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'No recipients' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'No recipients' in data:
                    pag.press("f2", interval=1)
                    pag.press("enter", presses=3, interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SEND_ACTION_RECIPIENTS_ACTION_NEW_ENTER' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_EMAIL_SEND_ACTION_RECIPIENTS_ACTION_NEW_ENTER' in data:
                        pag.write("test", interval=0.5)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while '(To) test' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if '(To) test' in data:
                            pag.press("f2", interval=1)
                            pag.press("enter", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", presses=2, interval=1)
                            data = readFromQueue(queue)

                            while 'No recipients' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'No recipients' in data:
                                pag.press("f2", interval=1)
                                pag.press("enter", presses=2, interval=1)
                                pag.press("s", interval=1)
                                pag.press("enter", interval=1)
                                pag.press("e", interval=1)
                                pag.press("enter", interval=1)
                                data = readFromQueue(queue)

                                while '(To) guideautomation2@outlook.com' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if '(To) guideautomation2@outlook.com' in data:
                                    print('Emails 6: Add and remove recipients to email from contact/email address\n'
                                          '>>> Result: PASS \n'
                                          '>>> Current String:', data)
                                    ws['I161'].value = 'PASS'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()

                                else:
                                    logging.critical("Emails 6: Failed to add/remove recipients from contact/email")
                                    print('Emails 6: Add and remove recipients to email from contact/email address\n'
                                          '>>> Result: FAIL \n'
                                          '>>> Current String:', data)
                                    ws['I161'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()

                            else:
                                logging.critical("Emails 6: Failed to remove a recipient")
                                print("Emails 6: failed to remove a recipient")
                                print(data)
                                ws['I161'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                pag.press("esc", presses=2, interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", interval=1)
                                returnhome()

                        else:
                            logging.critical("Emails 6: Failed to locate test(Added via email) in recipients")
                            print("Emails 6: test not present in recipients")
                            print(data)
                            ws['I161'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", presses=2, interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Emails 6: Failed to add recipient via email address")
                        print("Emails 6: failed to add recipient via email address")
                        print(data)
                        ws['I161'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=3, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:

                    logging.critical("Emails 6: Unexpected Recipient was present")
                    print("Emails 6: unexpected recipient was present")
                    print(data)
                    ws['I161'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:

                logging.critical("Emails 6: Failed to locate add or view recipients")
                print("Emails 6: failed to locate add or view recipients")
                print(data)
                ws['I161'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:

            logging.critical("Emails 6: Failed to enter body of the new email")
            print("Emails 6: failed to enter body of the new email")
            print(data)
            ws['I161'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 6: Failed to open recipient address")
        print("Emails 6: failed to add recipient address")
        print(data)
        ws['I161'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send7(queue):     # STATUS: TESTED
    # Change email subject line
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_SUBJECT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_SUBJECT' in data:
            pag.write("testing", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("c", interval=1)
            data = readFromQueue(queue)

            while 'Change subject' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Change subject' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_SEND_ACTION_CHANGE_SUBJECT' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_SEND_ACTION_CHANGE_SUBJECT' in data:
                    pag.write("changed", interval=0.5)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                        print('Emails 7: Change the email Subject \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I162'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 7: Failed to change the email subject line")
                        print('Emails 7: Change the email Subject \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I162'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 7: Failed to open change subject line")
                    print("Emails 7: failed to open change subject")
                    print(data)
                    ws['I162'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 7: Failed to locate change subject line")
                print("Emails 7: failed to locate change subject")
                print(data)
                ws['I162'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 7: Failed to add a new subject line")
            print("Emails 7: failed to add a new subject")
            print(data)
            ws['I162'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 7: Failed to open recipient address")
        print("Emails 7: failed to add recipient address")
        print(data)
        ws['I162'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send8(queue):     # STATUS: TESTED
    # Complete a spell check inside of an email
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
            pag.write("dis iz a twest", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)

            while 'Spell Check' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Spell Check' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_NEW_MESSAGE_SPELLCHECK' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE_SPELLCHECK' in data:
                    pag.press("d", presses=3, interval=1)
                    pag.press("enter", interval=1)
                    pag.press("t", interval=1)
                    data = readFromQueue(queue)

                    while 'test' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'test' in data:
                        pag.press("enter", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                            print('Emails 8: Spell check an email  \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I163'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                        else:
                            logging.critical("Emails 8: Failed to spell check an email")
                            print('Emails 8: Spell check an email  \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I163'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Emails 8: Failed to change twest to test")
                        print("Emails 8: failed to spell check twest to test")
                        print(data)
                        ws['I163'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 8: Failed to open spell checker")
                    print("Emails 8: failed to open spell checker")
                    print(data)
                    ws['I163'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 8: Failed to locate spell checker")
                print("Emails 8: failed to locate spell check")
                print(data)
                ws['I163'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 8: Failed to enter body of the email via send to address contact")
            print("Emails 8: failed to enter body of email via send to address book contact")
            print(data)
            ws['I163'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 8: Failed to open recipient address")
        print("Emails 8: failed to add recipient address")
        print(data)
        ws['I163'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_send9(queue):  # STATUS: TESTED
    # Save an email to drafts via F2
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("F2save2drafts@drafts.com", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("save to drafts", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Save to Drafts' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Save to Drafts' in data:
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.press("f", interval=1)
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                data = readFromQueue(queue)

                if 'Drafts' in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'F2save2drafts@drafts.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'F2save2drafts@drafts.com' in data:
                        print('Emails 9: Save a draft via F2 actions menu  \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I164'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("f2", interval=1)
                        pag.press("d", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", presses=2, interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 9: Failed to save draft via F2 actions menu")
                        print('Emails 9: Save a draft via F2 actions menu  \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I164'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Emails 9: Failed to locate drafts folder")
                    print("Emails 9: failed to locate drafts folder")
                    print(data)
                    ws['I164'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Emails 9: Failed to save draft and return to body of the email")
                print("Emails 9: failed to save to drafts and return to body of email")
                print(data)
                ws['I164'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 9: Failed to locate save to drafts")
            print("Emails 9: failed to locate save to drafts")
            print(data)
            ws['I164'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 9: Failed to open recipient address")
        print("Emails 9: failed to add recipient address")
        print(data)
        ws['I164'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send10(queue):        # STATUS: TESTED
    # Save email to draft via ESC
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("ESCsave2drafts@drafts.com", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("draft email via ESC", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'ESCsave2drafts@drafts.com'not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'ESCsave2drafts@drafts.com' in data:
            print('Emails 10: Save a draft via ESC method  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I165'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            returnhome()

        else:
            logging.critical("Emails 10: Failed to save a draft via ESC")
            print('Emails 10: Save a draft via ESC method  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I165'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails 10: Failed to open recipient address")
        print("Emails 10: failed to add recipient address")
        print(data)
        ws['I165'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send11(queue):        # STATUS: TESTED
    # Adds and removes an attachment from an email
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("guideautomation2@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("attachments", interval=0.5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("test", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("a", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Add or view attachments' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Add or view attachments' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'No Attachments' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No Attachments' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("f2", interval=1)
                    pag.press("r", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)

                    while 'No Attachments' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'No Attachments' in data:
                        print('Emails 11: Adds and removes an attachment from an email  \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I166'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                    else:
                        logging.critical("Emails 11: Failed to add and remove an attachment from an email")
                        print('Emails 11: Adds and removes an attachment from an email  \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I166'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Emails 11: Failed to add attachment to an email")
                    print("Emails 11: failed to add attachment to an email")
                    print(data)
                    ws['I166'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Emails 11: No Attachments Prompt is missing")
                print("Emails 11: Expected No Attachments prompt is missing")
                print(data)
                ws['I166'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=2, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 11: Failed to locate add or view attachments")
            print("Emails 11: failed to locate add or view attachments")
            print(data)
            ws['I166'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 11: Failed to open recipient address")
        print("Emails 11: failed to add recipient address")
        print(data)
        ws['I166'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        returnhome()


def email_send12(queue):    # STATUS: TESTED
    # Add all test data attachments to an email and check its received correctly
    delay()
    start = 0
    pag.press("enter", presses=4, interval=1)
    pag.write("guideautomation2@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("attachments", interval=0.5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_MESSAGE' in data:
        pag.write("test", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("a", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Add or view attachments' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Add or view attachments' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)

            while 'testdata' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'testdata' in data:
                pag.press("enter", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("j", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("p", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("p", presses=2, interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("r", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.press("t", interval=1)
                pag.press("enter", interval=1)
                pag.press("x", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("right", interval=1)
                    data = readFromQueue(queue)

                    while 'JPEG.jpg' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'JPEG.jpg' in data:
                        pag.press("right", interval=1)
                        data = readFromQueue(queue)

                        while 'PDF.pdf' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'PDF.pdf' in data:
                            pag.press("right", interval=1)
                            data = readFromQueue(queue)

                            while 'PNG.png' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'PNG.png' in data:
                                pag.press("right", interval=1)
                                data = readFromQueue(queue)

                                while 'RTF.rtf' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if 'RTF.rtf' in data:
                                    pag.press("right", interval=1)
                                    data = readFromQueue(queue)

                                    while 'TXT.txt' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'TXT.txt' in data:
                                        pag.press("right", interval=1)
                                        data = readFromQueue(queue)

                                        while 'XDOC.docx' not in data:
                                            data = readFromQueue(queue)
                                            if start < timeout:
                                                pag.sleep(10)
                                                start += 10
                                                if start == timeout:
                                                    print(data, ": Timed out after", timeout, "Seconds")
                                                    returnhome()
                                                    break

                                        if 'XDOC.docx' in data:
                                            pag.press("esc", interval=1)
                                            pag.press("f2", interval=1)
                                            pag.press("enter", interval=1)
                                            data = readFromQueue(queue)

                                            while 'GMENU_EMAIL_SENDING' in data:
                                                data = readFromQueue(queue)
                                                pag.sleep(2)

                                            if 'OK' in data:
                                                pag.press("enter", interval=1)
                                                pag.sleep(2)
                                                pag.press("esc", interval=1)
                                                pag.press("g", interval=1)
                                                data = readFromQueue(queue)

                                                while 'guideautomation2@outlook.com' not in data:
                                                    data = readFromQueue(queue)
                                                    if start < timeout:
                                                        pag.sleep(10)
                                                        start += 10
                                                        if start == timeout:
                                                            print(data, ": Timed out after", timeout, "Seconds")
                                                            returnhome()
                                                            break

                                                if 'guideautomation2@outlook.com' in data:
                                                    pag.press("enter", interval=1)
                                                    pag.press("i", interval=1)
                                                    pag.sleep(60)
                                                    pag.press("enter", interval=1)
                                                    data = readFromQueue(queue)

                                                    while 'guideautomation1@outlook.com' not in data:
                                                        data = readFromQueue(queue)
                                                        if start < timeout:
                                                            pag.sleep(10)
                                                            start += 10
                                                            if start == timeout:
                                                                print(data, ": Timed out after", timeout, "Seconds")
                                                                returnhome()
                                                                break

                                                    if 'guideautomation1@outlook.com' in data:
                                                        pag.press("f2", interval=1)
                                                        pag.press("v", interval=1)
                                                        pag.press("enter", interval=1)
                                                        data = readFromQueue(queue)

                                                        while 'DOC.doc' not in data:
                                                            data = readFromQueue(queue)
                                                            if start < timeout:
                                                                pag.sleep(10)
                                                                start += 10
                                                                if start == timeout:
                                                                    print(data, ": Timed out after", timeout, "Seconds")
                                                                    returnhome()
                                                                    break

                                                        if 'DOC.doc' in data:  # might need a condition for Archive
                                                            pag.press("esc", interval=1)
                                                            pag.press("f2", interval=1)
                                                            pag.press("m", interval=1)
                                                            pag.press("enter", presses=3, interval=1)
                                                            pag.sleep(5)
                                                            print('Emails 12: Sends all Attachments Types via Email  \n'
                                                                  '>>> Result: PASS \n'
                                                                  '>>> Current String:', data)
                                                            ws['I167'].value = 'PASS'
                                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                                            returnhome()

                                                        else:
                                                            logging.critical(
                                                                "Emails 12: Failed to send attachments via email")
                                                            print('Emails 12: Sends all Attachments Types via Email  \n'
                                                                  '>>> Result: FAIL \n'
                                                                  '>>> Current String:', data)
                                                            ws['I167'].value = 'FAIL'
                                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                                            returnhome()
                                                    else:
                                                        logging.critical(
                                                            "Emails 12: Failed to find email from guide automation 1"
                                                                        )
                                                        print("Emails 12: failed to find email"
                                                              " from guideautomation1@outlook.com")
                                                        print(data)
                                                        ws['I167'].value = 'FAIL'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()

                                                else:
                                                    logging.critical("Emails 12: Failed to locate "
                                                                     "guide automation 2")
                                                    print("Emails 12: failed to locate guideautomation2@outlook.com")
                                                    print(data)
                                                    ws['I167'].value = 'FAIL'
                                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                                    returnhome()
                                        else:
                                            logging.critical("Emails 12: Failed to locate attached DOCX")
                                            print("Emails 12: failed to locate DOCX")
                                            print(data)
                                            ws['I167'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            pag.press("esc", presses=2, interval=1)
                                            pag.press("right", interval=1)
                                            pag.press("enter", interval=1)
                                            returnhome()

                                    else:
                                        logging.critical("Emails 12: Failed to locate attached TXT")
                                        print("Emails 12: failed to locate TXT")
                                        print(data)
                                        ws['I167'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        pag.press("esc", presses=2, interval=1)
                                        pag.press("right", interval=1)
                                        pag.press("enter", interval=1)
                                        returnhome()

                                else:
                                    logging.critical("Emails 12: Failed to locate attached RTF")
                                    print("Emails 12: failed to locate RTF")
                                    print(data)
                                    ws['I167'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    pag.press("esc", presses=2, interval=1)
                                    pag.press("right", interval=1)
                                    pag.press("enter", interval=1)
                                    returnhome()
                            else:
                                logging.critical("Emails 12: Failed to locate attached PNG")
                                print("Emails 12: failed to locate PNG")
                                print(data)
                                ws['I167'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                pag.press("esc", presses=2, interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", interval=1)
                                returnhome()
                        else:
                            logging.critical("Emails 12: Failed to locate attached PDF")
                            print("Emails 12: failed to locate PDF")
                            print(data)
                            ws['I167'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("esc", presses=2, interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", interval=1)
                            returnhome()
                    else:
                        logging.critical("Emails 12: Failed to locate attached JPG")
                        print("Emails 12: failed to locate JPEG")
                        print(data)
                        ws['I167'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", presses=2, interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()
                else:
                    logging.critical("Emails 12: Failed to locate attached DOC")
                    print("Emails 12: failed to locate DOC")
                    print(data)
                    ws['I167'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()
            else:
                logging.critical("Emails 12: Failed to locate test data folder")
                print("Emails 12: failed to locate test data folder")
                print(data)
                ws['I167'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", presses=3, interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Emails 12: Failed to locate add or view attachments")
            print("Emails 12: failed to locate add or view attachments")
            print(data)
            ws['I167'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

    else:
        logging.critical("Emails 12: Failed to open recipient address")
        print("Emails 12: failed to add recipient address")
        print(data)
        ws['I167'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        # wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive1(queue):      # STATUS: TESTED
    # Manually check for new emails
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        pag.press("f2", interval=2)
        pag.sleep(5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        while 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK' in data:
            data = readFromQueue(queue)

        if 'OK' in data:
            pag.sleep(5)
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'GMENU_EMAIL_ACTIONS_NEWMAILCHECK_COMPLETE' in data:
            pag.sleep(5)
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 1: Failed to manually check emails")
            print('Email Receive 1: Manually check for emails \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I168'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 1: Failed to locate guideautomation2@outlook.com")
        print("Email Receive 1: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I168'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive2(queue):      # STATUS: TESTED
    # reply to an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            if 'Reply' in data:
                pag.press("enter", interval=1)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("reply email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 2: Reply to an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I169'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 2: Failed to reply to an email")
                                print('Email Receive 2: Reply to an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I169'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 2: Failed to locate guide automation1")
                            print("Email Receive 2: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I169'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 2: Failed to send an email")
                        print("Email Receive 2: failed to send email")
                        print(data)
                        ws['I169'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Receive 2: Failed to enter body of an email")
                    print("Email Receive 2: failed to enter body of email")
                    print(data)
                    ws['I169'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 2: Failed to locate reply in actions menu")
                print("Email Receive 2: failed to locate reply in actions menu")
                print(data)
                ws['I169'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 2: Failed to locate guide automation 1 in inbox")
            print("Email Receive 2: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I169'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 2: Failed to locate guide automation 2")
        print("Email Receive 2: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I169'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive3(queue):  # STATUS: TESTED
    # reply to all recipients in email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("r", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'Reply All' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Reply All' in data:
                pag.press("enter", interval=1)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("reply all email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 3: Reply All to an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I170'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 3: Failed to reply All to an email")
                                print('Email Receive 3: Reply All to an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I170'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 3: Failed to locate guide automation1")
                            print("Email Receive 3: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I170'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 3: Failed to send an email")
                        print("Email Receive 3: failed to send email")
                        print(data)
                        ws['I170'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Receive 3: Failed to enter body of an email")
                    print("Email Receive 3: failed to enter body of email")
                    print(data)
                    ws['I170'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 3: Failed to locate reply all in actions menu")
                print("Email Receive 3: failed to locate reply all in actions menu")
                print(data)
                ws['I170'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Email Receive 3: Failed to locate guide automation 1 in inbox")
            print("Email Receive 3: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I170'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 3: Failed to locate guide automation 2")
        print("Email Receive 3: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I170'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive4(queue):  # STATUS: TESTED
    # forward an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("f", interval=1)
            data = readFromQueue(queue)

            while 'Forward' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Forward' in data:
                pag.press("enter", presses=2, interval=1)
                pag.sleep(10)
                pag.write("guideautomation1@outlook.com", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)

                while 'GMENU_EMAIL_NEW_MESSAGE' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_EMAIL_NEW_MESSAGE' in data:
                    pag.write("forward email", interval=0.5)
                    pag.press("f2", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_EMAIL_SENDING' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        pag.press("esc", presses=2, interval=1)
                        data = readFromQueue(queue)

                        while 'guideautomation1@outlook.com' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'guideautomation1@outlook.com' in data:
                            pag.press("enter", interval=1)
                            pag.press("i", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'guideautomation2@outlook.com' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'guideautomation2@outlook.com' in data:
                                pag.press("Del", interval=1)
                                pag.press("right", interval=1)
                                pag.press("enter", presses=2, interval=1)

                                print('Email Receive 4: Forward an email \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I171'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Email Receive 4: Failed to forward an email")
                                print('Email Receive 4: Forward an email \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I171'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Email Receive 4: Failed to locate guide automation1")
                            print("Email Receive 4: failed to locate guideautomation1@outlook.com")
                            print(data)
                            ws['I171'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Receive 4: Failed to send an email")
                        print("Email Receive 4: failed to send email")
                        print(data)
                        ws['I171'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("esc", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", interval=1)
                        returnhome()

                else:
                    logging.critical("Email Receive 4: Failed to enter body of an email")
                    print("Email Receive 4: failed to enter body of email")
                    print(data)
                    ws['I171'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

            else:
                logging.critical("Email Receive 4: Failed to locate forward in actions menu")
                print("Email Receive 4: failed to locate forward in actions menu")
                print(data)
                ws['I171'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Email Receive 4: Failed to locate guide automation 1 in inbox")
            print("Email Receive 4: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I171'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 4: Failed to locate guide automation 2")
        print("Email Receive 4: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I171'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive5(queue):  # STATUS: TESTED
    # Add email contact to address book
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add sender to address book' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add sender to address book' in data:
                pag.press("enter", interval=1)
                pag.write("email", interval=0.5)
                pag.press("enter", interval=1)
                pag.write("contact", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                returnhome()
                pag.press("a", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("e", interval=1)
                data = readFromQueue(queue)

                while 'email contact' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'email contact' in data:
                    print('Email Receive 5: Add email contact to address book \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I172'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Receive 5: Failed to add a email contact to address book")
                    print('Email Receive 5: Add email contact to address book \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I172'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 5: Failed to locate Add contact in actions menu")
                print("Email Receive 5: failed to locate Add contact in actions menu")
                print(data)
                ws['I172'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 5: Failed to locate guide automation 1 in inbox")
            print("Email Receive 5: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I172'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 5: Failed to locate guide automation 2")
        print("Email Receive 5: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I172'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive6(queue):  # STATUS: TESTED
    # Move email to a different folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=2)
        pag.press("i", interval=2)
        pag.press("enter", interval=2)
        data = readFromQueue(queue)

        while 'guideautomation1@outlook.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'guideautomation1@outlook.com' in data:
            pag.press("f2", interval=2)
            pag.press("m", interval=2)
            data = readFromQueue(queue)

            while 'Move to folder' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Move to folder' in data:
                pag.press("enter", interval=1)
                pag.press("n", interval=1)
                data = readFromQueue(queue)

                while 'Notes' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Notes' in data:
                    pag.press("enter", presses=2, interval=1)
                    pag.sleep(3)
                    pag.press("esc", interval=1)
                    pag.press("f", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("n", interval=1)
                    pag.press("enter", interval=1)
                    pag.sleep(3)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        print('Email Receive 6: Move email to a different folder \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I173'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Email Receive 6: Failed to move an email to Notes folder")
                        print('Email Receive 6: Move email to a different folder \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I173'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

            else:
                logging.critical("Email Receive 6: Failed to locate Move email in actions menu")
                print("Email Receive 6: failed to locate Add contact in actions menu")
                print(data)
                ws['I173'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Receive 6: Failed to locate guide automation 1 in inbox")
            print("Email Receive 6: failed to locate 'guideautomation1@outlook.com' in inbox")
            print(data)
            ws['I173'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 6: Failed to locate guide automation 2")
        print("Email Receive 6: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I173'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive7(queue):  # STATUS: TESTED
    # View email attachments (JPG, PDF, DOC - All viewers)
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Folders' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Folders' in data:
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                pag.press("f2", interval=1)
                pag.press("v", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("enter", interval=1)

                    while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                        pag.press("esc", interval=1)
                        pag.press("j", interval=1)
                        data = readFromQueue(queue)

                        while 'JPEG.jpg' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'JPEG.jpg' in data:
                            pag.press("enter", interval=1)

                            while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                                pag.press("esc", interval=1)
                                pag.press("p", interval=1)
                                data = readFromQueue(queue)

                                while 'PDF.pdf' not in data:
                                    data = readFromQueue(queue)
                                    if start < timeout:
                                        pag.sleep(10)
                                        start += 10
                                        if start == timeout:
                                            print(data, ": Timed out after", timeout, "Seconds")
                                            returnhome()
                                            break

                                if 'PDF.pdf' in data:
                                    pag.press("enter", interval=1)

                                    while 'GMENU_ENTERTAINMENT_PDFVIEWER' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'GMENU_ENTERTAINMENT_PDFVIEWER' in data:
                                        print('Email Receive 7: View received attachments via email \n'
                                              '>>> Result: PASS \n'
                                              '>>> Current String:', data)
                                        ws['I174'].value = 'PASS'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()

                                    else:
                                        logging.critical("Email Receive 7: Failed to view attached emails")
                                        print('Email Receive 7: View received attachments via email \n'
                                              '>>> Result: FAIL \n'
                                              '>>> Current String:', data)
                                        ws['I174'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()

                                else:
                                    logging.critical("Email Receive 7: Failed to locate attached PDF")
                                    print("Email Receive 7: failed to locate attached PDF")
                                    print(data)
                                    ws['I174'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                        else:
                            logging.critical("Email Receive 7: Failed to locate attached JPEG")
                            print("Email Receive 7: failed to locate attached JPEG")
                            print(data)
                            ws['I174'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                else:
                    logging.critical("Email Receive 7: Failed to locate attached DOC")
                    print("Email Receive 7: failed to locate attached DOC")
                    print(data)
                    ws['I174'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 7: Failed to locate attachment email in archive folder")
                print("Emails Receive 7: failed to locate attachment email via archive folder")
                print(data)
                ws['I174'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Emails Receive 7: Failed to locate folder")
            print("Emails Receive 7: failed to locate folders")
            print(data)
            ws['I174'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Emails Receive 7: Failed to locate guide automation 2")
        print("Email Receive 7: failed to locate guide automation 2")
        print(data)
        ws['I174'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive8(queue):      # STATUS: TESTED
    # Blocks a sender via actions menu
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", interval=1)
        pag.press("k", interval=1)
        data = readFromQueue(queue)

        while 'Kieran.Baker@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Kieran.Baker@yourdolphin.com' in data:
            pag.press("f2", interval=1)
            pag.press("b", interval=1)
            data = readFromQueue(queue)

            while 'Block sender' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Block sender' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'Kieran.Baker@yourdolphin.com' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Kieran.Baker@yourdolphin.com' in data:
                    pag.press("enter", presses=2, interval=1)
                    print('Email Receive 8: Block email contact via actions menu \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I175'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Receive 8: Failed to block email contact via actions menu")
                    print('Email Receive 8: Block email contact via actions menu \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I175'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Receive 8: Failed to locate blocked sender")
                print("Email Receive 8: failed to locate Add contact in actions menu")
                print(data)
                ws['I175'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Email Receive 8: Failed to locate kieran.baker@yourdolphin.com in inbox")
            print("Email Receive 8: failed to locate 'kieran.baker@yourdolphin.com' in inbox")
            print(data)
            ws['I175'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 8: Failed to locate guide automation 2")
        print("Email Receive 8: failed to locate guideautomation2@outlook.com")
        print(data)
        ws['I175'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive9(queue):  # STATUS: TESTED
    # View details of a received email
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if 'GMENU_EMAIL_VIEW_DETAILS' in data:
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'Date' in data:
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 9: Failed to view details of a received email")
            print('Email Receive 9: View details of an email \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I176'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 9: Failed to locate details in the actions menu")
        print("Email Receive 9: failed to locate details via actions menu")
        print(data)
        ws['I176'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive10(queue):     # STATUS: TESTED
    # Mark email as Read / Unread
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("m", presses=2, interval=1)
    data = readFromQueue(queue)

    if 'Mark as read' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("m", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'Mark as unread' in data:
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails Receive 10: Failed to set email to read / unread")
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    elif 'Mark as unread' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("m", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'Mark as read' in data:
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Emails Receive 10: Failed to set email to read / unread")
            print('Email Receive 10: Mark email as read / unread \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I177'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 10: Failed to locate an unopened email")
        print("Email Receive 10: failed to locate unread email")
        print(data)
        ws['I177'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive11(queue):     # STATUS: TESTED
    # Search for specific email
    start = 0
    delay()
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("s", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_VIEW_ACTION_SEARCH' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_VIEW_ACTION_SEARCH' in data:
        pag.write("support@yourdolphin.com", interval=0.5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'support@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'support@yourdolphin.com' in data:
            print('Email Receive 11: Search for specific email \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I178'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Receive 11: Failed to search for 'Support'")
            print('Email Receive 11: Search for specific email \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I178'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Receive 11: Failed to open search email function")
        print("Email Receive 11: failed to open search email function")
        print(data)
        ws['I178'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive12(queue):     # STATUS: TESTED
    # Delete an email
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("i", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation1@outlook.com' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        pag.sleep(5)

    if 'guideautomation1@outlook.com' not in data:
        print('Email Receive 12: Delete email \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I179'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Email Receive 12: Failed to Delete Email(s) from Guide Automation 1")
        print('Email Receive 12: Delete email \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I179'].value = 'fail'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_receive13(queue):     # STATUS: TESTED
    # Save attachments from email and open via My saved attachments
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)

    while 'guideautomation2@outlook.com' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'Archive' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Archive' in data:
            pag.press("enter", interval=1)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                pag.press("f2", interval=1)
                pag.press("v", interval=1)
                pag.press("enter", interval=1)

                while 'DOC.doc' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'DOC.doc' in data:
                    pag.press("f2", interval=1)
                    pag.press("s", interval=1)
                    pag.press("enter", presses=3, interval=1)
                    pag.press("j", interval=1)
                    data = readFromQueue(queue)

                    if 'JPEG.jpg' in data:
                        pag.press("f2", interval=1)
                        pag.press("s", interval=1)
                        pag.press("enter", presses=3, interval=1)
                        pag.press("p", interval=1)
                        data = readFromQueue(queue)

                        if 'PDF.pdf' in data:
                            pag.press("f2", interval=1)
                            pag.press("s", interval=1)
                            pag.press("enter", presses=3, interval=1)
                            pag.press("esc", presses=3, interval=1)

                            while 'My saved attachments' not in data:
                                pag.press("m", interval=1)
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'My saved attachments' in data:
                                pag.press("enter", interval=1)
                                pag.press("d", interval=1)
                                data = readFromQueue(queue)
                                if 'DOC' in data:
                                    pag.press("enter", interval=1)

                                    while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                                        data = readFromQueue(queue)
                                        if start < timeout:
                                            pag.sleep(10)
                                            start += 10
                                            if start == timeout:
                                                print(data, ": Timed out after", timeout, "Seconds")
                                                returnhome()
                                                break

                                    if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                                        pag.press("esc", interval=1)
                                        pag.press("j", interval=1)
                                        data = readFromQueue(queue)
                                        if 'JPEG' in data:
                                            pag.press("enter", interval=1)

                                            while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                                                data = readFromQueue(queue)
                                                if start < timeout:
                                                    pag.sleep(10)
                                                    start += 10
                                                    if start == timeout:
                                                        print(data, ": Timed out after", timeout, "Seconds")
                                                        returnhome()
                                                        break

                                            if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                                                pag.press("esc", interval=1)
                                                pag.press("p", interval=1)
                                                data = readFromQueue(queue)
                                                if 'PDF' in data:
                                                    pag.press("enter", interval=1)

                                                    while 'GMENU_ENTERTAINMENT_PDFVIEWER' not in data:
                                                        data = readFromQueue(queue)
                                                        if start < timeout:
                                                            pag.sleep(10)
                                                            start += 10
                                                            if start == timeout:
                                                                print(data, ": Timed out after", timeout, "Seconds")
                                                                returnhome()
                                                                break

                                                    if 'GMENU_ENTERTAINMENT_PDFVIEWER' in data:
                                                        print('Email Receive 13: Open Saved attachments \n'
                                                              '>>> Result: PASS \n'
                                                              '>>> Current String:', data)
                                                        ws['I180'].value = 'PASS'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()

                                                    else:
                                                        logging.critical("Email Receive 13: "
                                                                         "Failed to open all saved attachments")
                                                        print('Email Receive 13: Open Saved attachments \n'
                                                              '>>> Result: FAIL \n'
                                                              '>>> Current String:', data)
                                                        ws['I180'].value = 'FAIL'
                                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                                        returnhome()
                                                else:
                                                    logging.critical("Email Receive 13:"
                                                                     "Failed to locate saved attachment PDF")
                                                    print("Email Receive 13: failed to locate saved attachment PDF")
                                                    print(data)
                                                    ws['I180'].value = 'FAIL'
                                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                                    returnhome()
                                            else:
                                                logging.critical("Email Receive 13: Failed to open image viewer")
                                                print("Email Receive 13: failed to open Image Viewer")
                                                print(data)
                                                ws['I180'].value = 'FAIL'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()
                                        else:
                                            logging.critical("Email Receive 13: Failed to locate"
                                                             "saved attachment JPEG")
                                            print("Email Receive 13: failed to locate saved attachment JPEG")
                                            print(data)
                                            ws['I180'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            returnhome()
                                    else:
                                        logging.critical("Email Receive 13: Failed to open Doc Viewer")
                                        print("Email Receive 13: failed to open Doc Viewer")
                                        print(data)
                                        ws['I180'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()
                                else:
                                    logging.critical("Email Receive 13: Failed to locate"
                                                     "saved attachment DOC")
                                    print("Email Receive 13: failed to locate saved attachment DOC")
                                    print(data)
                                    ws['I180'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()
                            else:
                                logging.critical("Email Receive 13: Failed to locate"
                                                 "my saved attachments")
                                print("Email Receive 13: failed to locate my saved attachments")
                                print(data)
                                ws['I180'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()
                        else:
                            logging.critical("Email Receive 13: Failed to locate"
                                             "attached PDF")
                            print("Email Receive 13: failed to locate attached PDF")
                            print(data)
                            ws['I180'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()
                    else:
                        logging.critical("Email Receive 13: Failed to locate attached JPEG")
                        print("Email Receive 13: failed to locate attached JPEG")
                        print(data)
                        ws['I180'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Email Receive 13: Failed to locate attached DOC")
                    print("Email Receive 13: failed to locate attached DOC file")
                    print(data)
                    ws['I180'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Email Receive 13: Failed to locate "
                                 "guide automation 1 email with attachments")
                print("Email Receive 13: failed to locate guide automation1 email with attachments")
                print(data)
                ws['I180'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Email Receive 13: Failed to locate archive folder")
            print("Email Receive 13: failed to locate Archive folder")
            print(data)
            ws['I180'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Email Receive 13: Failed to locate 'Guideautomation2@outlook.com")
        print("Email Receive 13: failed to locate 'guideautomation2@outlook.com'")
        print(data)
        ws['I180'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments1(queue):      # STATUS: TESTED
    # Sort documents
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'DOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'DOC' in data:
            pag.press("f2", interval=1)
            pag.press("s", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)

            while 'Sort by date' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Sort by date' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                data = readFromQueue(queue)

                while 'PDF' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'PDF' in data:
                    print('Email Attachments 1: Sort attachments \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I181'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Email Attachments 1: Failed to sort attachments by date")
                    print('Email Attachments 1: Sort attachments \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I181'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Attachments 1: Failed to locate sort by date")
                print("Email Attachments 1: failed to locate sort by date")
                print(data)
                ws['I181'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 1: Failed to locate saved attachment DOC")
            print("Email Attachments 1: failed to locate saved attachment DOC")
            print(data)
            ws['I181'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 1: Failed to locate my saved attachments folder")
        print("Email Attachments 1: failed to locate my saved attachments folder")
        print(data)
        ws['I181'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments2(queue):      # STATUS: TESTED
    # Search specific saved document
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("jpeg", interval=0.5)
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)

        while 'JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'JPEG' in data:
            print('Email Attachments 2: Search attachments \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I182'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Email Attachments 2: Failed to search attachments for JPEG")
            print('Email Attachments 2: Search attachments \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I182'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 2: Failed to locate my saved attachments folder")
        print("Email Attachments 2: failed to locate my saved attachments folder")
        print(data)
        ws['I182'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments3(queue):      # STATUS: TESTED
    # rename saved email attachment
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)

        while 'JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'renamed JPEG' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'renamed JPEG' in data:
                print('Email Attachments 3: Rename saved document \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I183'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Attachments 3: Failed to rename a saved attachment")
                print('Email Attachments 3: Rename saved document \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I183'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 3: Failed to locate saved attachment JPG")
            print("Email Attachments 3: failed to locate saved attachment JPG")
            print(data)
            ws['I183'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 3: Failed to locate my saved attachments folder")
        print("Email Attachments 3: failed to locate my saved attachments folder")
        print(data)
        ws['I183'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_attachments4(queue):      # STATUS: TESTED
    # delete saved attachments
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'My saved attachments' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'My saved attachments' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'renamed JPEG' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'renamed JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'Go to folder - GuideConnect' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Go to folder - GuideConnect' in data:
                print('Email Attachments 4: Delete saved document \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I184'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Attachments 4: Failed to delete saved document rename JPG")
                print('Email Attachments 4: Delete saved document \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I184'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Attachments 4: Failed to locate saved attachment renamed JPG")
            print("Email Attachments 4: failed to locate saved attachment JPG")
            print(data)
            ws['I184'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Attachments 4: Failed to locate my saved attachments folder")
        print("Email Attachments 4: failed to locate my saved attachments folder")
        print(data)
        ws['I184'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_folders1(queue):       # STATUS: TESTED
    # Add a custom folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Folders' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Folders' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_FOLDERS_ACTION_CREATE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_FOLDERS_ACTION_CREATE' in data:
            pag.write("added", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'added' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'added' in data:
                print('Email Folders 1: Create a new folder \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I185'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Folders 1: Failed to create new folder 'added'")
                print('Email Folders 1: Create a new folder \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I185'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Folders 1: Failed to open create a new folder")
            print("Email Folders 1: failed to open create new folder")
            print(data)
            ws['I185'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Folders 1: Failed to locate Folders folder")
        print("Email Folders 1: failed to locate Folders folder")
        print(data)
        ws['I185'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_folders2(queue):      # STATUS: TESTED
    # delete a custom folder
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Folders' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Folders' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'added' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'added' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            if 'added' not in data:
                print('Email Folders 2: delete a folder \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I186'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Folders 2: Failed to delete a custom folder")
                print('Email Folders 2: delete a folder \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I186'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Folders 2: Failed to locate added folder 'added' ")
            print("failed to locate added folder 'added'")
            print(data)
            ws['I186'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Folders 2: Failed to locate Folders folder")
        print("Email Folders 2: failed to locate Folders folder")
        print(data)
        ws['I186'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked1(queue):      # STATUS: TESTED
    # edit an existing blocked contact
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'Kieran.Baker@yourdolphin.com' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Kieran.Baker@yourdolphin.com' in data:
            pag.press("enter", interval=1)
            pag.write("changed", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'changed' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'changed' in data:
                print('Email Blocked 1: Edit an existing blocked contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I187'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 1: Failed to edit an existing blocked contact")
                print('Email Blocked 1: Edit an existing blocked contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I187'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 1: Failed to locate previous blocked contact"
                             "via inbox actions menu")
            print("Email Blocked 1: failed to locate added entry from inbox actions menu")
            print(data)
            ws['I187'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 1: Failed to locate Blocked senders folder")
        print("Email Blocked 1: failed to locate Blocked senders folder")
        print(data)
        ws['I187'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked2(queue):      # STATUS: TESTED
    # remove a blocked contact
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'changed' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'changed' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            pag.press("enter", presses=3, interval=1)
            data = readFromQueue(queue)

            while 'No items found' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No items found' in data:
                print('Email Blocked 2: Remove a blocked contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I188'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 1: Failed to remove an existing blocked contact")
                print('Email Blocked 1: Remove a blocked contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I188'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 2: Failed to locate previous blocked contact"
                             "via inbox actions menu")
            print("Email Blocked 2: failed to locate added entry from inbox actions menu")
            print(data)
            ws['I188'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 2: Failed to locate Blocked senders folder")
        print("Email Blocked 2: failed to locate Blocked senders folder")
        print(data)
        ws['I188'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked3(queue):      # STATUS: NOT TESTED
    # Add a blocked sender
    start = 0
    delay()
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)

    while 'Blocked email addresses' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Blocked email addresses' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_BLOCKED_SENDERS_ACTION_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_BLOCKED_SENDERS_ACTION_ADD' in data:
            pag.write("guideautomation1@outlook.com", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'guideautomation1@outlook.com' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'guideautomation1@outlook.com' in data:
                print('Email Blocked 3: Add a blocked contact via blocked contacts \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I189'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Email Blocked 3: Failed to add a blocked contact "
                                 "via blocked contacts feature")
                print('Email Blocked 3: Add a blocked contact via blocked contacts \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I189'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 3: Failed to open add blocked sender via blocked senders")
            print("Email Blocked 3: Failed to open add blocked sender via blocked senders")
            print(data)
            ws['I189'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 3: Failed to locate Blocked senders folder")
        print("Email Blocked 3: failed to locate Blocked senders folder")
        print(data)
        ws['I189'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def email_blocked4(queue):      # STATUS: NOT TESTED
    # Check blocked senders folder for blocked emails
    start = 0
    delay()
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
        pag.write("guideautomation2@outlook.com", interval=0.5)
        pag.press("enter", interval=1)
        pag.write("blocked", interval=0.5)
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_SENDING' in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            pag.press("g", interval=1)
            data = readFromQueue(queue)
            if 'guideautomation2@outlook.com' in data:
                pag.press("enter", interval=1)
                pag.press("f", interval=1)
                pag.press("enter", interval=1)
                pag.press("b", interval=1)
                data = readFromQueue(queue)
                if 'Blocked Senders' in data:
                    pag.sleep(30)
                    pag.press("enter", interval=1)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.press("del", interval=1)
                        pag.press("right", interval=1)
                        pag.press("enter", presses=2, interval=1)
                        data = readFromQueue(queue)
                        if 'No emails' in data:
                            print('Email Blocked 4: Blocked senders emails go to blocked folder \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I190'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Email Blocked 4: Failed to locate blocked sender email"
                                             "in blocked senders folder")
                            print('Email Blocked 4: Blocked senders emails go to blocked folder \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I190'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Email Blocked 4: Failed to locate expected blocked email in"
                                         "the blocked senders folder")
                        print("Email Blocked 4: failed to locate expected blocked email in blocked folder")
                        print(data)
                        ws['I190'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Email Blocked 4: failed to locate blocked senders folder")
                    print("Email Blocked 4: failed to locate blocked senders folder")
                    print(data)
                    ws['I190'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Email Blocked 4: Failed to locate guide automation 2 in email accounts")
                print("Email Blocked 4: failed to locate guide automation 2")
                print(data)
                ws['I190'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Email Blocked 4: Failed to send an email")
            print("Email Blocked 4: failed to send email")
            print(data)
            ws['I190'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Email Blocked 4: Failed to enter recipient email address")
        print("Email Blocked 4: failed to enter recipient email address")
        print(data)
        ws['I190'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def emailClean1(queue):
    delay()
    start = 0
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'guideautomation1@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)
        if 'Sent' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)

            while 'No emails' not in data:
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

                if 'No emails' in data:
                    logging.info("emailClean1: Email Sent folder cleaned")
                    print("emailClean1: Email Sent folder cleaned")
                    returnhome()

            else:
                logging.info("emailClean1: Failed to clean email sent folder")
                print("emailClean1: failed to clean email sent folder")
                returnhome()

        else:
            logging.critical("emailClean1: Failed to locate sent items folder")
            print("emailClean1: failed to locate sent items folder")
            returnhome()

    else:
        logging.critical("emailClean1: Failed to locate guideautomation1")
        print("emailClean1: failed to locate 'guideautomation1@outlook.com'")
        returnhome()


def emailClean2(queue):
    delay()
    start = 0
    pag.press("enter", interval=1)
    pag.press("g", interval=1)
    data = readFromQueue(queue)
    if 'guideautomation2@outlook.com' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Archive' in data:
            pag.press("enter", interval=1)

            while 'No emails' not in data:
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

                if 'No emails' in data:
                    pag.press("esc", interval=1)
                    pag.press("i", interval=1)
                    data = readFromQueue(queue)
                    if 'Inbox' in data:
                        pag.press("enter", interval=1)

                        while 'Kieran.Baker@yourdolphin.com' not in data:
                            pag.press("f2", interval=1)
                            pag.press("d", interval=1)
                            pag.press("enter", interval=1)
                            pag.press("right", interval=1)
                            pag.press("enter", presses=2, interval=1)

                            if 'Kieran.Baker@yourdolphin.com' in data:
                                pag.press("esc", interval=1)
                                pag.press("n", interval=1)
                                data = readFromQueue(queue)
                                if 'Notes' in data:
                                    pag.press("enter", interval=1)

                                    while 'No emails' not in data:
                                        pag.press("f2", interval=1)
                                        pag.press("d", interval=1)
                                        pag.press("enter", interval=1)
                                        pag.press("right", interval=1)
                                        pag.press("enter", presses=2, interval=1)

                                        if 'No emails' in data:
                                            pag.press("esc", interval=1)
                                            pag.press("s", interval=1)
                                            data = readFromQueue(queue)
                                            if 'Sent' in data:
                                                pag.press("enter", interval=1)

                                                while 'No emails' not in data:
                                                    pag.press("f2", interval=1)
                                                    pag.press("d", interval=1)
                                                    pag.press("enter", interval=1)
                                                    pag.press("right", interval=1)
                                                    pag.press("enter", presses=2, interval=1)

                                                    if 'No emails' in data:
                                                        logging.info("emailClean2: All email folders Cleaned")
                                                        print("emailClean2: All email folders Cleaned")
                                                        returnhome()

                                                else:
                                                    logging.critical("emailClean2: Failed to clean Sent items folder")
                                                    print("emailClean2: Failed to clean Sent items folder")
                                                    returnhome()

                                            else:
                                                logging.critical("emailClean2: Failed to locate sent items folder")
                                                print("emailClean2: failed to locate Sent items folder")
                                                returnhome()

                                    else:
                                        logging.critical("emailClean2: Failed to clear notes folder")
                                        print("emailClean2: failed to clear notes folder")
                                        returnhome()

                            else:
                                logging.critical("emailClean2: Failed to locate notes folder")
                                print("emailClean2: failed to locate notes folder")
                                returnhome()

                        else:
                            logging.critical("emailClean2: Failed to clean inbox folder")
                            print("emailClean2: Failed to clean inbox folder")
                            returnhome()

                    else:
                        logging.critical("emailClean2: Failed to locate inbox folder")
                        print("emailClean2: failed to locate inbox folder")
                        returnhome()
            else:
                logging.critical("emailClean2: Failed to clean archive folder")
                print("emailClean2: failed to clean archive folder")
                returnhome()

        else:
            logging.critical("emailClean2: Failed to locate Archive folder")
            print("emailClean2: failed to locate Archive folder")
            returnhome()

    else:
        logging.critical("emailClean2: Failed to locate guideautomation2")
        print("emailClean2:failed to locate 'guideautomation2@outlook.com'")
        returnhome()
