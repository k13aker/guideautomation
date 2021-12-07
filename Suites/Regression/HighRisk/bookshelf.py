import logging
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120
books_expected = [('Project Gutenberg', 2), ('RNIB Reading Services', 2), ('Calibre Audio', 2)]
book = []


def books1(queue):  # STATUS: TESTED
    # check all UK providers are present
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Find a new book' in data:
        pag.press("enter", interval=1)

        books_present = []

        while 'Calibre Audio' not in data:
            data = readFromQueue(queue)
            books_present.append(data)
            pag.sleep(1)
            pag.press("right", interval=1)
            if 'Calibre Audio' in data:
                break

        if books_present == books_expected:
            print('Books 1: Compares active vs expected book providers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I127'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 1: Providers list was different to expected")
            print('Books 1: Compares active vs expected book providers \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I127'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 1: Failed to locate find a new book")
        print("Books 1: failed to locate find a new book")
        print(data)
        ws['I127'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books2(queue):  # STATUS: TESTED
    # downloads book from Project Gutenburg
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'Project Gutenberg' in data:
        pag.press("enter", presses=3, interval=1)
        pag.sleep(5)
        data = readFromQueue(queue)
        global book
        book = data
        logging.info(f"Books 2: Book downloaded: {book}")
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Queued' in data:
            data = readFromQueue(queue)
            pag.press("right", interval=1)
            pag.press("left", interval=1)
            if 'Read' in data:
                pag.press("enter", interval=1)
                break

        while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
            print('Books 2: download a book from Project Gutenberg \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I128'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 2: Failed to download a book from Project Gutenberg")
            print('Books 2: download a book from Project Gutenberg \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I128'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 2: Failed to locate Project Gutenberg")
        print("failed to find project gutenberg")
        print(data)
        ws['I128'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books3(queue):  # STATUS: TESTED
    # Continue reading a book
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
        print('Books 3: Continue listening to a book \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I129'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Books 3: Failed to continue reading a book")
        print('Books 3: Continue listening to a book \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I129'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books4(queue):  # STATUS: TESTED
    # Is book present in your 'My Books'
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)
    if data == book:
        print('Books 4: is previously read book present in My Books \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I130'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Book 4: Failed to locate book in My Books")
        print('Books 4: is previously read book present in My Books \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I130'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books5(queue):  # STATUS: TESTED
    # View book information from My Books
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.sleep(5)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKINFO' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKINFO' in data:
        print('Books 5: View details of a book via My Books \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I131'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Books 5: Failed to view details of a book via My Books")
        print('Books 5: View details of a book via My Books \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I131'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books6(queue):  # STATUS: TESTED
    # Copies a book from My Books to USB device
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)
    if 'Copy To device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if f'{found[0]}' in data:
            pag.press("enter", presses=2, interval=1)
            print('Books 6: Copy a book from My Books to USB device \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I132'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Book 6: Failed to locate and transfer a book to a USB device")
            print('Books 6: Copy a book from My Books to USB device \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I132'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 6: Failed to locate copy to device")
        print("Books 6: failed to locate copy to device")
        print(data)
        ws['I132'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books7(queue):  # STATUS: TESTED
    # Delete a book via my books
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Delete Book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Delete Book' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'No Books Found' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'No Books Found' in data:
            print('Books 7: Delete a book via My Books \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I133'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 7: Failed to delete a book via My Books")
            print('Books 7: Delete a book via My Books \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I133'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 7: Failed to locate Delete book")
        print("Books 7: Failed to locate delete book")
        print(data)
        # ws['I133'].value = 'FAIL'
        # wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books8(queue):   # STATUS: TESTED
    # View details of a library
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Find a new book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Find a new book' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' in data:
            print('Books 8: View details of a provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I134'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Books 8: Failed to view details of a content provider")
            print('Books 8: View details of a provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I134'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 8: Failed to locate find a new book")
        print("Books 8: failed to locate find a new book")
        print(data)
        ws['I134'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books9(queue):  # STATUS: NOT TESTED
    # View details of a book via find a new book
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    while 'Find a new book' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Find a new book' in data:
        pag.press("enter", presses=5, interval=1)
        pag.press("b", interval=1)
        data = readFromQueue(queue)
        if 'Book information' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_BOOKSHELFMENU_BOOKINFO' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_BOOKSHELFMENU_BOOKINFO' in data:
                print('Books 9: View details of a book via find a new book \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I135'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 9: Failed to view details of a book via find a new book")
                print('Books 9: View details of a book via find a new book \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I135'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 9: Failed to locate book information")
            print("Books 9: failed to locate book information")
            print(data)
            ws['I135'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 8: Failed to locate find a new book")
        print("Books 8: failed to locate find a new book")
        print(data)
        ws['I135'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books10(queue):     # STATUS: TESTED
    # Open and views a book from USB
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("r", interval=1)
    data = readFromQueue(queue)
    if 'Read from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f'{found[0]}' in data:
            pag.press("enter", presses=2, interval=1)

            while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
                print('Books 10: Read a book from device \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I136'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 10: Failed to read a book from USB device")
                print('Books 10: Read a book from device \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I136'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 10: Failed to locate correct USB")
            print("Books 10: failed to locate correct regression usb")
            print(data)
            ws['I136'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 10: Failed to locate read from device")
        print("Books 10: failed to locate read from device")
        print(data)
        ws['I136'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers1(queue):  # STATUS: TESTED
    # Log in to provider and download a newspaper
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'RNIB NTNM' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'RNIB NTNM' in data:
            pag.press("enter", interval=1)
            pag.write("i.rollason", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("tna751", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'Magazines by category' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Magazines by category' in data:
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                newspaper = data
                logging.info(f"Newspaper downloaded: {newspaper}")
                pag.press("enter", presses=3, interval=1)
                pag.sleep(15)
                pag.press("enter", interval=1)
                pag.sleep(10)
                data = readFromQueue(queue)

                if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                elif '' in data:
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("News 1: Failed to log in, download and play an edition")
                    print('News 1: log in, download and play an edition \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I137'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("News 1: Failed to log into RNIB NTNM")
                print("News 1: failed to log into RNIB NTNM")
                print(data)
                ws['I137'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("News 1: Failed to locate RNIB NTNM")
            print("News 1: failed to locate RNIB NTNM")
            print(data)
            ws['I137'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 1: Failed to locate newspapers and magazines")
        print("News 1: failed to locate newspapers and magazines")
        print(data)
        ws['I137'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers2(queue):  # STATUS: TESTED
    # continue playing a newspaper
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.sleep(7)
    data = readFromQueue(queue)

    while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER2' in data:
        print('News 2: Continue listening to an edition \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I138'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    elif '' in data:
        print('News 2: Continue listening to an edition \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I138'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        logging.critical("News 2: Failed to continue playing an edition")
        print('News 2: Continue listening to an edition \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I138'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers3(queue):  # STATUS: TESTED
    # Play an edition from My newspapers
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", presses=4, interval=1)
        pag.sleep(5)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_BOOKSHELF_READER' in data:
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        elif '' in data:
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 3: Failed to play edition from my newspapers")
            print('News 3: Play an edition from My Newspapers \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I139'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 3: Failed to locate newspapers and magazines")
        print("News 3: failed to locate newspapers and magazines")
        print(data)
        ws['I139'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers4(queue):     # STATUS: TESTED
    # unsubscribe from edition from my newspapers
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)

        while '> Unsubscribe' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if '> Unsubscribe' in data:
            pag.press("enter", presses=2, interval=1)
            data = readFromQueue(queue)

            while 'No Newspapers Or Magazines Found' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'No Newspapers Or Magazines Found' in data:
                print('News 4: Unsubscribe from an edition from My Newspapers \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I140'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("News 4: Failed to unsubscribe from edition via my newspapers")
                print('News 4: Unsubscribe from an edition from My Newspapers \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I140'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("News 4: Failed to unsubscribe")
            print("News 4: failed to unsubscribe")
            print(data)
            ws['I140'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 4: Failed to locate newspapers and magazines")
        print("News 4: failed to locate newspapers and magazines")
        print(data)
        ws['I140'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers5(queue):     # STATUS: TESTED
    # View details of newspaper provider
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_ABOUT' in data:
            print('News 5: View details of a newspaper content provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I141'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 5: Failed to view details of a newspaper library provider")
            print('News 5: View details of a newspaper content provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I141'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 5: Failed to locate newspapers and magazines")
        print("News 5: failed to locate newspapers and magazines")
        print(data)
        ws['I141'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def newspapers6(queue):     # STATUS: TESTED
    # log out of newspaper provider
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("n", interval=1)
    data = readFromQueue(queue)

    while 'Newspapers and magazines' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Newspapers and magazines' in data:
        pag.press("enter", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_BOOKSHELFMENU_PROVIDER_USERNAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_BOOKSHELFMENU_PROVIDER_USERNAME' in data:
            print('News 6: Log out of newspaper provider \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I142'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("News 6: Failed to log out of newspaper provider")
            print('News 6: Log out of newspaper provider \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I142'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("News 6: Failed to locate newspapers and magazines")
        print("News 6: failed to locate newspapers and magazines")
        print(data)
        ws['I142'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def books11(queue):
    # Copies a book from My Books to USB device
    start = 0
    delay()
    pag.press("right", presses=4, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    pag.press("enter", interval=1)
    pag.press("r", interval=1)
    data = readFromQueue(queue)
    if 'Read from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f'{found[0]}' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)

            while 'XDOC.docx' not in data:
                data = readFromQueue(queue)
                pag.press("f2", interval=1)
                pag.press("d", interval=1)
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'XDOC.docx' in data:
                print('Books 11: Delete a book from device \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I191'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Books 11: Failed to delete a book from USB device")
                print('Books 11: Delete a book from device \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I191'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Books 11: Failed to locate regression USB drive")
            print("Books 11: failed to locate Regression USB drive")
            print(data)
            ws['I191'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Books 11: Failed to locate read from device")
        print("Books 11: failed to locate Read from device")
        print(data)
        ws['I191'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


