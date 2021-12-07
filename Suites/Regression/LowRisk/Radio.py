import logging
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

# wb = load_workbook('')
# ws = wb['']

radio_expected = [("Absolute Radio 80's", 2), ("Absolute Radio 90's", 2), ('Absolute\xa0Radio', 2),
                  ('Absolute\xa0Radio\xa060s', 2), ('Absolute\xa0Radio\xa070s', 2),
                  ('Absolute\xa0Radio\xa0Classic Rock', 2), ('BBC - World Service', 2),
                  ('BBC Local Radio Berkshire', 2), ('BBC Local Radio Bristol', 2),
                  ('BBC Local Radio Cambridge', 2), ('BBC Local Radio Cornwall', 2),
                  ('BBC Local Radio Coventry & Warwickshire', 2), ('BBC Local Radio Cumbria', 2),
                  ('BBC Local Radio Derby', 2), ('BBC Local Radio Devon', 2), ('BBC Local Radio Essex', 2),
                  ('BBC Local Radio Gloucestershire', 2), ('BBC local Radio Guernsey', 2),
                  ('BBC Local Radio Hereford & Worcester', 2), ('BBC Local Radio Humberside', 2),
                  ('BBC Local Radio Jersey', 2), ('BBC Local Radio Kent', 2), ('BBC Local Radio Lancashire', 2),
                  ('BBC Local Radio Leeds', 2), ('BBC Local Radio Leicester', 2), ('BBC Local Radio Lincolnshire', 2),
                  ('BBC Local Radio London', 2), ('BBC Local Radio Manchester', 2), ('BBC Local Radio Merseyside', 2),
                  ('BBC Local Radio Newcastle', 2), ('BBC Local Radio Norfolk', 2), ('BBC Local Radio Northampton', 2),
                  ('BBC Local Radio Oxford', 2), ('BBC Local Radio Sheffield', 2), ('BBC Local Radio Shropshire', 2),
                  ('BBC Local Radio Solent', 2), ('BBC Local radio Somerset', 2), ('BBC Local Radio Stoke', 2),
                  ('BBC Local Radio Suffolk', 2), ('BBC Local Radio Sussex', 2), ('BBC Local Radio Tees', 2),
                  ('BBC Local Radio Three Counties', 2), ('BBC Local Radio Wiltshire', 2), ('BBC Local Radio York', 2),
                  ('BBC Radio 1', 2), ('BBC Radio 2', 2), ('BBC Radio 3', 2), ('BBC Radio 4', 2), ('BBC Radio 5', 2),
                  ('BBC Radio 6', 2), ('BBC World Service', 2), ('BBC\xa0World\xa0Service News', 2),
                  ('British Comedy Radio GB', 2), ('Classic Fm', 2), ('Heart London', 2), ('Lbc London', 2),
                  ('LBC News', 2), ('Nottingham & Derby roots radio', 2), ('R.T.E. Raidió na Gaeltachta', 2),
                  ('RNIB Connect radio', 2), ('Smooth Scotland', 2), ('Talksport', 2), ('UCB UK', 2)]


def radio1(queue):   # STATUS: TESTED
    # Play a radio station from favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=4, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
        pag.press("enter", interval=1)
        print('Radio 1: Play a station from favourites \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I115'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        pag.press("enter", interval=1)
        logging.critical("Radio 1: Failed to play a station from favourites")
        print('Radio 1: Play a station from favourites \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I115'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio2(queue):  # STATUS: TESTED
    # Continue listening to a previous station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(5)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_RADIO_LISTEN' in data:
        pag.press("enter", interval=1)
        print('Radio 2: Continue Playing a station \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I116'].value = 'PASS'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()

    else:
        pag.press("enter", interval=1)
        logging.critical("Radio 2: Failed to continue playing a radio station")
        print('Radio 2: Continue Playing a station \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I116'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio3(queue):  # STATUS: TESTED
    # Add a station to favourites via favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.press("a", interval=1)
        data = readFromQueue(queue)
        if "Absolute Radio 80's" in data:
            print('Radio 3: Add a radio station to favourites via favourites \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I117'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            print('Radio 3: Add a radio station to favourites via favourites \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I117'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 3: Failed to add a radio station to favourites via favourites")
        print("Radio 3: failed to navigate to favourites")
        print(data)
        ws['I117'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio4(queue):  # STATUS: TESTED
    # Add a custom radio station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("a", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_RADIO_CUSTOM' in data:
            pag.write("test", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("http://stream.live.vc.bbcmedia.co.uk/bbc_radio_cornwall", interval=0.5)
            pag.press("enter", presses=2, interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)
            if 'test' in data:
                print('Radio 4: Add a custom radio station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I118'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 4: Failed to add a custom radio station")
                print('Radio 4: Add a custom radio station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I118'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 4: Failed to enter add custom station")
            print("Radio 4: Failed to enter add custom station")
            print(data)
            ws['I118'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 4: Failed to navigate to favourites")
        print("Radio 4: Failed to enter favourites")
        print(data)
        ws['I118'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio5(queue):  # STATUS: TESTED
    # Remove a radio station from favourites
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'RNIB Connect radio' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'Remove selected from favourites' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'BBC Radio 4' in data:
                    print('Radio 5: Remove a station from favourites \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I119'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Radio 5: Failed to remove a radio station from favourites")
                    print('Radio 5: Remove a station from favourites \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I119'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Radio 5: Failed to locate remove from favourites")
                print("Radio 5: Failed to locate remove from favourites")
                print(data)
                ws['I119'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 5: Failed to locate RNIB Radio ")
            print("Radio 5: Failed to locate RNIB Radio")
            print(data)
            ws['I119'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 5: Failed to locate favourites")
        print("Radio 5: Failed to locate favourites")
        print(data)
        ws['I119'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio6(queue):  # STATUS: TESTED
    # Views details of a radio favourite
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Favourites' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Name' in data:
            print('Radio 6: View details of a radio favourite \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I120'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Radio 6: Failed to view details of a radio favourite")
            print('Radio 6: View details of a radio favourite \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I120'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 6: Failed to locate favourites")
        print("Radio 6: Failed to locate favourites")
        print(data)
        ws['I120'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio7(queue):  # STATUS: TESTED
    # Remove a favourite podcast from play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("right", interval=1)
        pag.press("left", interval=1)
        data = readFromQueue(queue)
        if "Remove From favourites" in data:
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("esc", interval=1)
            pag.press("f", interval=1)
            pag.press("enter", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)
            if 'BBC Radio 4' in data:
                print('Radio 7: Removes a radio favourite via Play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I121'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 7: Failed to remove a radio favourite via play a new station")
                print('Radio 7: Removes a radio favourite via Play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I121'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 7: Failed to locate remove from favourites via play a new station")
            print("Radio 7: Failed to locate remove from favourites")
            print(data)
            ws['I121'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 7: Failed to locate play a new station")
        print("Radio 7: Failed to locate play a new station")
        print(data)
        ws['I121'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio8(queue):  # STATUS: TESTED
    # Adds a radio to favourites via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("c", interval=1)
        data = readFromQueue(queue)
        if 'Classic Fm' in data:
            pag.press("f2", interval=1)
            pag.press("enter", presses=2, interval=1)
            pag.press("esc", interval=1)
            pag.press("f", interval=1)
            pag.press("enter", interval=1)
            pag.press("c", interval=1)
            data = readFromQueue(queue)
            if 'Classic Fm' in data:
                print('Radio 8: Adds a station to favourites via play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I122'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 8: Failed to addd a station to favourites via play a new station")
                print('Radio 8: Adds a station to favourites via play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I122'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 8: Failed to locate Classic FM")
            print("Radio 8: Failed to locate Classic FM")
            print(data)
            ws['I122'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 8: Failed to locate play a new station")
        print("Radio 8: Failed to locate play a new station")
        print(data)
        ws['I122'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio9(queue):  # STATUS: TESTED
    # Searches for a specific station via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ENTERTAINMENT_RADIO_STATIONS_SEARCH_EDIT' in data:
            pag.write("talk", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Talksport' in data:
                print('Radio 9: Searches for a specific station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I123'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 9: Failed to search for a specific station")
                print('Radio 9: Searches for a specific station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I123'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 9: Failed to enter search radio list")
            print("Radio 9: Failed to enter search radio list")
            print(data)
            ws['I123'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 9: Failed to locate play a new station")
        print("Radio 9: Failed to locate play a new station")
        print(data)
        ws['I123'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio10(queue):  # STATUS: TESTED
    # Changes the Language list to Czech
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'Switch language' in data:
            pag.press("enter", interval=1)
            pag.press("c", presses=2, interval=1)
            data = readFromQueue(queue)
            if 'Czech' in data:
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'ABradio.cz\xa0Radio\xa0Folk' in data:
                    print('Radio 10: Changes radio language list to Czech \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I124'].value = 'PASS'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Radio 10: Failed to change language list to Czech")
                    print('Radio 10: Changes radio language list to Czech \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I124'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Radio 10: Failed to locate Czech in language list")
                print("Radio 10: Failed to locate Czech in language list")
                print(data)
                ws['I124'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 10: Failed to locate switch language")
            print("Radio 10: Failed to locate switch language")
            print(data)
            ws['I124'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 10: Failed to locate play a new station")
        print("Radio 10: Failed to locate play a new station")
        print(data)
        ws['I124'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio11(queue):  # STATUS: TESTED
    # Views details of station via play a new station
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)
        if 'Details' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Name' in data:
                print('Radio 11: Views details of station via play a new station \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I125'].value = 'PASS'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Radio 11: Failed to view details of station via play a new station")
                print('Radio 11: Views details of station via play a new station \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I125'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Radio 11: Failed to locate details via play a new station")
            print("Radio 11: Failed to locate details")
            print(data)
            ws['I125'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 11: Failed to locate play a new station")
        print("Radio 11: Failed to locate play a new station")
        print(data)
        ws['I125'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def radio12(queue):  # STATUS: TESTED
    # iterates through active radio list and compares vs expected
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("p", interval=1)
    data = readFromQueue(queue)
    if 'Play a new station' in data:
        pag.press("enter", interval=1)

        radio_present = []

        while 'UCB UK' not in data:
            data = readFromQueue(queue)
            radio_present.append(data)
            pag.sleep(1)
            pag.press("right", interval=1)
            if 'UCB UK' in data:
                break

        if radio_present == radio_expected:
            print('Radio 12: Compares active vs expected radio stations \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I126'].value = 'PASS'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Radio 12: Failed to match active vs expected radio list")
            print('Radio 12: Compares active vs expected radio stations \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I126'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Radio 12: Failed to locate play a new station")
        print("Radio 12: Failed to locate play a new station")
        print(data)
        ws['I126'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()
