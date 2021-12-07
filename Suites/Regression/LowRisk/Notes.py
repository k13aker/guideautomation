import logging

import pyautogui as pag
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

wb = load_workbook('Test Reports/Automated Test cases.xlsx')
ws = wb['Notes']

# STATUS: ALL CURRENT TEST CASES TESTED
# TODO: OPTION TO ADD MORE TEST CASES


def tnote1(queue):  # STATUS: TESTED
    # with no notes present, does the prompt say no notes available?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    pag.sleep(3)
    data = readFromQueue(queue)
    if 'No notes available' in data:
        print('Notes 1: with no notes present, does the prompt say no notes available? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I62'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Notes 1: No available notes prompt missing")
        print('Notes 1: with no notes present, does the prompt say no notes available? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I62'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote2(queue): # STATUS: TESTED
    # Can you create a new text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Add new text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Add new text note' in data:
        pag.press("enter", interval=1)
        pag.write("text note", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("this is a text note", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'text note' in data:
            print('Notes 2: Can you create and view a text note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I63'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 2: Failed to create and view a text note")
            print('Notes 2: Can you create and view a text note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I63'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 2: Failed to add a new text note")
        print("Notes 2: failed to add a new text note")
        print(data)
        ws['I63'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote3(queue):  # STATUS: TESTED
    # can you rename an existing text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'text note' in data:
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_NOTES_MAIN_RENAME_TEXT_TITLE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_NOTES_MAIN_RENAME_TEXT_TITLE' in data:
            pag.write("renamed", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.sleep(5)
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'renamed' in data:
                print('Notes 3: Can you rename an existing text note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I64'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Notes 3: Failed to rename existing text note")
                print('Notes 3: Can you rename an existing text note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I64'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 3: Failed to rename text note")
            print("Notes 3: failed to Rename text note")
            print(data)
            ws['I64'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 3: Text note not found")
        print("Notes 3: text note not found in notes")
        print(data)
        ws['I64'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote4(queue):  # STATUS: TESTED
    # can you edit a text note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("edited ", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(3)
        data = readFromQueue(queue)
        if 'OK' in data:
            pag.press("enter", interval=1)
            print('Notes 4: Can you edit a text note? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I65'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 4: Failed to edit a text note")
            print('Notes 4: Can you edit a text note? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I65'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 4: Failed to enter text note")
        print("Notes 4: Failed to enter text note")
        print(data)
        ws['I65'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote5(queue):  # STATUS: TESTED
    # Can you send a note via an email?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)

        while 'Send as email' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Send as email' in data:
            pag.press("enter", interval=1, presses=3)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Text note", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("text note attached", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_EMAIL_SENDING' in data:
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'OK' in data:
                pag.press("enter", interval=1)
                returnhome()
                pag.sleep(3)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)

                while 'guideautomation2@outlook.com'not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter", interval=1)
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.sleep(3)
                        pag.press("f2", interval=1)
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'renamed.txt' in data:
                            print('Notes 5: Can you send a note via email? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I66'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Notes 5: Failed to send text note via attachment")
                            print('Notes 5: Can you send a note via email? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I66'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Notes 5: Failed to locate email with attachment")
                        print("Notes 5: failed to locate email with attachment")
                        print(data)
                        ws['I66'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Notes 5: Failed to enter guide automation 2")
                    print("Notes 5: Failed to enter guide automation 2")
                    print(data)
                    ws['I66'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Notes 5: Failed to send email with attachment")
                print("Notes 5: Failed to send email with attachment")
                print(data)
                ws['I66'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 5: Failed to navigate to send as email")
            print("Notes 5: failed to navigate to send as email")
            print(data)
            ws['I66'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 5: Unable to find text note")
        print("Notes 5: unable to find text note")
        print(data)
        ws['I66'].value = 'FAIL'
        wb.save('Test chReports/Automated Test cases.xlsx')
        returnhome()


def tnote6(queue):  # STATUS: TESTED
    # Can you view details and delete a note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'renamed' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'renamed' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.sleep(3)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_NOTES_MAIN_ACTIONS_DETAILS' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_NOTES_MAIN_ACTIONS_DETAILS' in data:
            pag.sleep(3)
            pag.press("esc", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", presses=2, interval=1)
            pag.press("enter", interval=1, presses=3)
            data = readFromQueue(queue)
            if 'No notes available' in data:
                print('Notes 6: Can you view details and then delete a note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I67'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Notes 6: Failed to view details and delete a note")
                print('Notes 6: Can you view details and then delete a note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I67'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Notes 6: Unable to view details of a note")
            print("Notes 6: unable to view details of note")
            print(data)
            ws['I67'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 6: Unable to find text note")
        print("Notes 6: unable to find text note")
        print(data)
        ws['I67'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def tnote7(queue):  # STATUS: TESTED
    # Can you create a note without saving?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Add new text note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Add new text note' in data:
        pag.press("enter", interval=1)
        pag.write("text note", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("this is a text note", interval=0.5)
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1, presses=2)
        data = readFromQueue(queue)
        if 'No notes available' in data:
            print('Notes 7: Can you create a text note without saving? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I68'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Notes 7: Failed to ESC a text note without saving ")
            print('Notes 7: Can you create a text note without saving? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I68'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Notes 7: Failed to add a new text note")
        print("Notes 7: Failed to add a new text note")
        print(data)
        ws['I68'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def anote1(queue):  # STATUS: TESTED
    # Can you create a new audio note?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_NOTES_MAIN_VOICE_TITLE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_NOTES_MAIN_VOICE_TITLE' in data:
        pag.write("voice note", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(5)
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("v")
        data = readFromQueue(queue)
        if 'voice note' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            data = readFromQueue(queue)
            if '' in data:
                print('Audio Notes 1: Can you create and view an audio note? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I69'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Audio Notes 1: Failed to create a new audio note")
                print('Audio Notes 1: Can you create and view an audio note? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I69'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Audio notes 1: Failed to find voice note")
            print("Audio Notes 1: failed to find voice note")
            print(data)
            ws['I69'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Audio notes 1: Failed to add new voice note")
        print("Audio Notes 1: failed to add new voice note")
        print(data)
        ws['I69'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def anote2(queue):  # STATUS: TESTED
    # can you send a audio note via email?
    delay()
    start = 0
    pag.press("right", presses=7, interval=1)
    pag.press("enter")
    pag.sleep(3)
    data = readFromQueue(queue)

    while 'voice note' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'voice note' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)

        while 'Send as email' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Send as email' in data:
            pag.press("enter", interval=1, presses=3)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("Audio note", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("audio note attached", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_EMAIL_SENDING' in data:
                print("Email Sending...")
                pag.sleep(5)
                data = readFromQueue(queue)

            if 'OK' in data:
                pag.press("enter", interval=1)
                returnhome()
                pag.sleep(3)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)

                while 'guideautomation2@outlook.com' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter")
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation1@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation1@outlook.com' in data:
                        pag.sleep(3)
                        pag.press("f2")
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'voice note.mp3' in data:
                            print('Audio note 2: Can you send a note via email? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I70'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Audio note 2: Failed to send audio note via email")
                            print('Audio note 2: Can you send a note via email? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I70'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Audio note 2: Failed to locate email with attachement")
                        print("Audio note 2: failed to locate email with attachment")
                        print(data)
                        ws['I70'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Audio note 2: Failed to enter guide automation")
                    print("Audio note 2: Failed to enter guide automation 2")
                    print(data)
                    ws['I70'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Audio note 2: Failed to send email with attachment")
                print("Audio note 2: Failed to send email with attachment")
                print(data)
                ws['I70'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Audio note 2: Failed to navigate to send as email")
            print("Audio note 2: failed to navigate to send as email")
            print(data)
            ws['I70'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Audio note 2: Unable to find audio note")
        print("Audio note 2: unable to find audio note")
        print(data)
        ws['I70'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()
