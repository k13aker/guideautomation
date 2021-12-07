import logging

from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

# wb = load_workbook('Test Reports/Automated Test cases.xlsx')
# ws = wb['Address & Calendar']


def calendar1(queue):  # STATUS: TESTED
    # Are you able to add a new event to your calendar?
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)  # address book
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("enter", presses=3, interval=1)
    data = readFromQueue(queue)
    pag.sleep(2)

    while "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if "GMENU_CALENDAR_ADD_EVENT_STEPS_DETAILS" in data:
        # TEST CASE 1: - TODAY'S EVENTS
        pag.write("tc1", interval=0.5)
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)  # DAILY REPEAT
        pag.press("enter", presses=4, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(2)
        data = readFromQueue(queue)
        pag.sleep(5)

        if "Event name" in data:
            print('Calendar 1 - TC1: Calendar - Todays Events \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I58'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')

            # TEST CASE 2: Upcoming Events
            pag.press("esc", presses=2, interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            pag.sleep(1)
            if "Upcoming events" in data:
                pag.press("enter")
                data = readFromQueue(queue)
                pag.sleep(5)
                if 'tc1' in data:
                    pag.sleep(2)
                    print('Calendar 1 - TC2: Calendar - Upcoming Events \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I58'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')

                    # TEST CASE 3: Calendar - View Calendar
                    pag.press("esc", interval=1)
                    pag.press("right", presses=2, interval=1)
                    pag.press("enter", presses=3, interval=1)
                    data = readFromQueue(queue)
                    pag.sleep(5)
                    if 'Event name' in data:
                        pag.sleep(2)
                        print('Calendar 1 - TC3: Calendar - View Calendar \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I58'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Calendar 1 (TC3) - Failed to find calendar event")
                        print('Calendar 1 - TC3: Failed to find calendar event \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I58'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Calendar 1 (TC2) - Failed to find upcoming event")
                    print('Calendar 1 - TC2: Failed to find the upcoming event \n'
                          '>>> Current String:', data)
                    ws['I58'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Calendar 1 (TC2) - Failed to enter upcoming events")
                print('Calendar 1 - TC2: Failed to enter upcoming events \n'
                      '>>> Current String:', data)
                ws['I58'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 1 (TC1) - Failed to locate event name")
            print('Calendar 1 - TC1: Failed to locate event name \n'
                  '>>> Current String:', data)
            ws['I58'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 1: Failed to add a calendar event")
        print('Calendar 1: Failed to add a calendar event \n'
              '>>> Current String:', data)
        ws['I58'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar2(queue):   # STATUS: TESTED
    # edit details of an event
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", interval=1, presses=4)
        data = readFromQueue(queue)

        if 'GMENU_CALENDAR_EDIT_EVENT_UPDATE_DETAILS' in data:
            pag.write("update", interval=0.5)
            pag.press("enter", interval=1)

            print('Calendar 2: View and edit a calendar entry \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I59'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Calendar 2: Failed to view and edit calendar entry")
            print('Calendar 2: View and edit a calendar entry \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I59'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 2: Failed to navigate to calendar")
        print("Calendar 2: Failed to navigate to Calendar")
        print(data)
        ws['I59'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar3(queue):   # STATUS: TESTED
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        data = readFromQueue(queue)
        if 'Send details as attachment' in data:
            pag.press("enter", presses=3, interval=1)
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("sent from calendar", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                returnhome()
                delay()
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                data = readFromQueue(queue)
                if 'guideautomation2@outlook.com' in data:
                    pag.press("enter", interval=1)
                    pag.press("i", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("f2", interval=1)
                    pag.press("v", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'calendar.ics' in data:
                        print('Calendar 3: Send event via email \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I60'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Calendar 3: Failed to send event via email")
                        print('Calendar 3: Send event via email \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I60'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Calendar 3: Failed to find Email acccount")
                    print("Calendar 3: failed to find 'guideautomation2@outlook.com'")
                    print(data)
                    ws['I60'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Calendar 3: Failed to send event as email")
                print("Calendar 3: Failed to send event as email")
                print(data)
                ws['I60'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 3: Failed to locate send event as attachment")
            print("Calendar 3: Failed to locate send event as attachment")
            print(data)
            ws['I60'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 3: Failed to navigate to Calendar")
        print("Calendar 3: Failed to navigate to Calendar")
        print(data)
        ws['I60'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calendar4(queue):   # STATUS: TESTED
    # Can you delete a calendar event?
    start = 0
    delay()
    pag.press("right", presses=5, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    if 'Calendar' in data:
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)
        if 'Event name' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'No events on this day' in data:
                print('Calendar 4: Can you delete a calendar entry? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I61'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Calendar 4: Failed to delete calendar entry")
                print('Calendar 4: Can you delete a calendar entry? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I61'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Calendar 4: Failed to find pre existing event")
            print("Calendar 4: Failed to find pre existing event")
            print(data)
            ws['I61'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Calendar 4: Failed to navigate to calendar")
        print("Calendar 4: Failed to navigate to Calendar")
        print(data)
        ws['I61'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address1(queue):    # STATUS: TESTED
    # Can you add a contact to your address book?
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'GMENU_ADDRESSBOOK_ADD' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_ADDRESSBOOK_ADD' in data:
        pag.write("Joe", interval=0.5)  # first name
        pag.press("enter", interval=1)
        pag.write("bloggs", interval=0.5)  # surname
        pag.press("enter", interval=1)
        pag.write("0123456789101", interval=0.5)  # telephone number
        pag.press("enter", interval=1)
        pag.write("22", interval=0.5)  # house number / number
        pag.press("enter", interval=1)
        pag.write("Dolphin street", interval=0.5)  # street name
        pag.press("enter", interval=1)
        pag.write("dolphin town", interval=0.5)  # town name
        pag.press("enter", interval=1)
        pag.write("dolphin county", interval=0.5)  # county name
        pag.press("enter", interval=1)
        pag.write("B45 9XE", interval=0.5)  # postcode
        pag.press("enter", interval=1)
        pag.write("United kingdom", interval=0.5)  # country
        pag.press("enter", interval=1)
        pag.write("joebloggs@dolphin.co.uk", interval=0.5)  # email
        pag.press("enter", interval=1)
        pag.write("Dolphin computer access", interval=0.5)  # company name
        pag.press("enter", interval=1)
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)
        if "Joe bloggs" in data:
            print('Address 1: Can you add a new contact? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I49'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 1: failed to add new contact")
            print('Address 1: Can you add a new contact? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I49'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 1: Failed to enter address book")
        print("Address 1: Failed to enter address book")
        print(data)
        ws['I49'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address2(queue):    # STATUS: TESTED
    # Edit an address book contact
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("j", interval=1)
    data = readFromQueue(queue)

    while 'Joe bloggs' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Joe bloggs' in data:
        pag.press("enter", presses=2, interval=1)
        pag.write("edited", interval=0.5)  # first name
        pag.press("enter", interval=1)
        pag.press("down", interval=1)
        pag.press("enter", interval=1)
        pag.write("user", interval=0.5)  # surname
        pag.press("enter", interval=1)
        pag.press("down", presses=2, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit phone number", interval=0.5)  # telephone number
        pag.press("enter", interval=1)
        pag.press("down", presses=3, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit house number", interval=0.5)  # house name / number
        pag.press("enter", interval=1)
        pag.press("down", presses=4, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit street name", interval=0.5)  # street name
        pag.press("enter", interval=1)
        pag.press("down", presses=5, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit town", interval=0.5)  # town
        pag.press("enter", interval=1)
        pag.press("down", presses=6, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit county", interval=0.5)  # county
        pag.press("enter", interval=1)
        pag.press("down", presses=7, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit postcode", interval=0.5)  # postcode
        pag.press("enter", interval=1)
        pag.press("down", presses=8, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit country", interval=0.5)  # country
        pag.press("enter", interval=1)
        pag.press("down", presses=9, interval=1)
        pag.press("enter", interval=1)
        pag.write("guideautomation2@outlook.com", interval=0.5)  # email
        pag.press("enter", interval=1)
        pag.press("down", presses=10, interval=1)
        pag.press("enter", interval=1)
        pag.write("edit company", interval=0.5)  # company
        pag.press("enter", interval=1)
        pag.press("esc", interval=1)
        pag.press("e", interval=1)

        data = readFromQueue(queue)
        if "edited user" in data:
            print('Address 2: Edit an existing contact \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I50'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 2: Failed to edit existing contact")
            print('Address 2: Edit an existing contact \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I50'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 2: Failed to locate user 'Joe Bloggs")
        print("Address 2: Failed to locate user 'Joe Bloggs'")
        print(data)
        ws['I50'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address3(queue):    # STATUS: TESTED
    # Send an email to address book contact
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        if 'GMENU_EMAIL_NEW_SUBJECT' in data:
            pag.write("Sent from Address book", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("this is a test email", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                pag.press("enter", interval=1)
                print('Address 3: Send an email to address contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I51'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 3: Failed to send an email to address contact")
                print('Address 3: Send an email to address contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I51'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 3: Failed to enter send email")
            print("Address 3: failed to enter send email")
            print(data)
            ws['I51'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 3: Failed to locate user 'Edited user'")
        print("Address 3: Failed to locate user 'Edited user'")
        print(data)
        ws['I51'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address4(queue):    # STATUS: TESTED
    # Invite to video calling
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'GMENU_EMAIL_INVITE_MESSAGE' in data:
            pag.press("enter", presses=4, interval=1)
            data = readFromQueue(queue)
            if 'Dolphin Computer Access Ltd.' in data:
                print('Address 4: Send an email to address contact \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I52'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 4: Failed to send email to address contact")
                print('Address 4: Send an email to address contact \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I52'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 4: Failed to enter invite to video calling")
            print("Address 4: failed to enter invite to video calling")
            print(data)
            ws['I52'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 4: Failed to locate 'Edited user'")
        print("Address 4: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I52'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address5(queue):    # STATUS: TESTED
    # Compose a letter from address book
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'GMENU_DOCUMENTS_VIEW' in data:
            print('Address 5: Compose a letter \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I53'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 5: Failed to compose a letter")
            print('Address 5: Compose a letter \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I53'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 5: Failed to locate user 'Edited user'")
        print("Address 5: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I53'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address6(queue):    # STATUS: TESTED
    # Sends a VCF as attachment, deletes existing entry and then imports it to address book
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)

    while 'edited user' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'edited user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=3, interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_EMAIL_NEW_TO_TYPE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_EMAIL_NEW_TO_TYPE' in data:
            pag.write("guideautomation2@outlook.com", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("sent from address", interval=0.5)
            pag.press("enter", interval=1)
            pag.write("test", interval=0.5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_EMAIL_SENDING' in data:
                data = readFromQueue(queue)

            while 'OK' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'OK' in data:
                pag.press("enter", interval=1)
                pag.press("e", interval=1)
                data = readFromQueue(queue)

                if 'edited user' in data:
                    pag.press("f2", interval=1)
                    pag.press("d", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    returnhome()
                    delay()
                    pag.press("enter", interval=1)
                    pag.press("g", interval=1)
                    data = readFromQueue(queue)

                    while 'guideautomation2@outlook.com' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'guideautomation2@outlook.com' in data:
                        pag.press("enter", interval=1)
                        pag.press("i", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("f2", interval=1)
                        pag.press("v", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'contacts.vcf' in data:
                            pag.press("enter", interval=1)
                            pag.press("f2", interval=1)
                            pag.press("enter", presses=3, interval=1)
                            returnhome()
                            pag.press("a", interval=1)
                            pag.press("enter", presses=2, interval=1)
                            pag.press("e", interval=1)
                            data = readFromQueue(queue)
                            if 'edited user' in data:
                                print('Address 6: Send VCF via email, delete existing instance and then import \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I54'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Address 6: Failed to send VCF via email")
                                print('Address 6: Send VCF via email, delete existing instance and then import \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I54'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Address 6: Contact VCF not attached to email")
                            print("Address 6: contact VCF not attached to email")
                            print(data)
                            ws['I54'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Address 6: Failed to find email address")
                        print("Address 6: Failed to find guideautomation2@outlook.com")
                        print(data)
                        ws['I54'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Address 6: Failed to delete edited user")
                    print("Address 6: Failed to delete edited user")
                    print(data)
                    ws['I54'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Address 6: Failed to send a new email")
                print("Address 6: Failed to send a new email")
                print(data)
                ws['I54'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Address 6: Failed to enter new email")
            print("Address 6: Failed to enter new email")
            print(data)
            ws['I54'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 6: Failed to locate user 'Edited user'")
        print("Address 6: Failed to locate user 'Edited Bloggs'")
        print(data)
        ws['I54'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address7(queue):    # STATUS: TESTED
    # Edit an imported contact details
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("e", interval=1)
    data = readFromQueue(queue)
    if 'edited user' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'GMENU_ADDRESSBOOK_EDIT' in data:
            pag.write("imported", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            data = readFromQueue(queue)
            if 'imported user' in data:
                print('Address 7: Edit details of imported user \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I55'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Address 7: Failed to edit details of imported user")
                print('Address 7: Edit details of imported user \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I55'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Address 7: Failed edit imported contact")
            print("Address 7: Failed to edit imported contact")
            print(data)
            ws['I55'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 7: Failed to find imported contact")
        print("Address 7: failed to find imported contact")
        print(data)
        ws['I55'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address8(queue):    # STATUS: TESTED
    # sort contacts by last name
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("i", interval=1)
    data = readFromQueue(queue)
    if 'imported user' in data:
        pag.press("f2", interval=1)
        pag.press("s", presses=4, interval=1)
        pag.press("enter", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", interval=1)
        pag.press("u", interval=1)
        data = readFromQueue(queue)
        if 'user, imported' in data:
            print('Address 8: sort contacts by Last name \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I56'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 8: Failed to sort contacts by last name")
            print('Address 8: sort contacts by Last name \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I56'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 8: Failed to find imported user")
        print("Address 8: failed to find imported user")
        print(data)
        ws['I56'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Address9(queue):    # STATUS: TESTED
    # Delete multiple address book contacts
    delay()
    start = 0
    pag.press("right", presses=5, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press(" ")
    pag.press("right", interval=1)
    pag.press(" ")
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)
    if 'Delete' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        if 'No contacts' in data:
            print('Address 9: Delete multiple contacts \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I57'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Address 9: Failed to delete multiple contacts")
            print('Address 9: Delete multiple contacts \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I57'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Address 9: Failed to locate delete")
        print("Address 9: failed to locate delete")
        print(data)
        ws['I57'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()
