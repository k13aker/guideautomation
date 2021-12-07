import logging
import pyautogui as pag
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120


def Settings1(queue):  # STATUS: NEEDS TESTING
    # Changes files to save to windows folder and checks its working
    delay()
    start = 0
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("a", interval=1)
    data = readFromQueue(queue)

    while 'Advanced' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Advanced' in data:
        pag.press("enter", interval=1)
        pag.press("s", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Save files to' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Save files to' in data:
            pag.press("enter", interval=1)
            pag.press("w", interval=1)
            data = readFromQueue(queue)

            while 'Windows folders' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Windows folders' in data:
                pag.press("enter", interval=1)
                pag.press("s", presses=2, interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'Windows folders' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Windows folders' in data:
                    returnhome()
                    pag.press("L", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("N", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_DOCUMENTS_VIEW' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_DOCUMENTS_VIEW' in data:
                        pag.write("t", interval=0.5)
                        pag.press("esc", interval=1)
                        pag.press("enter", interval=1)
                        pag.write("Win Folders document", interval=0.5)
                        pag.press("enter", presses=2, interval=1)
                        returnhome()
                        pag.press("t", interval=1)
                        pag.press("enter", interval=1)
                        pag.press("f", interval=1)
                        data = readFromQueue(queue)
                        if 'File explorer' in data:
                            pag.press("enter", presses=3, interval=1)

                            while 'Win Folders document' not in data:
                                data = readFromQueue(queue)
                                pag.press("w", interval=1)

                            if 'Win Folders document' in data:
                                print('Settings 1: Auto Save files to windows folders \n'
                                      '>>> Result: PASS \n'
                                      '>>> Current String:', data)
                                ws['I193'].value = 'PASS'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                            else:
                                logging.critical("Settings 1: Failed to locate file in windows folder")
                                print('Settings 1: Auto Save files to windows folders \n'
                                      '>>> Result: FAIL \n'
                                      '>>> Current String:', data)
                                ws['I193'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Settings 1: Failed to locate file explorer")
                            print("Settings 1: Failed to locate file explorer")
                            print(data)
                            ws['I193'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Settings 1: Failed to launch create new document")
                        print("Settings 1: Failed to launch create new document")
                        print(data)
                        ws['I193'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Settings 1: Failed to change save files to windows folders")
                    print("Settings 1: Failed to change Save files to windows folders")
                    print(data)
                    ws['I193'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Settings 1: Failed to locate windows folders settings")
                print("Settings 1: Failed to locate Windows Folders Setting")
                print(data)
                ws['I193'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Settings 1: Failed to locate save files to settings")
            print("Settings 1: Failed to locate Save Files to Setting")
            print(data)
            ws['I193'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Settings 1: Failed to locate advanced settings")
        print("Settings 1: Failed to locate Advanced settings")
        print(data)
        ws['I193'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Settings2(queue):   # STATUS: NEEDS TESTING
    # Tests Max downloaded podcast limit
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)

    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("d", interval=1)
            data = readFromQueue(queue)

            while 'Desert Island Discs with Kirsty Young' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Desert Island Discs with Kirsty Young' in data:
                pag.press("enter", interval=1)
                pag.sleep(10)
                added = []

                while len(added) < 10:
                    pag.press("f2", interval=1)
                    pag.press("d", presses=2, interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.sleep(2)
                    data = readFromQueue(queue)
                    added.append(data)
                    pag.press("right", interval=1)

                    if len(added) == 10:
                        break

                pag.press("esc", interval=1, presses=2)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'My downloaded podcast episodes' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'My downloaded podcast episodes' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(10)
                    present = []

                    while len(present) < 10:
                        data = readFromQueue(queue)
                        present.append(data)
                        pag.press("right", interval=1)

                        if len(present) == 10:
                            break

                    if added == present:
                        print('Settings 2: Add the correct Podcast limit (10) \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws[''].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Settings 2: Failed to add 10 podcasts to max param 10")
                        print('Settings 2: Add the correct Podcast limit (10) \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws[''].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Settings 2: Failed to locate My downloaded Podcasts")
                    print("Settings 2: Failed to locate My downloaded Podcasts")
                    print(data)
                    ws[''].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Settings 2: Failed to locate Desert Island Discs with Kirsty Young")
                print("Settings 2: Failed to locate Desert Island Discs with Kirsty Young ")
                print(data)
                ws[''].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Settings 2: Failed to locate play a podcasts")
            print("Settings 2: Failed to locate play a podcast")
            print(data)
            ws[''].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Settings 2: Failed to open podcasts")
        print("Settings 2: Failed to open Podcasts")
        print(data)
        ws[''].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Settings3(queue):
    # Exceeds Max download podcast limit
    pass


def Settings4(queue):
    # Increases podcast limit and adds more podcasts
    pass