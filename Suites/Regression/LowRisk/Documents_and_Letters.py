from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120
import logging


def Docs1(queue):  # STATUS: TESTED
    # Can you save a document using ESC and view it recent documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document", interval=0.5)
        pag.press("esc", interval=1)  # saving document
        pag.press("enter", interval=1)
        pag.write("1", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '1' in data:
            pag.press("enter", interval=1)
            pag.sleep(3)
            print('Docs1: Can you save a document using ESC and view it recent documents \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I71'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs1: Can you save a document using ESC and view it in recent documents")
            print('Docs1: Can you save a document using ESC and view it in recent documents \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I71'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 1: Failed to open a new document")
        print("Docs 1: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I71'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs2(queue):  # STATUS: TESTED
    # Can you save a document using F2 save and view it in my documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("this is a test document", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        pag.write("2", interval=0.5)
        pag.press("enter", interval=1, presses=2)
        pag.press("esc", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        pag.press("2", interval=1)
        data = readFromQueue(queue)
        while '2' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if '2' in data:
            pag.press("enter", interval=1)
            pag.sleep(3)
            print('Docs2: Can you save a document using F2 save and view it in my documents \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I72'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs2: Failed to save and view document from my documents")
            print('Docs2: Can you save a document using F2 save and view it in my documents \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I72'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 2: Failed to open a new document")
        print("Docs 2: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I72'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs3(queue):    # STATUS: TESTED
    # Spell checks a document
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.write("thi iz a speling chek", interval=0.5)
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'GMENU_DOCUMENTS_SPELLCHECK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DOCUMENTS_SPELLCHECK' in data:
            pag.press("enter", interval=1, presses=5)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Docs3: Can you spell check a created document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I73'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Docs 3: Failed to sepll check a document")
                print('Docs3: Can you spell check a created document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I73'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Docs 3: Failed to spell check document")
            print("Docs3: Failed to Spell check document \n"
                  ">>> Current String:", data)
            ws['I73'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            pag.sleep(3)
            pag.press("esc", presses=2)
            pag.press("right")
            pag.press("enter")
            returnhome()

    else:
        logging.critical("Docs 3: Failed to open a new document")
        print("Docs3: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I73'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs4(queue):   # STATUS: TESTED
    # Looks up a dictionary word from document
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DOCUMENTS_VIEW' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DOCUMENTS_VIEW' in data:
        pag.press("f2", interval=1)
        pag.press("l", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DICTIONARYMODULE_MAIN' in data:
            pag.write("document", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' in data:
                pag.sleep(5)
                print('Docs4: Can you look up a word in a created document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I74'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs 4: Failed to look up word from created document")
                print('Docs4: Can you look up a word in a created document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I74'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 4: Failed to open look up word")
            print("Docs4: Failed to open look up word \n"
                  ">>> Current String:", data)
            ws['I74'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 4: Failed to open a new document")
        print("Docs4: Navigation failed to open a new document \n"
              ">>> Current String:", data)
        ws['I74'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs5(queue):   # STATUS: TESTED
    # search for a document from my documents
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.write("2", interval=1)
    pag.press("enter", interval=1)
    pag.sleep(3)
    pag.press("right", interval=1)
    data = readFromQueue(queue)
    while '2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '2' in data:
        print('Docs5: Can you search for a specific document from my documents? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I75'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Docs 5: Failed to search for specific document in my documents")
        print('Docs5: Can you search for a specific document from my documents? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I75'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs6(queue):   # STATUS: TESTED
    # Can you rename a document?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("2", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.write("rename", interval=0.5)
    pag.press("enter", interval=2, presses=2)
    pag.press("r", interval=1)
    data = readFromQueue(queue)

    while 'rename2' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'rename2' in data:
        print('Docs6: Can you rename a document from my documents? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I76'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Docs 6: Failed to rename a document from my documents")
        print('Docs6: Can you rename a document from my documents? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I76'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs7(queue):   # STATUS: TESTED
    # Can you delete a document?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("r", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Delete selected document' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Delete selected document' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=2, presses=2)
        pag.sleep(5)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'Go to folder - GuideConnect' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Go to folder - GuideConnect' in data:
            print('Docs7: Can you delete a document from my documents? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I77'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Docs7: Failed to delete a document from my documents")
            print('Docs7: Can you delete a document from my documents? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I77'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs7: Failed to navigate to delete document")
        print('Docs7: Failed to navigate to delete document \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I77'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs8(queue):   # STATUS: TESTED
    # Opens a DOC file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'DOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'DOC' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs8: Can you view a .DOC file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I78'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs8: Failed to view DOC file")
                print('Docs8: Can you view a .DOC file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I78'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs8: Failed to locate DOC file from test data")
            print("Docs 8: failed to locate DOC.doc test file")
            print(data)
            ws['I78'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs8: Failed to find test data folder")
        print("Docs 8: failed to find test data folder")
        print(data)
        ws['I78'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs9(queue):   # STATUS: TESTED
    # Opens a PDF file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'PDF' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'PDF' in data:
            pag.press("enter", interval=1)

            while 'GMENU_DOCUMENTS_PDF_VIEWER' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_PDF_VIEWER' in data:
                pag.sleep(10)
                print('Docs9: Can you view a .PDF file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I79'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs9: Failed to open a PDF file")
                print('Docs9: Can you view a .PDF file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I79'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs9: Failed to locate PDF test file")
            print("Docs 9: failed to locate PDF.pdf test file")
            print(data)
            ws['I79'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs9: Failed to find test data folder")
        print("Docs 9: failed to find test data folder")
        print(data)
        ws['I79'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs10(queue):   # STATUS: TESTED
    # Opens a RTF file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'RTF' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'RTF' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs10: Can you view a .RTF file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I80'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs10: Failed to view a RTF File")
                print('Docs10: Can you view a .RTF file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I80'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs10: Failed to locate RTF test file")
            print("Docs 10: failed to locate RTF.rtf test file")
            print(data)
            ws['I80'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs10: Failed to find test data folder")
        print("Docs 10: failed to find test data folder")
        print(data)
        ws['I80'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs11(queue):  # STATUS: TESTED
    # Opens a DOCX file
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'testdata' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'testdata' in data:
        pag.press("enter", interval=1)
        pag.press("x", interval=1)
        data = readFromQueue(queue)

        while 'XDOC' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'XDOC' in data:
            pag.press("enter", interval=1)
            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_DOCUMENTS_VIEW' in data:
                pag.sleep(5)
                print('Docs11: Can you view a .DOCX file? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I81'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Docs11: Failed to view a DOCX file")
                print('Docs11: Can you view a .DOCX file? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I81'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 11: Failed to locate DOCX test file")
            print("Docs 11: failed to locate XDOC.docx test file")
            print(data)
            ws['I81'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 11: Failed to find test data folder")
        print("Docs 11: failed to find test data folder")
        print(data)
        ws['I81'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs12(queue):  # STATUS: TESTED
    # Opens DOC from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)
    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'DOC' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs12: Can you open a DOC file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I82'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs12: Failed to open DOC file from USB")
                    print('Docs12: Can you open a DOC file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I82'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs12: Failed to DOC file from the USB")
                print("Docs12: Failed to find DOC.doc from the USB")
                print(data)
                ws['I82'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs12: Failed to find correct Regression USB")
            print("Docs12: Failed to find correct Regression USB")
            print(data)
            ws['I82'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs12: Failed to enter open my devices")
        print("Docs12: failed to enter open from device")
        print(data)
        ws['I82'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs13(queue):  # STATUS: TESTED
    # Opens PDF from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("p", interval=1)
            data = readFromQueue(queue)

            while 'PDF' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'PDF' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_PDF_VIEWER' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_PDF_VIEWER' in data:
                    pag.sleep(10)
                    print('Docs13: Can you open a PDF file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I83'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs13: Failed to open PDF from USB")
                    print('Docs13: Can you open a PDF file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I83'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 13: Failed to find PDF on USB")
                print("Docs 13: Failed to find PDF.pdf from USB")
                print(data)
                ws['I83'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 13: Failed to find correct Regression USB")
            print("Docs 13: Failed to find correct Regression USB")
            print(data)
            ws['I83'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 13: Failed to enter open from device")
        print("Docs 13: failed to enter open from device")
        print(data)
        ws['I83'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs14(queue):  # STATUS: TESTED
    # Opens RTF from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'RTF' in data:
                pag.press("enter", interval=1)

                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs14: Can you open a RTF file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I84'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs14: Failed to find RTF file from USB")
                    print('Docs14: Can you open a RTF file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I84'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 14: Failed to find RTF from USB")
                print("Docs 14: Failed to find RTF.rtf from USB")
                print(data)
                ws['I84'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 14: Failed to find correct Regression USB")
            print("Docs 14: Failed to find correct Regression USB")
            print(data)
            ws['I84'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 14: Failed to enter open from device")
        print("Docs 14: failed to enter open from device")
        print(data)
        ws['I84'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Docs15(queue):  # STATUS: TESTED
    # Opens DOCX from USB
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("o", interval=1)
    data = readFromQueue(queue)

    while 'Open from device' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Open from device' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]} [Regression]" in data:
            pag.press("enter", interval=1, presses=2)
            pag.press("x", interval=1)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)
                while 'GMENU_DOCUMENTS_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_DOCUMENTS_VIEW' in data:
                    pag.sleep(5)
                    print('Docs15: Can you open a DOCX file from a USB? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I85'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Docs 15: Failed to open DOCX from USB")
                    print('Docs15: Can you open a DOCX file from a USB? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I85'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Docs 15: Failed to find DOCX from USB")
                print("Docs15: Failed to find XDOC.docx from USB")
                print(data)
                ws['I85'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Docs 15: Failed to find correct Regression USB")
            print("Docs15: Failed to find correct Regression USB")
            print(data)
            ws['I85'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Docs 15: Failed to enter open from device")
        print("Docs15: failed to enter open from device")
        print(data)
        ws['I85'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter1(queue):  # STATUS: TESTED
    # Can you manually enter your sender address send a letter to a address book contact?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_COREMODULE_SETTINGS_DOCUMENT_ADDRESS_SENDER_FIRST_NAME' in data:
        pag.write("joe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bloggs", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("22", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("dolphin street", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("bromsgrove", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("worcestershire", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("b45 9xe", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("united kingdom", interval=0.5)
        pag.press("enter", presses=4, interval=0.5)
        pag.write("this is a test letter", interval=0.5)
        pag.press("esc", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("letter 1", interval=0.5)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=0.5)
        pag.press("enter", interval=0.5)
        pag.write("L", interval=0.5)
        data = readFromQueue(queue)
        if 'letter 1' in data:
            pag.press("enter", interval=1)
            pag.sleep(2)
            print('Letter1: Can you manually enter your sender address and'
                  ' send a letter to a address book contact? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I86'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Letter 1: Failed to add manual sender address and compose a letter")
            print('Letter1: Can you manually enter your sender address and'
                  ' send a letter to a address book contact? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I86'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letter 1: Failed to open add senders address")
        print("Letter 1: failed to open add senders address")
        print(data)
        ws['I86'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter2(queue):  # STATUS: TESTED
    # Can you create a letter using manually enter address
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("t", interval=1)
    data = readFromQueue(queue)

    while 'Type address manually' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Type address manually' in data:
        pag.press("enter", interval=1)

        while 'GMENU_DOCUMENTS_RECIPIENTS_ADDRESS_FIRST_NAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DOCUMENTS_RECIPIENTS_ADDRESS_FIRST_NAME' in data:
            pag.write("joe", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("bloggs", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("dolphin", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("22", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("dolphin street", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("bromsgrove", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("b45 9xe", interval=0.5)
            pag.press("enter", interval=0.5)
            pag.write("united kingdom", interval=0.5)
            pag.press("enter", interval=0.5, presses=2)
            pag.write("this is a letter 2", interval=0.5)
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)
            pag.write("manual", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("m", interval=1)
            pag.press("enter", interval=1)
            pag.press("m", interval=1)
            data = readFromQueue(queue)

            while 'manual' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'manual' in data:
                print('Letter2: Can you manually enter a letter address? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I87'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 2: Failed to manually enter a letter address")
                print('Letter2: Can you manually enter a letter address? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I87'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letter 2: Failed to enter type manual address")
            print("Letters 2: failed to enter type manual address")
            print(data)
            ws['I87'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letters 2: Failed to navigate to type manual address")
        print("Letters 2: failed to navigate to type manual address")
        print(data)
        ws['I87'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter3(queue):  # STATUS: TESTED
    # can you change the sender address via f2 menu?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)

    while 'Change address' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Change address' in data:
        pag.press("enter", interval=1, presses=3)
        data = readFromQueue(queue)

        while 'Dolphin Computer Access Ltd.' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Dolphin Computer Access Ltd.' in data:
            pag.press("enter", interval=2, presses=2)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Letter3: Can you change the sender address via f2 menu? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I88'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 3: Failed to change sender address via F2 menu")
                print('Letter3: Can you change the sender address via f2 menu? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I88'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letters 3: Failed to find dolphin in address book")
            print("Letters 3: Failed to find Dolphin in address book")
            print(data)
            ws['I88'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letters 3: Failed to navigate to change address")
        print("Letters 3: Failed to navigate to change address")
        print(data)
        ws['I88'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def letter4(queue):  # STATUS: TESTED
    # can you change the recipient address via f2 menu?
    start = 0
    delay()
    pag.press("right", presses=1, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("c", interval=1)
    data = readFromQueue(queue)

    while 'Change address' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Change address' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1, presses=2)
        data = readFromQueue(queue)

        while 'Dolphin Computer Access Ltd.' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Dolphin Computer Access Ltd.' in data:
            pag.press("enter", interval=2, presses=2)
            data = readFromQueue(queue)

            while 'GMENU_DOCUMENTS_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_DOCUMENTS_VIEW' in data:
                print('Letter4: Can you change the recipient address via f2 menu? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I89'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Letter 4: Failed to change recipients address via F2 menu")
                print('Letter4: Can you change the recipient address via f2 menu? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I89'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Letter 4: Failed to find dolphin in address book")
            print("Letter4: Failed to find Dolphin in address book")
            print(data)
            ws['I89'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Letter 4: Failed to navigate to change address")
        print("Letter4: Failed to navigate to change address")
        print(data)
        ws['I89'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()
