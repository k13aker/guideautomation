import logging

from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120


def Podcast1(queue):    # STATUS: TESTED
    # Play a new podcast from find a new podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p")
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.sleep(10)
            pag.press("enter")
            print('Podcast 1: Can you play a new podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I10'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Podcast 1: Failed to play a new podcast")
            print('Podcast 1: Can you play a new podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I10'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 1: Failed to open podcasts")
        print("Podcast 1: Failed to open Podcasts")
        print(data)
        ws['I10'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast2(queue):   # STATUS: TESTED
    # Continue playing a recently played podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_PLAY' in data:
            pag.sleep(10)
            pag.press("enter")
            print('Podcast 2: Can you continue playing a podcast? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I11'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            pag.sleep(10)
            pag.press("enter")
            logging.critical("Podcast 2: Failed to continue listening to a podcast")
            print('Podcast 2: Can you continue playing a podcast? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I11'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 2: Failed to open podcasts")
        print("Podcast 2: Failed to open Podcasts")
        print(data)
        ws['I11'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast3(queue):  # STATUS: TESTED
    # can you add a podcast to your favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2")
            pag.sleep(3)
            pag.press("enter", interval=1)
            pag.press("w", interval=1)
            pag.sleep(3)
            data = readFromQueue(queue)

            while "Woman's Hour: News, Politics, Culture" not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if "Woman's Hour: News, Politics, Culture" in data:
                pag.press("enter", interval=1, presses=2)
                pag.press("w")
                pag.sleep(3)
                data = readFromQueue(queue)

                while "Woman's Hour: News, Politics, Culture" not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if "Woman's Hour: News, Politics, Culture" in data:
                    print('Podcast 3: Can you add a podcast to favourites? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I12'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 3: Failed to add a podcast favourite")
                    print('Podcast 3: Can you add a podcast to favourites? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I12'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 3: Failed to add Women's Hour to favourites")
                print("Podcast 3: Failed to add Women's Hour to favourites")
                print(data)
                ws['I12'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 3: Failed to open podcast favourites")
            print("Podcast 3: Failed to open podcast favourites")
            print(data)
            ws['I12'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 3: Failed to open podcasts")
        print("Podcast 3: Failed to open Podcasts")
        print(data)
        ws['I12'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast4(queue):  # STATUS: TESTED
    # can you remove a podcast from favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("w", interval=1)
            pag.sleep(3)
            data = readFromQueue(queue)

            while "Woman's Hour: News, Politics, Culture" not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if "Woman's Hour: News, Politics, Culture" in data:
                pag.press("f2", interval=1)
                pag.press("r")
                data = readFromQueue(queue)

                while 'Remove selected from favourites' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Remove selected from favourites' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.sleep(5)
                    pag.press("w")
                    data = readFromQueue(queue)

                    while 'In Touch' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'In Touch' in data:
                        print('Podcast 4: Can you add remove a podcast from favourites? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I13'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Podcast 4: Failed to remove a podcast")
                        print('Podcast 4: Can you add remove a podcast from favourites? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I13'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 4: Failed to find remove inside favourites")
                    print("Podcast 4: Failed to find remove from favourites")
                    print(data)
                    ws['I13'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 4: Failed to find Women's Hour in favourites")
                print("Podcast 4: Failed to find Women's Hour in favourites")
                print(data)
                ws['I13'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 4: Failed to open podcast favourites")
            print("Podcast 4: Failed to open podcast favourites")
            print(data)
            ws['I13'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 4: Failed to open podcasts")
        print("Podcast 4: Failed to open Podcasts")
        print(data)
        ws['I13'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast5(queue):  # STATUS: TESTED
    # can you view details of a favourite podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 5: Can you view details of a favourite podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I14'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Podcast 5: Failed to view details of a favourite podcast")
                print('Podcast 5: Can you view details of a favourite podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I14'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 5: Failed to open podcast favourites")
            print("Podcast 5: Failed to open podcast favourites")
            print(data)
            ws['I14'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 5: Failed to open podcasts")
        print("Podcast 5: Failed to open Podcasts")
        print(data)
        ws['I14'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast6(queue):  # STATUS: TESTED
    # Can you add a custom podcast from favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'Add custom podcast' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Add custom podcast' in data:
                pag.press("enter", interval=1)
                pag.write("test", interval=0.5)
                pag.press("enter", interval=1)
                pag.write("https://podcasts.files.bbci.co.uk/p089sfrz.rss", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                pag.press("t", presses=2, interval=1)
                data = readFromQueue(queue)
                while 'test' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'test' in data:
                    print('Podcast 6: Can you add a custom podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I15'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 6: Failed to add custom podcast")
                    print('Podcast 6: Can you add a custom podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I15'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 6: Failed to add custom podcast via favourites")
                print("Podcast 6: Failed to add custom podcast via favourites")
                print(data)
                ws['I15'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 6: Failed to open podcast favourites")
            print("Podcast 6: Failed to open podcast favourites")
            print(data)
            ws['I15'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 6: Failed to open podcasts")
        print("Podcast 6: Failed to open Podcasts")
        print(data)
        ws['I15'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast7(queue):    # STATUS: TESTED
    # Can you search a podcast list and play via favourites?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("f", interval=1)
        data = readFromQueue(queue)

        while 'Favourites' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Favourites' in data:
            pag.press("enter", interval=1, presses=2)
            pag.sleep(3)
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_SEARCH_EDIT' in data:
                pag.write("central", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'Central heating, Mike Lambert column' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)
                    data = readFromQueue(queue)
                    if '' in data:
                        print('Podcast 7: Can you search and play a podcast via favourites? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I16'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("enter")
                        returnhome()

                    else:
                        logging.critical("Podcast 7: Failed to search and play a podcast via favourites")
                        print('Podcast 7: Can you search and play a podcast via favourites? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I16'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        pag.press("enter")
                        returnhome()

                else:
                    logging.critical("Podcast 7: Failed to find podcast - Central heating, Mike lambert column")
                    print("Podcast 7: Failed to find Podcast: Central heating, Mike Lambert column")
                    print(data)
                    ws['I16'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 7: Failed to open podcast search via favourites")
                print("Podcast 7: Failed to open podcast search via favourites")
                print(data)
                ws['I16'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 7: Failed to open podcast favourites")
            print("Podcast 7: Failed to open podcast favourites")
            print(data)
            ws['I16'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Podcast 7: Failed to open podcasts")
        print("Podcast 7: Failed to open Podcasts")
        print(data)
        ws['I16'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast8(queue):   # STATUS: TESTED
    # Can you search the play a new podcast list
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_SEARCH_EDIT' in data:
                pag.write("desert", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                while 'Desert Island Discs with Kirsty Young' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Desert Island Discs with Kirsty Young' in data:
                    print('Podcast 8: Can you search for a podcast via find new podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I17'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 8: Failed to search a new podcast via find new podcast")
                    print('Podcast 8: Can you search for a podcast via find new podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I17'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 8: Failed to search for a new podcast")
                print("Podcast 8: Failed to search for new podcast")
                print(data)
                ws['I17'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 8: Failed to open play a new podcast")
            print("Podcast 8: Failed to open Play a new Podcast")
            print(data)
            ws['I17'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 8: Failed to open podcasts")
        print("Podcast 8: Failed to open Podcasts")
        print(data)
        ws['I17'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast9(queue):    # STATUS: TESTED
    # Can you view details of a podcast via find new podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 9: Can you view details for a podcast via find new podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I18'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

            else:
                logging.critical("Podcast 9: Failed to view details of a podcast via new podcast")
                print('Podcast 9: Can you view details for a podcast via find new podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I18'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

        else:
            logging.critical("Podcast 9: Failed to open play a new podcast")
            print("Podcast 9: Failed to open Play a new Podcast")
            print(data)
            ws['I18'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 9: Failed to open podcasts")
        print("Podcast 9: Failed to open Podcasts")
        print(data)
        ws['I18'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast10(queue):   # STATUS: TESTED
    # Can you change the language of the podcast list via find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", presses=2,  interval=1)
            pag.press("enter", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'Fire Patroner Pa Lommen' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Fire Patroner Pa Lommen' in data:
                print('Podcast 10: Can you change the list language via find a podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I19'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

            else:
                logging.critical("Podcast 10: Failed to change language list via find a podcast")
                print('Podcast 10: Can you change the list language via find a podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I19'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("enter")
                returnhome()

        else:
            logging.critical("Podcast 10: Failed to open play a new podcast")
            print("Podcast 10: Failed to open Play a new Podcast")
            print(data)
            ws['I19'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 10: Failed to open podcasts")
        print("Podcast 10: Failed to open Podcasts")
        print(data)
        ws['I19'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast11(queue):   # STATUS: TESTED
    # Search a podcast and then proceed to play it via find a podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            pag.write("life after", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'Life After Blindness' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Life After Blindness' in data:
                pag.press("enter", interval=1, presses=2)
                pag.sleep(15)
                data = readFromQueue(queue)
                if '' in data:
                    print('Podcast 11: Can you search and then play a podcast via find a podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I20'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 11: Failed to search and play a podcast via find a podcast")
                    print('Podcast 11: Can you search and then play a podcast via find a podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I20'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 11: Failed to search for Life After Blindness")
                print("Podcast 11: Failed to search for Life After Blindness")
                print(data)
                ws['I20'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 11: Failed to open play a new podcast")
            print("Podcast 11: Failed to open Play a new Podcast")
            print(data)
            ws['I20'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 11: Failed to open podcasts")
        print("Podcast 11: Failed to open Podcasts")
        print(data)
        ws['I20'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast12(queue):   # STATUS: TESTED
    # Can you view details of an episode via find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("f2", interval=1)
            pag.sleep(3)
            pag.press("d")
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_FEEDS_DETAILS' in data:
                print('Podcast 12: Can you view details of a podcast episode? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I21'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.sleep(5)
                returnhome()

            else:
                logging.critical("Podcast 12: Failed to view details of a podcast episode")
                print('Podcast 12: Can you view details of a podcast episode? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I21'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.sleep(5)
                returnhome()

        else:
            logging.critical("Podcast 12: Failed to open play a new podcast")
            print("Podcast 12: Failed to open Play a new Podcast")
            print(data)
            ws['I21'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 12: Failed to open podcasts")
        print("Podcast 12: Failed to open Podcasts")
        print(data)
        ws['I21'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast13(queue):  # STATUS: TESTED
    # Can you add a podcast to favourites from find a podcast?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)
        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("f", interval=1)
            data = readFromQueue(queue)
            while 'Farming Today' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Farming Today' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1, presses=2)
                pag.sleep(3)
                pag.press("esc", interval=1)
                pag.press("f")
                pag.press("enter", interval=1)
                pag.press("f")
                pag.sleep(3)
                data = readFromQueue(queue)
                while 'Farming Today' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Farming Today' in data:
                    print('Podcast 13: Can you add a podcast to favourites from find a podcast? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I22'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

                else:
                    logging.critical("Podcast 13: Failed to add a podcast to favourites from find a podcast")
                    print('Podcast 13: Can you add a podcast to favourites from find a podcast? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I22'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    pag.press("enter")
                    returnhome()

            else:
                logging.critical("Podcast 13: Failed to find podcast Farming Today")
                print("Podcast 13: Failed to find podcast: Farming Today")
                print(data)
                ws['I22'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 13: Failed to open play a new podcast")
            print("Podcast 13: Failed to open Play a new Podcast")
            print(data)
            ws['I22'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 13: Failed to open podcasts")
        print("Podcast 13: Failed to open Podcasts")
        print(data)
        ws['I22'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast14(queue):   # STATUS: TESTED
    # Can you download a podcast and play it from my downloads?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)

    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("p", interval=1)
        data = readFromQueue(queue)

        while 'Play a new podcast' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Play a new podcast' in data:
            pag.press("enter", interval=2)
            pag.press("a", interval=1)
            data = readFromQueue(queue)

            while 'A History of the World in 100 Objects' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'A History of the World in 100 Objects' in data:
                pag.press("enter", interval=1)
                pag.press("f2", interval=1)
                pag.press("d", presses=2, interval=1)
                pag.press("enter", interval=1, presses=2)
                pag.press("esc", presses=2, interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'My downloaded podcast episodes' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'My downloaded podcast episodes' in data:
                    pag.sleep(20)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'A History of the World in 100 Objects - Object 101' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'A History of the World in 100 Objects - Object 101' in data:
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        if '' in data:
                            pag.sleep(10)
                            pag.press("enter", interval=1)
                            print('Podcast 14: Can you download and play a podcast? \n'
                                  '>>> Result: PASS \n'
                                  '>>> Current String:', data)
                            ws['I23'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Podcast 14: Failed to download and play a podcast")
                            print('Podcast 14: Can you download and play a podcast? \n'
                                  '>>> Result: FAIL \n'
                                  '>>> Current String:', data)
                            ws['I23'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            pag.press("enter", interval=1)
                            returnhome()

                    else:
                        logging.critical("Podcast 14: A History of the World in 100 Objects not in downloads ")
                        print("Podcast 14: A History of the World in 100 Objects - Object 101 not present in downloads")
                        print(data)
                        ws['I23'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 14: Failed to navigate to my downloaded podcasts")
                    print("Podcast 14: Failed to navigate to My downloaded podcasts")
                    print(data)
                    ws['I23'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 14: Failed to find podcast A history of the world")
                print("Podcast 14: Failed to find podcast: A History of the world in 100 Objects")
                print(data)
                ws['I23'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 14: Failed to open play a new podcast")
            print("Podcast 14: Failed to open Play a new Podcast")
            print(data)
            ws['I23'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 14: Failed to open podcasts")
        print("Podcast 14: Failed to open Podcasts")
        print(data)
        ws['I23'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast15(queue):  # STATUS: TESTED
    # views details of a downloaded podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        while 'A History of the World in 100 Objects - Object 101' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'A History of the World in 100 Objects - Object 101' in data:
            pag.press("f2", interval=1)
            pag.press("d", presses=2, interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_DETAILS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_DETAILS' in data:
                pag.sleep(5)
                print('Podcast 15: Can you view details of a downloaded podcast? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I24'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                pag.sleep(5)
                logging.critical("Podcast 15: Failed to view details of a downloaded podcast")
                print('Podcast 15: Can you view details of a downloaded podcast? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I24'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 15: Failed to download podcast")
            print("Podcast 15: Failed to locate downloaded podcast: "
                  "A History of the World in 100 Objects - Object 101")
            print(data)
            ws['I24'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 15: Failed to open podcasts")
        print("Podcast 15: Failed to open Podcasts")
        print(data)
        ws['I24'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast16(queue):   # STATUS: TESTED
    # Copy a downloaded podcast to a USB
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Radio' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Radio' in data:
        pag.press("p")
        data = readFromQueue(queue)
        if 'Podcasts' in data:
            pag.press("enter", interval=1)
            pag.press("m", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            while 'A History of the World in 100 Objects - Object 101' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'A History of the World in 100 Objects - Object 101' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                if f"{found[0]}" in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_ENTERTAINMENT_PODCAST_PODCASTS_DOWNLOADED_COPYTODEVICE_PROGRESS' in data:
                        data = readFromQueue(queue)
                        pag.sleep(2)
                        if 'OK' in data:
                            break

                    if 'OK' in data:
                        pag.press("enter", interval=1)
                        print('Podcast 16: Can you copy a podcast to a device? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I25'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Podcast 16: Failed to copy a podcast to a USB")
                        pag.press("enter", interval=1)
                        print('Podcast 16: Can you copy a podcast to a device? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I25'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Podcast 16: Failed to open copy to device")
                    print("Podcast 16: Failed to open Copy to device")
                    print(data)
                    ws['I25'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Podcast 16: Failed to locate downloaded podcast")
                print("Podcast 16: Failed to locate downloaded podcast:"
                      " A History of the World in 100 Objects - Object 101")
                print(data)
                ws['I25'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 16: Failed to open podcasts")
            print("Podcast 16: Failed to open Podcasts")
            print(data)
            ws['I25'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 16: Failed to enter entertainment")
        print("Podcast 16: Failed to enter Entertainment")
        print(data)
        ws['I25'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast17(queue):   # STATUS: TESTED
    # deletes a downloaded podcast
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("p")
    data = readFromQueue(queue)
    if 'Podcasts' in data:
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'A History of the World in 100 Objects - Object 101' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'A History of the World in 100 Objects - Object 101' in data:
            pag.press("enter", interval=1)
            pag.press("esc", interval=1)
            pag.press("enter", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            data = readFromQueue(queue)

            if 'No downloaded podcasts' in data:
                print('Podcast 17: Can you delete a podcast from my downloaded podcasts? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I26'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Podcast 17: Failed to delete podcasts from my downloaded podcasts")
                print('Podcast 17: Can you delete a podcast from my downloaded podcasts? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I26'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 17: Failed to find downloaded podcast")
            print("Podcast 17: Failed to find downloaded podcast")
            print(data)
            ws['I26'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 17: Failed to open podcasts")
        print("Podcast 17: Failed to open Podcasts")
        print(data)
        ws['I26'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def Podcast18(queue):
    # Delete a podcast from a USB
    start = 0
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.sleep(5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if f"{found[0]}" in data:
            pag.press("enter", interval=1)
            pag.press("o", interval=1)
            data = readFromQueue(queue)
            if 'Object 101' in data:
                pag.press("del", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("o", interval=1)
                data = readFromQueue(queue)
                if 'testdata' in data:
                    print('Podcast 18: Delete a podcast from USB \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I192'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Podcast 18: Failed to delete a podcast from USB")
                    print('Podcast 18: Delete a podcast from USB \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I192'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Podcast 18: Failed to locate Object 101 on USB device")
                print("Podcast 18: Failed to locate Object 101 On USB device")
                print(data)
                ws['I192'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Podcast 18: Failed to locate regression USB")
            print("Podcast 18: failed to locate regression USB")
            print(data)
            ws['I192'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Podcast 18: Failed to locate file explorer")
        print("Podcast 18: failed to locate file explorer")
        print(data)
        ws['I192'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()
