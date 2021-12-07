import logging

import pyautogui as pag
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120


def scanner_text1(queue):  # STATUS: TESTED
    # Can you select a scanner to use?
    start = 0
    delay()
    pag.press("right", presses=10, interval=1)
    pag.press("enter", interval=1)
    pag.press("s", interval=1)
    data = readFromQueue(queue)

    while 'Scanner' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scanner' in data:
        pag.press("enter", interval=1, presses=2)
        pag.sleep(5)
        data = readFromQueue(queue)
        if 'No devices found.  Please make sure that your scanner' \
           ' is connected to the computer and switched on.' in data:
            logging.critical("Scanner 1: Failed to find scanner attached to system")
            print("Scanner 1: No Scanner attached to the system")
            print(data)
            ws['I143'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        elif 'No devices found.  Please make sure that your scanner' \
             ' is connected to the computer and switched on.' not in data:
            data = readFromQueue(queue)
            scanner = data
            if scanner == data:
                pag.press("enter", interval=1)
                print(f"Scanner found: {scanner}")
                print("Scanner 1: Can you select your Scanner \n"
                      "Result: PASS \n"
                      "Current String:", data)
                ws['I143'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 1: Failed to find your attached scanner")
                print("Scanner 1: Can you select your Scanner \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I143'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 1: Failed to select a scanner")
            print("Scanner 1: Can you select your Scanner \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I143'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 1: Failed to navigate to scanner settings")
        print("Scanner 1: failed to navigate to Scanner settings")
        print(data)
        ws['I143'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text2(queue):    # STATUS: TESTED
    # can you scan for text (1 page)
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)
        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'Read Scanned Pages' in data:
                break

        data = readFromQueue(queue)
        if "Read Scanned Pages" in data:
            pag.sleep(2)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_SCANNER_OCR_PROGRESS' in data:
                data = readFromQueue(queue)
                if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                    break
            pag.sleep(5)

            print('Scanner Text 2: Can you scan 1 page for text? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I144'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
            pag.press("esc", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1)
            returnhome()

        else:
            logging.critical("Scanner Text 2: Failed to scan a single page for text")
            print('Scanner Text 2: Can you scan 1 page for text? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I144'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Scanner Text 2: Failed to locate scan for text")
        print("Scanner Text 2: Failed to locate Scan for text")
        print(data)
        ws['I144'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text3(queue):   # STATUS: TESTED
    # Can you scan for text (2 pages)
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_SCANNER_SCAN_TEXT_PROGRESS' in data:
            data = readFromQueue(queue)
            pag.sleep(3)
            if 'Read Scanned Pages' in data:
                pag.press("right", interval=1)
                data = readFromQueue(queue)
                break

            elif 'GMENU_SCANNER_SCAN_TEXT_COMPLETE' in data:
                pag.press("right", interval=1)
                data = readFromQueue(queue)
                break

        if "Scan Another Page" in data:
            pag.press("enter", interval=1)
            while 'Read Scanned Pages' not in data:
                data = readFromQueue(queue)

            if 'Read Scanned Pages' in data:
                pag.sleep(3)
                pag.press("enter", interval=1)

                while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                    data = readFromQueue(queue)

                if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                    pag.sleep(15)
                    print('Scanner Text 3: Can you scan 2 pages for text? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I145'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
                    pag.press("esc", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", interval=1)
                    returnhome()

                else:
                    logging.critical("Scanner Text 3: Failed to scan Multiple pages for text")
                    print('Scanner Text 3: Can you scan 2 pages for text? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I145'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Scanner Text 3: Failed to locate read scanned pages")
                print("Scanner Text 3: Failed to locate Read Scanned Pages")
                print(data)
                ws['I145'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Scanner Text 3: Failed to locate read scanned pages")
            print("Scanner Text 3: Failed to locate Read Scanned Pages")
            print(data)
            ws['I145'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Scanner Text 3: Failed to locate scan for text")
        print("Scanner Text 3: Failed to locate Scan for text")
        print(data)
        ws['I145'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text4(queue):    # STATUS: TESTED
    # Can you scan a single document after scanning multiple?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Read Scanned Pages' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if "Read Scanned Pages" in data:
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                pag.sleep(10)
                print('Scanner Text 4: Can you scan 1 page for text after scanning multiple? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I146'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Scanner Text 4: Failed to scan a single page after multiple pages")
                print('Scanner Text 4: Can you scan 1 page for text after scanning multiple?  \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I146'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner Text 4: Failed to locate read scanned pages")
            print("Scanner Text 4: failed to locate read scanned pages")
            print(data)
            ws['I146'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner Text 4: Failed to locate scan for text")
        print("Scanner Text 4: Failed to locate scan for text")
        print(data)
        ws['I146'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner_text5(queue):    # STATUS: TESTED
    # Can you save a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Scan for text' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for text' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'Read Scanned Pages' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if "Read Scanned Pages" in data:
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                pag.press("f2", interval=1)
                pag.press("enter", interval=1)
                pag.write("5scanner", interval=0.5)
                pag.press("enter", interval=1, presses=2)
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                pag.press("esc", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'My scanned documents' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'My scanned documents' in data:
                    pag.press("enter", interval=1)
                    pag.press("5", interval=0.5)
                    data = readFromQueue(queue)

                    while '5scanner' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if '5scanner' in data:
                        print('Scanner 5: Can you save a scanned document? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I147'].value = 'PASS'
                        wb.save('Test Reports/Automation Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Scanner Text 5: Failed to save a scanned document")
                        print('Scanner 5: Can you save a scanned document? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I147'].value = 'FAIL'
                        wb.save('Test Reports/Automation Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Scanner Text 5: Failed to navigate to my scanned documents")
                    print("Scanner 5: Failed to navigate to my scanned documents")
                    print(data)
                    ws['I147'].value = 'FAIL'
                    wb.save('Test Reports/Automation Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Scanner Text 5: Failed to open scanned document")
                print("Scanner 5: Failed to open scanned document")
                print(data)
                ws['I147'].value = 'FAIL'
                wb.save('Test Reports/Automation Test cases.xlsx')
                returnhome()
        else:
            logging.critical("Scanner Text 5: Failed to find read scanned pages")
            print("Scanner 5: Failed to find Read scanned pages")
            print(data)
            ws['I147'].value = 'FAIL'
            wb.save('Test Reports/Automation Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner Text 5: Failed to locate scan for text ")
        print("Scanner 5: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I147'].value = 'FAIL'
        wb.save('Test Reports/Automation Test cases.xlsx')
        returnhome()


def scan_image1(queue):  # STATUS: TESTED
    # Can you scan for an image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            print('Scan Image 1: Can you scan for an image? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I148'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Scan Image 1: Failed to scan for an image")
            print('Scan Image 1: Can you scan for an image? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I148'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 1: Failed to locate scan for image")
        print("Scan image 1: Failed to locate scan for image")
        print(data)
        ws['I148'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scan_image2(queue):   # STATUS: TESTED
    # can you save a scanned image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("s", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.write("scanned image", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("esc", presses=2)
            pag.sleep(3)
            pag.press("m", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("s", interval=1)
            data = readFromQueue(queue)
            if 'scanned image' in data:
                print('Scan Image 2: Can you save a scanned image? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I149'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scan Image 2: Failed to save scanned image")
                print('Scan Image 2: Can you save a scanned image? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I149'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scan Image 2: Failed to load scanned image")
            print("Scan image 2: Failed to Succesfully load scanned image")
            print(data)
            ws['I149'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 2: Failed to locate Scan for image")
        print("Scan image 2: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I149'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scan_image3(queue):  # STATUS: TESTED
    # can you scan for text inside of a scanned image?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("right", interval=1)
    data = readFromQueue(queue)

    while 'Scan for image' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Scan for image' in data:
        pag.sleep(2)
        pag.press("enter", presses=2, interval=1)

        while "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" not in data:
            data = readFromQueue(queue)

        if "GMENU_SCANNER_SCAN_IMAGE_COMPLETE" in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)

            while 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_OCR_DOCUMENT_VIEW' in data:
                print('Scan Image 3: Can you scan for text in a scanned image? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I150'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

            else:
                logging.critical("Scan Image 3: Failed to scan for text inside a scanned image")
                print('Scan Image 3: Can you scan for text in a scanned image? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I150'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                pag.press("esc", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", interval=1)
                returnhome()

        else:
            logging.critical("Scan Image 3: Failed to load scanned image")
            print("Scan image 3: Failed to Succesfully load scanned image")
            print(data)
            ws['I150'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scan Image 3: Failed to locate scan for image")
        print("Scan image 3: Failed to enter the scanner module, check scanner is attached")
        print(data)
        ws['I150'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner9(queue):  # STATUS: TESTED
    # Can you rename a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)
    if 'My scanned documents' in data:
        pag.press("enter", interval=1)
        pag.press("5", interval=1)
        data = readFromQueue(queue)

        if '5scanner' in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'renamed 5scanner' in data:
                print('Scanner9: Can you rename a saved scanned document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I151'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 9: Failed to rename a scanned document ")
                print('Scanner9: Can you rename a saved scanned document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I151'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 9: Failed to find a saved scanned document")
            print("Scanner9: Failed to find a saved scanned document")
            print(data)
            ws['I151'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 9: Failed to locate my scanned documents")
        print("Scanner 9: failed to locate my scanned documents")
        print(data)
        ws['I151'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def scanner10(queue):  # STATUS: TESTED
    # Can you delete a scanned document?
    start = 0
    delay()
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)
    if 'My scanned documents' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'renamed 5scanner' in data:
            pag.sleep(5)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            if 'Go to folder - GuideConnect' in data:
                print('Scanner 10: Can you delete a saved scanned document? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I152'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Scanner 10: Failed to delete a saved scanned document")
                print('Scanner 10: Can you delete a saved scanned document? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I152'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Scanner 10: Failed to find a saved scanned document")
            print("Scanner 10: Failed to find a saved scanned document")
            print(data)
            ws['I152'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Scanner 10: Failed to locate my scanned documents")
        print("Scanner 10: failed to locate my scanned documents")
        print(data)
        ws['I152'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera1(queue):  # STATUS: TESTED
    # Can you capture an image, save and then view it?
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(5)
    pag.press("c", interval=1)
    data = readFromQueue(queue)
    if 'Camera' in data:
        pag.press("enter", interval=1)

        while 'GMENU_SHOW_CAMERA_FEED' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_SHOW_CAMERA_FEED' in data:
            pag.press("enter", interval=1)
            start = 0

            while 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_SCANNER_SCAN_IMAGE_COMPLETE' in data:
                pag.sleep(5)
                pag.press("f2", interval=1)
                pag.press("s", presses=2, interval=1)  # save image
                pag.press("enter", interval=1)
                pag.write("camera test", interval=0.5)
                pag.press("enter", presses=2, interval=1)
                pag.press("esc", presses=2, interval=1)
                pag.press("m", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'My pictures' in data:
                    pag.press("enter", interval=1)
                    pag.press("c", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                        print('Camera 1: Can you capture an image, save and view it? \n'
                              '>>> Result: PASS \n'
                              '>>> Current String:', data)
                        ws['I153'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Camera 1: Failed to capture, save and view an image")
                        print('Camera 1: Can you capture an image, save and view it? \n'
                              '>>> Result: FAIL \n'
                              '>>> Current String:', data)
                        ws['I153'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("Camera 1: Failed to locate my pictures")
                    print("Camera 1: Failed to locate My pictures")
                    print(data)
                    ws['I153'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Camera 1: Failed to load captured image ")
                print("Camera 1: Failed to load GMENU_SCANNER_SCAN_IMAGE_COMPLETE")
                print("Current String:", data)
                ws['I153'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Camera 1: Failed to load camera feed")
            print("Camera 1: Camera failed to load GMENU_SHOW_CAMERA_FEED")
            print("Current String:", data)
            ws['I153'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("Camera 1: Failed to locate camera")
        print("Camera 1: failed to locate camera")
        print(data)
        ws['I153'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera2(queue):  # STATUS: TESTED
    # Can you rename a saved image?
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("m", interval=1, presses=2)
    data = readFromQueue(queue)
    if 'My pictures' in data:
        pag.press("enter", interval=1)
        pag.press("c", interval=1)
        data = readFromQueue(queue)
        if 'camera test' in data:
            pag.press("f2", interval=1)
            pag.press("enter", interval=1)
            pag.write("renamed ", interval=0.5)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'renamed camera test' in data:
                print('Camera 2: Can you rename a saved camera picture? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I154'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Camera 2: Failed to rename a saved captured image")
                print('Camera 2: Can you rename a saved camera picture? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I154'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.exception("Camera 2: Failed to locate saved picture")
            print("Camera 2: failed to locate saved picture")
            print("Current String:", data)
            ws['I154'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Camera 2: Failed to locate my pictures")
        print("Camera 2: failed to locate my pictures")
        print(data)
        ws['I154'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def camera3(queue):     # STATUS: TESTED
    # Deletes a saved captured image
    delay()
    start = 0
    pag.press("right", presses=3, interval=1)
    pag.press("enter", interval=1)
    pag.sleep(10)
    pag.press("m", interval=1, presses=2)
    data = readFromQueue(queue)
    if 'My pictures' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'renamed camera test' in data:
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press("enter", interval=1, presses=2)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'Go to folder - GuideConnect' in data:
                print('Camera 3: Can you delete a saved camera picture? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I155'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Camera 3: Failed to delete a saved camera picture")
                print('Camera 3: Can you delete a saved camera picture? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I155'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Camera 3: Failed to locate saved picture")
            print("Camera 3: failed to locate saved picture")
            print("Current String:", data)
            ws['I155'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Camera 3: Failed to locate My Pictures folder")
        print("Camera 3: failed to locate my pictures")
        print(data)
        ws['I155'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()
