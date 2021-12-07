import logging

from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

wb = load_workbook('')
ws = wb['']


def file_1(queue):  # STATUS: TESTED
    # Locate a DOCX file from Test data and opens it
    try:
        testdata()

    except OSError as test_data:
        print("Error: Files already Exist")

    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("x", interval=1)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_1: Can you navigate File explorer to find a file? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I90'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 1: Failed to find PDF file")
                    print('File_1: Can you navigate File explorer to find a file? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I90'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 1: Failed to find PDF")
                print("File_1: Failed to find PDF.pdf")
                print(data)
                ws['I90'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 1: Failed to locate test data folder")
            print("File_1: Failed to locate Test data folder")
            print(data)
            ws['I90'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 1: Failed to locate file explorer")
        print("File_1: Failed to locate File explorer")
        print(data)
        ws['I90'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_2(queue):   # STATUS: TESTED
    # Can you open an RTF
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'RTF' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_2: Can you open an RTF with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I91'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 2: Failed to open RTF with file explorer")
                    print('File_2: Can you open an RTF with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I91'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 2: Failed to find RTF")
                print("File_2: Failed to find RTF.rtf")
                print(data)
                ws['I91'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 2: Failed to locate test data folder")
            print("File_2: Failed to locate Test data folder")
            print(data)
            ws['I91'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 2: Failed to locate file explorer")
        print("File_2: Failed to locate File explorer")
        print(data)
        ws['I91'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_3(queue):   # STATUS: TESTED
    # Opens TXT file from file explorer
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)
        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("t", interval=1)
            data = readFromQueue(queue)
            if 'TXT' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_3: Can you open an TXT with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I92'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 3: Failed to open a TXT with file explorer")
                    print('File_3: Can you open an TXT with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I92'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 3: Failed to find TXT")
                print("File_3: Failed to find TXT.txt")
                print(data)
                ws['I92'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 3: Failed to locate test data folder")
            print("File_3: Failed to locate Test data folder")
            print(data)
            ws['I92'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 3: Failed to locate file explorer")
        print("File_3: Failed to locate File explorer")
        print(data)
        ws['I92'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_4(queue):    # STATUS: TESTED
    # Opens JPG
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("j", interval=1)
            data = readFromQueue(queue)
            if 'JPEG' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                    print('File_4: Can you open an JPG with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I93'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 4: Failed to open a JPG with file explorer")
                    print('File_4: Can you open an JPG with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I93'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 4: Failed to find JPG")
                print("File_4: Failed to find JPG.jpg")
                print(data)
                ws['I93'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 4: Failed to locate test data folder")
            print("File_4: Failed to locate Test data folder")
            print(data)
            ws['I93'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 4: Failed to locate file explorer")
        print("File_4: Failed to locate File explorer")
        print(data)
        ws['I93'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_5(queue):   # STATUS: TESTED
    # Opens PNG
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("p", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'PNG' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_IMAGEVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_IMAGEVIEWER' in data:
                    print('File_5: Can you open an PNG with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I94'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 5: Failed to open PNG with file explorer")
                    print('File_5: Can you open an PNG with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I94'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 5: Failed to find PNG")
                print("File_5: Failed to find PNG.png")
                print(data)
                ws['I94'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 5: Failed to locate test data folder")
            print("File_5: Failed to locate Test data folder")
            print(data)
            ws['I94'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 5: Failed to locate file explorer")
        print("File_5: Failed to locate File explorer")
        print(data)
        ws['I94'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_6(queue):  # STATUS: TESTED
    # Opens DOC file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("d", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'DOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_6: Can you open an DOC with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I95'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 6: Failed to open DOC file with file explorer")
                    print('File_6: Can you open an DOC with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I95'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 6: Failed to find DOC ")
                print("File_6: Failed to find DOC.doc")
                print(data)
                ws['I95'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 6: Failed to locate test data folder")
            print("File_6: Failed to locate Test data folder")
            print(data)
            ws['I95'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 6: Failed to locate file explorer")
        print("File_6: Failed to locate File explorer")
        print(data)
        ws['I95'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_7(queue):  # STATUS: TESTED
    # Opens XDOC
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1, presses=2)
        pag.press("t", interval=1)
        data = readFromQueue(queue)

        while 'testdata' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("x", interval=1, presses=2)
            data = readFromQueue(queue)
            if 'XDOC' in data:
                pag.press("enter", interval=1)

                while 'GMENU_ENTERTAINMENT_DOCVIEWER' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'GMENU_ENTERTAINMENT_DOCVIEWER' in data:
                    print('File_7: Can you open an DOCX with File explorer? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I96'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 7: Failed to open DOCX with file explorer")
                    print('File_7: Can you open an DOCX with File explorer? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I96'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 7: Failed to find DOCX")
                print("File_7: Failed to find DOCX.docx")
                print(data)
                ws['I96'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 7: Failed to locate test data folder")
            print("File_7: Failed to locate Test data folder")
            print(data)
            ws['I96'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 7: Failed to locate file explorer")
        print("File_7: Failed to locate File explorer")
        print(data)
        ws['I96'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_8(queue):  # STATUS: TESTED
    # Move a selected file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=4, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("t", interval=1)
    pag.press("enter", interval=1)
    pag.press("d", interval=1)
    pag.press("f2", interval=1)
    pag.press("m", interval=1)
    data = readFromQueue(queue)

    while 'Move Selected File' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Move Selected File' in data:
        pag.press("enter", interval=1)
        pag.press("g", interval=1)
        pag.press("enter", interval=1)
        pag.press("g", interval=1)
        pag.press("enter", interval=1)
        pag.press("m", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("esc", presses=2, interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)
        if 'DOC' in data:
            pag.sleep(2)

            print("File_8: Can you move a file with file explorer \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I97'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("File 8: Failed to move DOC file with file explorer")
            print("File_8: Can you move a file with file explorer \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I97'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 8: Failed to navigate to Move file")
        print("File_8: Navigation was not present on Move File \n"
              "Current String:", data)
        ws['I97'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_9(queue):  # STATUS: TESTED
    # can you rename a file
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("j", interval=1)
        data = readFromQueue(queue)

        if 'JPEG' in data:
            pag.press("f2", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)

            while 'Rename Selected File' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Rename Selected File' in data:
                pag.press("enter", interval=1)
                pag.write("rename", interval=0.5)
                pag.press("enter", interval=1, presses=2)
                pag.press("r", interval=1)
                data = readFromQueue(queue)
                if 'renameJPEG' in data:
                    print("File_9: Can you rename a file from file explorer? \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I98'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 9: Failed to rename a file ")
                    print("File_9: Can you rename a file from file explorer? \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I98'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 9: Failed to locate rename file")
                print("File_9: Failed to locate rename file \n"
                      "Current String:", data)
                ws['I98'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 9: Failed to find JPEG to rename")
            print("File_9: Failed to find JPEG to rename \n"
                  "Current String:", data)
            ws['I98'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 9: Failed to enter file explorer")
        print("File_9: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I98'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_10(queue):  # STATUS: TESTED
    # Copy a file with file explorer
    delay()
    start = 0
    x = os.environ['USERPROFILE'] + "\\Documents\\GuideConnect\\Music\\DOC.doc"
    y = os.environ['USERPROFILE'] + "\\Documents\\GuideConnect\\Documents\\testdata\\DOC.doc"
    try:
        shutil.copy(x, y)
    except:
        pass

    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        if 'DOC' in data:
            pag.press("f2", interval=1)
            pag.press("c", interval=1, presses=2)
            data = readFromQueue(queue)

            while 'Copy Selected File' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Copy Selected File' in data:
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                pag.press("enter", interval=1)
                pag.press("g", interval=1)
                pag.press("enter", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)

                while 'Music' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'Music' in data:
                    pag.press("enter", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.press("esc", presses=2, interval=1)
                    pag.press("m", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("d", interval=1)
                    data = readFromQueue(queue)
                    if 'DOC' in data:
                        print("File_10: Can you copy a file from file explorer? \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I99'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 10: Failed to copy a file from file explorer")
                        print("File_10: Can you copy a file from file explorer? \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I99'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 10: Failed to find music folder")
                    print("File_10: Failed to find music folder \n"
                          "Current String:", data)
                    ws['I99'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
        else:
            logging.critical("File 10: Failed to find DOC to copy")
            print("File_10: Failed to find DOC to copy \n"
                  "Current String:", data)
            ws['I99'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("File 10: Failed to enter file explorer")
        print("File_10: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I99'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_11(queue):     # STATUS: TESTED
    # delete multiple files and locate them in deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("t", interval=1)
        data = readFromQueue(queue)
        if 'testdata' in data:
            pag.press("enter", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            pag.press(" ", interval=1)
            pag.press("f2", interval=1)
            pag.press("d", interval=1)
            data = readFromQueue(queue)
            if 'Delete Selected Items' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                pag.press("esc", presses=2, interval=1)
                pag.press("d", presses=2, interval=1)
                data = readFromQueue(queue)

                if 'Deleted items' in data:
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'DOC.doc' in data:
                        pag.press("right", interval=1)
                        data = readFromQueue(queue)
                        if 'PDF.pdf' in data:
                            pag.press("right", interval=1)
                            data = readFromQueue(queue)
                            if 'PNG.png' in data:
                                pag.press("right", interval=1)
                                data = readFromQueue(queue)
                                if 'renameJPEG.jpg' in data:
                                    pag.press("right", interval=1)
                                    data = readFromQueue(queue)
                                    if 'RTF.rtf' in data:
                                        pag.press("right", interval=1)
                                        data = readFromQueue(queue)
                                        if 'TXT.txt' in data:
                                            pag.press("right", interval=1)
                                            data = readFromQueue(queue)
                                            if 'XDOC.docx' in data:
                                                print("File_11: Can you view multiple deleted items in deleted items \n"
                                                      "Result: PASS \n"
                                                      "Current String:", data)
                                                ws['I100'].value = 'PASS'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()

                                            else:
                                                logging.critical("File 11: Failed to locate all files in deleted items")
                                                print("File_11: Can you view multiple deleted items in deleted items \n"
                                                      "Result: FAIL \n"
                                                      "Current String:", data)
                                                ws['I100'].value = 'FAIL'
                                                wb.save('Test Reports/Automated Test cases.xlsx')
                                                returnhome()

                                        else:
                                            logging.critical("File 11: Failed to find TXT in deleted items")
                                            print("File_11: failed to find TXT in deleted items \n"
                                                  "Current String:", data)
                                            ws['I100'].value = 'FAIL'
                                            wb.save('Test Reports/Automated Test cases.xlsx')
                                            returnhome()
                                    else:
                                        logging.critical("File 11: Failed to find RTF in deleted items")
                                        print("File_11: failed to find RTF in deleted items \n"
                                              "Current String:", data)
                                        ws['I100'].value = 'FAIL'
                                        wb.save('Test Reports/Automated Test cases.xlsx')
                                        returnhome()
                                else:
                                    logging.critical("File 11: Failed to find JPG in deleted items")
                                    print("File_11: failed to find JPG in deleted items \n"
                                          "Current String:", data)
                                    ws['I100'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()
                            else:
                                logging.critical("File 11: Failed to find PNG in deleted items")
                                print("File_11: failed to find PNG in deleted items \n"
                                      "Current String:", data)
                                ws['I100'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()
                        else:
                            logging.critical("File 11: Failed to find PDF in deleted items")
                            print("File_11: failed to find PDF in deleted items \n"
                                  "Current String:", data)
                            ws['I100'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()
                    else:
                        logging.critical("File 11: Failed to find DOC in deleted items")
                        print("File_11: failed to find DOC in deleted items \n"
                              "Current String:", data)
                        ws['I100'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()
                else:
                    logging.critical("File 11: Failed to find deleted items")
                    print("File_11: failed to find deleted items \n"
                          "Current String:", data)
                    ws['I100'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("File 11: Failed to locate deleted selected items")
                print("File_11: failed to locate deleted selected items \n"
                      "Current String:", data)
                ws['I100'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()
        else:
            logging.critical("File 11: Failed to find test data folder")
            print("File_11: failed to find test data folder \n"
                  "Current String:", data)
            ws['I100'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()
    else:
        logging.critical("File 11: Failed to enter file explorer")
        print("File_11: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I100'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_12(queue):     # STATUS: TESTED
    # restore multiple files from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'DOC.doc' in data:
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            if 'PDF.pdf' in data:
                pag.press(" ", interval=1)
                pag.press("f2", interval=1)
                pag.press("m", interval=1)
                data = readFromQueue(queue)
                if 'Move selected items' in data:
                    pag.press("enter", interval=1)
                    pag.press("a", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)
                    if 'PNG.png' in data:
                        print("File_12: Restore multiple files from deleted items to attachments \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I101'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 12: Failed to restore DOC and PDF files from deleted items")
                        print("File_12: Restore multiple files from deleted items to attachments \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I101'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 12: Failed to locate move selected items")
                    print("File_12: Failed to locate move selected items \n"
                          "Current String:", data)
                    ws['I101'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 12: Failed to find PDF in deleted items")
                print("File_12: Failed to find PDF in deleted items \n"
                      "Current String:", data)
                ws['I101'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 12: Failed to find DOC in deleted items")
            print("File_12: Failed to find DOC in deleted items \n"
                  "Current String:", data)
            ws['I101'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 12: Failed to enter file explorer")
        print("File_12: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I101'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_13(queue):     # STATUS: TESTED
    # Delete multiple files from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'PNG.png' in data:
            pag.press(" ", interval=1)
            pag.press("right", interval=1)
            data = readFromQueue(queue)
            if 'renameJPEG.jpg' in data:
                pag.press(" ", interval=1)
                pag.press("f2", interval=1)
                pag.press("r", interval=1)
                data = readFromQueue(queue)

                if 'Remove selected items' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)
                    data = readFromQueue(queue)
                    if 'RTF.rtf' in data:
                        print("File_13: Delete multipe files from deleted items \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I102'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("File 13: Failed to delete PNG and JPG from deleted items")
                        print("File_13: Delete multipe files from deleted items \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I102'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("File 13: Failed to locate remove selected items")
                    print("File_13: Failed to locate remove selected items \n"
                          "Current String:", data)
                    ws['I102'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 13: Failed to find JPG in deleted items")
                print("File_13: Failed to find JPG in deleted items \n"
                      "Current String:", data)
                ws['I102'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 13: Failed to find PNG in deleted items")
            print("File_13: Failed to find PNG in deleted items \n"
                  "Current String:", data)
            ws['I102'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 13: Failed to enter file explorer")
        print("File_13: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I102'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def file_14(queue):     # STATUS: TESTED
    # Remove all items from deleted items
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    while 'File explorer' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'File explorer' in data:
        pag.press("enter", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'RTF.rtf' in data:
            pag.press("f2", interval=1)
            pag.press("r", presses=2, interval=1)
            data = readFromQueue(queue)
            if 'Remove all' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'No items found' in data:
                    print("File_14: Remove all items from deleted items \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I103'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("File 14: Failed to remove all from deleted items")
                    print("File_14: Remove all items from deleted items \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I103'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("File 14: Failed to locate remove all")
                print("File_14: Failed to locate remove all \n"
                      "Current String:", data)
                ws['I103'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("File 14: Failed to find RTF in deleted items")
            print("File_14: Failed to find RTF in deleted items \n"
                  "Current String:", data)
            ws['I103'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("File 14: Failed to enter file explorer")
        print("File_14: Failed to enter file explorer \n"
              "Current String:", data)
        ws['I103'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


# TODO: USB TESTS NEEDED