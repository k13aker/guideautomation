import logging

from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

wb = load_workbook('')
ws = wb['']


def music_1(queue):  # STATUS: TESTED
    start = 0
    delay()
    # Can you import music?
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=4, interval=1)
    data = readFromQueue(queue)
    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
        data = readFromQueue(queue)

    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if '9TAILS & guccihighwaters - NOVEMBER' in data:
        print('Music 1: Can you import Music? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I27'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 1: Failed to import music")
        print('Music 1: Can you import Music? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I27'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_2(queue):  # STATUS: TESTED
    # ALBUM VIEW: Can you play a track from album?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1, presses=4)
    pag.sleep(10)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 2: ALBUM VIEW: Can you play a track? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I28'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 2: Failed to play a track")
        pag.press("enter", interval=1)
        print('Music 2: ALBUM VIEW: Can you play a track? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I28'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_3(queue):     # STATUS: TESTED
    # Can you continue listening to a song?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(10)
    data = readFromQueue(queue)
    if 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 3: ALBUM VIEW: Can you continue playing a track? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I29'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 3: Failed to continue playing a track")
        pag.press("enter", interval=1)
        print('Music 3: ALBUM VIEW: Can you continue playing a track? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I29'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_4(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you view and change album details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while '9TAILS & guccihighwaters - NOVEMBER' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '9TAILS & guccihighwaters - NOVEMBER' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        data = readFromQueue(queue)

        while 'Details' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Details' in data:
            pag.press("enter", presses=2, interval=1)
            pag.press("u", interval=1)
            data = readFromQueue(queue)

            while 'Unknown Album [55CC5344]' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'Unknown Album [55CC5344]' in data:
                pag.press("enter", presses=2, interval=1)
                data = readFromQueue(queue)
                if 'Unknown Album [55CC5344]' in data:
                    print('Music 4: ALBUM VIEW: Can you view and change the album details? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I30'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Music 4: Failed to change album details")
                    print('Music 4: ALBUM VIEW: Can you view and change the album details? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I30'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 4: Failed to switch albums")
                print('Music 4: ALBUM VIEW: Failed to switch albums \n'
                      '>>> Current String:', data)
                ws['I30'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 4: Failed to open details")
            print('Music 4: ALBUM VIEW: Failed to open details \n'
                  '>>> Current String:', data)
            ws['I30'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 4: Failed to locate 9Tails Album")
        print('Music 4: ALBUM VIEW: Failed to locate 9Tails album \n'
              '>>> Current String:', data)
        ws['I30'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_5(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you rename an album?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'Unknown Album [55CC5344]' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Unknown Album [55CC5344]' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)

        while 'Rename' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Rename' in data:
            pag.press("enter", interval=1)
            pag.write("A", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'A' in data:
                print('Music 5: ALBUM VIEW: Can you rename an existing album? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I31'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 5: Failed to rename existing album")
                print('Music 5: ALBUM VIEW: Can you rename an existing album? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I31'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 5: Failed to locate rename album")
            print('Music 5: ALBUM VIEW: Failed to locate rename album \n'
                  '>>> Current String:', data)
            ws['I31'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 5: Failed to find album")
        print('Music 5: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I31'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_6(queue):     # STATUS: TESTED
    # ALBUM VIEW: Can you add a new album in details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'A' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'A' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ALBUM_ACTIONS_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ALBUM_ACTIONS_ADD' in data:
            pag.write("Added", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Added' in data:
                print('Music 6: ALBUM VIEW: Can you add a new album via details \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I32'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 6: Failed to add new album details")
                print('Music 6: ALBUM VIEW: Can you add a new album via details \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I32'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 6: Failed to add new album")
            print('Music 6: ALBUM VIEW: Failed to add new album \n'
                  '>>> Current String:', data)
            ws['I32'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 6: Failed to find album")
        print('Music 6: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I32'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_7(queue):     # STATUS: TESTED
    # ALBUM VIEW: Perform an album search
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'A' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'A' in data:
        pag.press("f2", interval=1)
        pag.press("s", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' in data:
            pag.write("Added", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'Added' in data:
                print('Music 7: ALBUM VIEW: Can you search for a specific album? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I33'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 7: Failed to search for specific album")
                print('Music 7: ALBUM VIEW: Can you search for a specific album? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I33'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 7: Failed to open search album")
            print('Music 7: ALBUM VIEW: Failed to open search album \n'
                  '>>> Current String:', data)
            ws['I33'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 7: Failed to find album")
        print('Music 7: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I33'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_8(queue):     # STATUS: TESTED
    # ALBUM VIEW (Track list) - Can you view and change all the details?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while '9TAILS & guccihighwaters - for once (prod. notmorgn)' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if '9TAILS & guccihighwaters - for once (prod. notmorgn)' in data:
        pag.press("f2", interval=1)
        pag.press("d", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TITLE' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TITLE' in data:
            pag.write("changed", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("a", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while '9TAILS Archive' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if '9TAILS Archive' in data:
                pag.press("enter", interval=1)
                pag.press("a", presses=2, interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)

                while 'A' not in data:
                    data = readFromQueue(queue)
                    if start < timeout:
                        pag.sleep(10)
                        start += 10
                        if start == timeout:
                            print(data, ": Timed out after", timeout, "Seconds")
                            returnhome()
                            break

                if 'A' in data:
                    pag.press("a", interval=1)
                    pag.press("enter", interval=1)
                    pag.press("t", interval=1)
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)

                    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TRACK' not in data:
                        data = readFromQueue(queue)
                        if start < timeout:
                            pag.sleep(10)
                            start += 10
                            if start == timeout:
                                print(data, ": Timed out after", timeout, "Seconds")
                                returnhome()
                                break

                    if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_TRACK' in data:
                        pag.write("7", interval=0.5)
                        pag.press("enter", interval=1)
                        pag.press("g", interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)

                        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_GENRE' not in data:
                            data = readFromQueue(queue)
                            if start < timeout:
                                pag.sleep(10)
                                start += 10
                                if start == timeout:
                                    print(data, ": Timed out after", timeout, "Seconds")
                                    returnhome()
                                    break

                        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_GENRE' in data:
                            pag.write("changed", interval=0.5)
                            pag.press("enter", interval=1)
                            pag.press("y", interval=1)
                            pag.press("enter", interval=1)
                            data = readFromQueue(queue)

                            while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_YEAR' not in data:
                                data = readFromQueue(queue)
                                if start < timeout:
                                    pag.sleep(10)
                                    start += 10
                                    if start == timeout:
                                        print(data, ": Timed out after", timeout, "Seconds")
                                        returnhome()
                                        break

                            if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_YEAR' in data:
                                pag.write("2021", interval=0.5)
                                pag.press("enter", interval=1)
                                pag.press("esc", presses=2, interval=1)
                                pag.press("a", interval=1)
                                pag.press("enter", interval=1)
                                data = readFromQueue(queue)
                                if 'changed' in data:
                                    print('Music 8: ALBUM VIEW: Can you edit the track list details? \n'
                                          '>>> Result: PASS \n'
                                          '>>> Current String:', data)
                                    ws['I34'].value = 'PASS'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                                else:
                                    logging.critical("Music 8: Failed to edit track list details")
                                    print('Music 8: ALBUM VIEW: Can you edit the track list details? \n'
                                          '>>> Result: FAIL \n'
                                          '>>> Current String:', data)
                                    ws['I34'].value = 'FAIL'
                                    wb.save('Test Reports/Automated Test cases.xlsx')
                                    returnhome()

                            else:
                                logging.critical("Music 8: Failed to open edit year")
                                print('Music 8: ALBUM VIEW: Failed to open edit year \n'
                                      '>>> Current String:', data)
                                ws['I34'].value = 'FAIL'
                                wb.save('Test Reports/Automated Test cases.xlsx')
                                returnhome()

                        else:
                            logging.critical("Music 8: Failed to open change genre")
                            print('Music 8: ALBUM VIEW: Failed to open change genre \n'
                                  '>>> Current String:', data)
                            ws['I34'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Music 8: Failed to open edit track number")
                        print('Music 8: ALBUM VIEW: Failed to open edit track number \n'
                              '>>> Current String:', data)
                        ws['I34'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Music 8: Failed to open change album")
                    print('Music 8: ALBUM VIEW: Failed to open change album \n'
                          '>>> Current String:', data)
                    ws['I34'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 8: Failed to open change artist")
                print('Music 8: ALBUM VIEW: Failed to open change artist \n'
                      '>>> Current String:', data)
                ws['I34'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 8: Failed to open edit song title")
            print('Music 8: ALBUM VIEW: Failed to open edit song title \n'
                  '>>> Current String:', data)
            ws['I34'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 8: Failed to find album")
        print('Music 8: ALBUM VIEW: Failed to find album \n'
              '>>> Current String:', data)
        ws['I34'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_9(queue):     # STATUS: TESTED
    # ALBUM VIEW (Track list) - Can you delete a track?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("a", interval=1)
    data = readFromQueue(queue)

    while 'Added' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Added' in data:
        pag.press("enter", interval=1)
        pag.press("f2", interval=1)
        pag.press("d", presses=2, interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

        while 'No items found' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'No items found' in data:
            print('Music 9: ALBUM VIEW: Can you delete a track?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I35'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 9: Failed to delete a track")
            print('Music 9: ALBUM VIEW: Can you delete a track?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I35'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 9: Failed to locate album 'Added'")
        print('Music 9: ALBUM VIEW: Failed to locate Album "Added" \n'
              '>>> Current String:', data)
        ws['I35'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_10(queue):        # STATUS: TESTED
    # ALBUM VIEW: Search for a specific track in an album
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("s", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_NEXT_EDIT' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_NEXT_EDIT' in data:
        pag.write("scars", interval=0.5)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9TAILS & guccihighwaters - scars (prod. notmorgn)' in data:
            print('Music 10: ALBUM VIEW: Can you search for a specific track?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I36'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 10: Failed to search for specific track")
            print('Music 10: ALBUM VIEW: Can you search for a specific track?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I36'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 10: Failed to open search for song")
        print('Music 10: ALBUM VIEW: Failed to open search for song \n'
              '>>> Current String:', data)
        ws['I36'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_11(queue):    # STATUS: TESTED
    # ALBUM VIEW: Can you remove albums then import via the F2 feature
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    for _ in range(4):
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'Remove' in data:
            pag.press("enter", presses=3, interval=1)

        else:
            print("failed to locate remove")
            break

    data = readFromQueue(queue)
    if 'No items found' in data:
        pag.press("f2", interval=1)
        pag.press("i", interval=1)
        pag.press("enter", presses=2, interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
            data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if '9TAILS & guccihighwaters - NOVEMBER' in data:
                print('Music 11: ALBUM VIEW: Can you remove all albums and then import via the f2 menu  \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I37'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 11: Failed to remove all albums and re import via f2 menu")
                print('Music 11: ALBUM VIEW: Can you remove all albums and then import via the f2 menu  \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I37'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 11: Failed to import music")
            print('Music 11: ALBUM VIEW: Failed to import music \n'
                  '>>> Current String:', data)
            ws['I37'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 11: Failed to delete all albums")
        print('Music 11: ALBUM VIEW: Failed to delete all albums \n'
              '>>> Current String:', data)
        ws['I37'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_12(queue):    # STATUS: TESTED
    # ALBUM VIEW: Can you change your view?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("a", interval=1)
    data = readFromQueue(queue)

    while 'Artist view' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Artist view' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9TAILS Archive' in data:
            print('Music 12: ALBUM VIEW: Can you change to artist view?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I38'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 12: Failed to change to artist view")
            print('Music 12: ALBUM VIEW: Can you change to artist view?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I38'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 12: Failed to switch artist view")
        print('Music 12: ALBUM VIEW: Failed to switch to Artist View \n'
              '>>> Current String:', data)
        ws['I38'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_13(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you play a track?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(10)
    data = readFromQueue(queue)
    if '' in data:
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I39'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    elif 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I39'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 13: Failed to play a track")
        pag.press("enter", interval=1)
        print('Music 13: ARTIST VIEW: Can you play a track?  \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I39'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_14(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you add and change the artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_ADD' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_ADD' in data:
            pag.write("tester", interval=0.5)
            pag.press("enter", interval=1)
            pag.press("enter", presses=2, interval = 1)
            data = readFromQueue(queue)
            if 'tester' in data:
                print('Music 14: ARTIST VIEW: Can you add and change the artist details? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I40'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 14: Failed to change artist details")
                print('Music 14: ARTIST VIEW: Can you add and change the artist details? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I40'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 14: Failed to add a new artist")
            print('Music 14 ARTIST VIEW: Failed to add a new artist \n'
                  '>>> Current String:', data)
            ws['I40'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 14: Failed to locate view details")
        print('Music 14 ARTIST VIEW: Failed to locate view details \n'
              '>>> Current String:', data)
        ws['I40'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_15(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you rename an artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("d", interval=1)
    data = readFromQueue(queue)

    while 'Details' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Details' in data:
        pag.press("enter", presses=2, interval=1)
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_RENAME' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY_EDIT_ARTIST_ACTIONS_RENAME' in data:
            pag.write("renamed", interval=0.5)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)
            if 'renamed' in data:
                print('Music 15: ARTIST VIEW: Can you rename the artist details? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I41'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 15: Failed to rename artist details")
                print('Music 15: ARTIST VIEW: Can you rename the artist details? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I41'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 15: Failed to rename artist")
            print('Music 15 ARTIST VIEW: Failed to rename artist \n'
                  '>>> Current String:', data)
            ws['I41'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 15: Failed to locate view details")
        print('Music 15 ARTIST VIEW: Failed to locate view details \n'
              '>>> Current String:', data)
        ws['I41'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_16(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you remove an artist?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    for _ in range(2):
        pag.press("f2", interval=1)
        pag.press("r", interval=1)
        pag.press("enter", presses=3, interval=1)
        data = readFromQueue(queue)

    if 'No items found' in data:
        print('Music 16: ARTIST VIEW: Can you remove an artist? \n'
              '>>> Result: PASS \n'
              '>>> Current String:', data)
        ws['I42'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Music 16: Failed to remove an artist")
        print('Music 16: ARTIST VIEW: Can you remove an artist? \n'
              '>>> Result: FAIL \n'
              '>>> Current String:', data)
        ws['I42'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_17(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you perform a search?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_ENTERTAINMENT_MUSIC_LIBRARY' in data:
        data = readFromQueue(queue)

    while 'OK' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'OK' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        if '9TAILS Archive' in data:
            pag.press("f2", interval=1)
            pag.press("s", interval=1)
            pag.press("enter", interval=1)
            data = readFromQueue(queue)

            while 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_ENTERTAINMENT_MUSIC_SEARCH_EDIT' in data:
                pag.write("9", interval=0.5)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                pag.press("right", interval=1)
                if '9TAILS Archive' in data:
                    print('Music 17: ARTIST VIEW: Can you import music and then search for an artist? \n'
                          '>>> Result: PASS \n'
                          '>>> Current String:', data)
                    ws['I43'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Music 17: Failed to import music and search for an artist")
                    print('Music 17: ARTIST VIEW: Can you import music and then search for an artist? \n'
                          '>>> Result: FAIL \n'
                          '>>> Current String:', data)
                    ws['I43'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Music 17: Failed to search for an artist")
                print('Music 17 ARTIST VIEW: Failed to search for an artist \n'
                      '>>> Current String:', data)
                ws['I43'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 17: Failed to locate album")
            print('Music 17 ARTIST VIEW: Failed to locate album \n'
                  '>>> Current String:', data)
            ws['I43'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 17: Failed to import music")
        print('Music 17 ARTIST VIEW: Failed to import music \n'
              '>>> Current String:', data)
        ws['I43'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_18(queue):    # STATUS: TESTED
    # ARTIST VIEW: Can you change to folder view?
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("enter", interval=1)
    pag.press("f", interval=1)
    data = readFromQueue(queue)
    if 'Folder view' in data:
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'C:' in data:
            print('Music 18: ARTIST VIEW: Can you change to folder view?  \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            ws['I44'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Music 18: Failed to change to folder view")
            print('Music 18: ARTIST VIEW: Can you change to folder view?  \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            ws['I44'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 18: Failed to switch to folder view")
        print('Music 18: ALBUM VIEW: Failed to switch to folder View \n'
              '>>> Current String:', data)
        ws['I44'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def music_19(queue):    # STATUS: TESTED
    # FOLDER VIEW: Search through folders to play a track
    start = 0
    delay()
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    pag.press("m", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)
    if 'C:' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("m", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if '9tails' in data:
            pag.press("enter", presses=2, interval=1)
            pag.sleep(10)
            data = readFromQueue(queue)
            if '' in data:
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            elif 'GMENU_ENTERTAINMENT_MUSIC_LISTEN_PLAY' in data:
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: PASS \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Music 19: Failed to play a track")
                pag.press("enter", interval=1)
                print('Music 19: FOLDER VIEW: Can you play a track? \n'
                      '>>> Result: FAIL \n'
                      '>>> Current String:', data)
                ws['I45'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Music 19: Failed to locate 9Tails folder")
            print('Music 19: FOLDER VIEW: Failed to locate 9Tails folder \n'
                  '>>> Current String:', data)
            ws['I45'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Music 19: Failed to be in the folder view")
        print('Music 19: FOLDER VIEW: Failed to be in folder View \n'
              '>>> Current String:', data)
        ws['I45'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


# TODO: ADD ANY USB TESTS