import logging

from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120


def about(queue):   # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Check my Premium Plan' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Check my Premium Plan' in data:
        pag.press("enter", interval=1)
        pag.sleep(2)
        print("About 1: Check Premium plan status \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['I104'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Tools 1: Failed to check premium plan status")
        print("About 1: Check Premium plan status \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I104'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def calc(queue):    # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_CALCULATOR_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_CALCULATOR_MAIN' in data:
        # TEST CASE 1: 50,000 + 90,000 = 140,000
        pag.press("5", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("+")
        pag.press("9", interval=0.5)
        pag.press("0", presses=4, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 2: 666 / 22 = 30.272
        pag.press("6", presses=3, interval=0.5)
        pag.press("/")
        pag.press("2", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 3: 200 - 11 = 189
        pag.press("2", interval=0.5)
        pag.press("0", presses=2, interval=0.5)
        pag.press("-")
        pag.press("1", presses=2, interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)
        pag.press("backspace", interval=1)

        # TEST CASE 4: 497 x 15 = 7,455
        pag.press("4", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("*")
        pag.press("1", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("enter", interval=1)
        pag.sleep(2)

        print("Calc 1: Use the calculator to add, multiply, divide and minus \n"
              "Result: PASS \n"
              "Current String:", data)
        ws['I105'].value = 'PASS'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()

    else:
        logging.critical("Tools 2: Failed to use the calculator")
        print("Calc 1: Use the calculator to add, multiply, divide and minus \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I105'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def dict_valid(queue):  # STATUS: TESTED
    delay()
    start = 0
    # TEST CASE 1: Search a valid valid dictionary entry = test
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("test", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_DICTIONARYMODULE_RESULT_DEFINITIONS' in data:
            print("Dictionary Valid 1: Can you search for a valid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I106'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 3: Failed to search a valid dictionary term")
            print("Dictionary Valid 1: Can you search for a valid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I106'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 3: Failed to open dictionary")
        print("Dictionary Valid 1: Failed to open dictionary  \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I106'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def dict_invalid(queue):    # STATUS: TESTED
    # TEST CASE 2: Can you search an invalid search term in dictionary
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", interval=1)
    pag.press("enter", interval=1)
    data = readFromQueue(queue)

    while 'GMENU_DICTIONARYMODULE_MAIN' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_DICTIONARYMODULE_MAIN' in data:
        pag.write("twest", interval=1)
        pag.press("enter", interval=1)
        pag.sleep(10)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            print("Dictionary Invalid 1: Can you search for an invalid search term in dictionary? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I107'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 4: Failed to search an invalid dictionary term")
            print("Dictionary Invalid 1: Can you search for an invalid search term in dictionary? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I107'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 4: Failed to open dictionary")
        print("Dictionary Invalid 1: Failed to open dictionary \n"
              "Result: FAIL \n"
              "Current String:", data)
        ws['I107'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def gstart_vids(queue):  # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=3, interval=1)
    pag.sleep(2)
    data = readFromQueue(queue)

    while 'Keyboard and mouse video' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Keyboard and mouse video' in data:
        pag.press("enter", interval=1)
        pag.sleep(5)
        pag.press("esc", interval=1)
        pag.press("R", interval=1)
        data = readFromQueue(queue)
        if 'Remote control video' in data:
            pag.press("enter", interval=1)
            pag.sleep(5)
            pag.press("esc", interval=1)
            pag.press("T", interval=1)
            data = readFromQueue(queue)
            if 'Touch screen video' in data:
                pag.press("enter", interval=1)
                pag.sleep(5)
                pag.press("esc", interval=1)
                pag.press("V", interval=1)
                data = readFromQueue(queue)
                if 'Voice input video' in data:
                    pag.press("enter", interval=1)
                    pag.sleep(5)

                    print("Getting Started Videos: Can you play the Getting Started Videos \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I108'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 5: Failed to find Voice Video")
                    print("Getting Started Videos: Voice video missing")
                    print("Current String:", data)
                    ws['I108'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 5: Failed to find Touch Video")
                print("Getting Started Videos: Touch video missing")
                print("Current String:", data)
                ws['I108'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 5: Failed to find Remote video")
            print("Getting Started Videos: Remote video missing")
            print("Current String:", data)
            ws['I108'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 5: Failed to find M&K Video")
        print("Getting Started Videos: Keyboard and mouse video missing \n"
              "Result: FAIL -  Keyboard video not present \n"
              "Current String:", data)
        ws['I108'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def learn_keyboard(queue):  # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)

    while 'GMENU_TYPING_TUTOR_DIALOGUE' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'GMENU_TYPING_TUTOR_DIALOGUE' in data:
        pag.press("1", interval=0.5)
        pag.press("2", interval=0.5)
        pag.press("3", interval=0.5)
        pag.press("4", interval=0.5)
        pag.press("5", interval=0.5)
        pag.press("6", interval=0.5)
        pag.press("7", interval=0.5)
        pag.press("8", interval=0.5)
        pag.press("9", interval=0.5)
        pag.press("q", interval=0.5)
        pag.press("w", interval=0.5)
        pag.press("e", interval=0.5)
        pag.press("r", interval=0.5)
        pag.press("t", interval=0.5)
        pag.press("y", interval=0.5)
        pag.press("u", interval=0.5)
        pag.press("i", interval=0.5)
        pag.press("o", interval=0.5)
        pag.press("p", interval=0.5)
        pag.press("a", interval=0.5)
        pag.press("s", interval=0.5)
        pag.press("d", interval=0.5)
        pag.press("f", interval=0.5)
        pag.press("g", interval=0.5)
        pag.press("h", interval=0.5)
        pag.press("j", interval=0.5)
        pag.press("k", interval=0.5)
        pag.press("l", interval=0.5)
        pag.press("z", interval=0.5)
        pag.press("x", interval=0.5)
        pag.press("c", interval=0.5)
        pag.press("v", interval=0.5)
        pag.press("b", interval=0.5)
        pag.press("n", interval=0.5)
        pag.press("m", interval=0.5)
        pag.press("enter", interval=0.5)
        data = readFromQueue(queue)
        if 'Learn the Keyboard' in data:
            print("LTK 1: Can you use the typing tutor? \n"
                  "Result: PASS \n"
                  "Current String:", data)
            ws['I109'].value = 'PASS'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

        else:
            logging.critical("Tools 6: Failed to use typing tutor")
            print("LTK 1: Can you use the typing tutor? \n"
                  "Result: FAIL \n"
                  "Current String:", data)
            ws['I109'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 6: Failed to enter Learn the keyboard")
        print("LTK 1: Could not enter learn the keyboard \n"
              "Current String:", data)
        ws['I109'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def shortcut(queue):    # STATUS: TESTED
    delay()
    start = 0
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=2, interval=1)
    data = readFromQueue(queue)
    while 'Training' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Training' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        data = readFromQueue(queue)
        while 'Shortcut keys' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'Shortcut keys' in data:
            pag.press("enter", interval=1)
            while 'GMENU_COREMODULE_SHORTCUTKEYS' not in data:
                data = readFromQueue(queue)
                if start < timeout:
                    pag.sleep(10)
                    start += 10
                    if start == timeout:
                        print(data, ": Timed out after", timeout, "Seconds")
                        returnhome()
                        break

            if 'GMENU_COREMODULE_SHORTCUTKEYS' in data:
                print("Shortcuts 1: Can you view the shortcuts documentation? \n"
                      "Result: PASS \n"
                      "Current String:", data)
                ws['I110'].value = 'PASS'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

            else:
                logging.critical("Tools 7: Failed to view shortcut keys")
                print("Shortcuts 1: Can you view the shortcuts documentation? \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I110'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 7: Failed to enter shortcut keys")
            print("Shortcuts 1: failed to enter shortcut keys")
            print(data)
            ws['I110'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 7: Failed to navigate to Training function")
        print("Shortcuts 1: failed to navigate to training")
        print("data")
        ws['I110'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def backup(queue):  # STATUS: TESTED
    # Backs up Guide Connect
    start = 0
    delay()

    # ADD A TEST NOTE
    pag.press("right", presses=7, interval=1)
    pag.press("enter", interval=1)
    pag.press("f2", interval=1)
    pag.press("a", interval=1)
    pag.press("enter", interval=1)
    pag.write("restore", interval=0.5)
    pag.press("enter", interval=1)
    pag.press("esc", interval=1)
    pag.press("enter", presses=2, interval=1)
    returnhome()

    # BEGIN BACK UP TEST
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)
    if 'Backup and restore' in data:
        pag.press("enter", presses=3, interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)

        while 'OK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'OK' in data:
            returnhome()
            delay()
            pag.press("n", interval=1)
            pag.press("enter", interval=1)
            pag.press("r", interval=1)
            data = readFromQueue(queue)
            if 'restore' in data:
                pag.press("f2", interval=1)
                pag.press("d", presses=2, interval=1)
                pag.press("enter", presses=3, interval=1)
                data = readFromQueue(queue)
                if 'No notes available' in data:
                    print("Back up 1: Back up GC \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I111'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 8: Failed to back up GC")
                    print("Back up 1: Back up GC")
                    print(data)
                    ws['I111'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()
            else:
                logging.critical("Tools 8: Failed to locate Restore Note")
                print("Back up 1: Failed to locate note 'Restore'")
                print(data)
                ws['I111'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 8: Failed to Back up GC")
            print("Back up 1: Failed to back up GC")
            print(data)
            ws['I111'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 8: Failed to locate back up and restore")
        print("failed to locate back and restore")
        print(data)
        ws['I111'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def restore(queue): # STATUS: TESTED
    # Backs up Guide Connect
    start = 0
    launch = 300
    delay()
    pag.press("right", presses=9, interval=1)
    pag.press("enter", interval=1)
    pag.press("b", interval=1)
    data = readFromQueue(queue)
    if 'Backup and restore' in data:
        pag.press("enter", interval=1)
        pag.press("r", interval=1)
        data = readFromQueue(queue)
        if 'Restore GuideConnect' in data:
            pag.press("enter", presses=2, interval=1)
            pag.press("right", interval=1)
            pag.press("enter", presses=2, interval=1)

            while 'Emails' not in data:
                data = readFromQueue(queue)
                if start < launch:
                    pag.sleep(10)
                    start += 10
                    if start == launch:
                        print(data, ": Timed out after", launch, "Seconds")
                        returnhome()
                        break

            if 'Emails' in data:
                delay()
                pag.press("n", interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'restore' in data:
                    print("Restore 1: restore GC \n"
                          "Result: PASS \n"
                          "Current String:", data)
                    ws['I112'].value = 'PASS'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

                else:
                    logging.critical("Tools 9: Failed to restore GC")
                    print("Restore 1: restore GC \n"
                          "Result: FAIL \n"
                          "Current String:", data)
                    ws['I112'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 9: Failed to find Emails in data on relaunch")
                print("Restore 1: restore GC \n"
                      "Result: FAIL \n"
                      "Current String:", data)
                ws['I112'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 9: Failed to locate restore")
            print("Restore 1: failed to locate restore")
            print(data)
            ws['I112'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 9: Failed to locate back up and restore")
        print("failed to locate back and restore")
        print(data)
        ws['I112'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def reset_settings(queue):  # STATUS: TESTED
    # Resets settings only
    start = 0
    launch = 300
    delay()
    # CHANGE TILE VIEW TO OFF
    pag.press("s", presses=2, interval=1)
    pag.press("enter", presses=2, interval=1)
    pag.press("t", presses=2, interval=1)
    data = readFromQueue(queue)
    if 'Tile view mode' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", interval=1)
        pag.press("t", presses=2, interval=1)
        pag.press("enter", interval=1)
        data = readFromQueue(queue)
        if 'Off' in data:
            returnhome()
            pag.press("t", interval=1)
            pag.press("enter", interval=1)
            pag.press("b", interval=1)
            data = readFromQueue(queue)
            if 'Backup and restore' in data:
                pag.press("enter", interval=1)
                pag.press("r", presses=2, interval=1)
                pag.press("enter", interval=1)
                data = readFromQueue(queue)
                if 'Reset settings only' in data:
                    pag.press("enter", interval=1)
                    pag.press("right", interval=1)
                    pag.press("enter", presses=2, interval=1)

                    while 'Emails' not in data:
                        data = readFromQueue(queue)
                        if start < launch:
                            pag.sleep(10)
                            start += 10
                            if start == launch:
                                print(data, ": Timed out after", launch, "Seconds")
                                returnhome()
                                break

                    if 'Emails' in data:
                        pag.press("s", presses=2, interval=1)
                        pag.press("enter", presses=2, interval=1)
                        pag.press("t", presses=2, interval=1)
                        pag.press("enter", interval=1)
                        data = readFromQueue(queue)
                        if 'On' in data:
                            print("Reset 1: Reset Settings Only \n"
                                  "Result: PASS \n"
                                  "Current String:", data)
                            ws['I113'].value = 'PASS'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                        else:
                            logging.critical("Tools 10: Failed to reset GC Settings only")
                            print("Reset 1: Reset Settings Only \n"
                                  "Result: FAIL \n"
                                  "Current String:", data)
                            ws['I113'].value = 'FAIL'
                            wb.save('Test Reports/Automated Test cases.xlsx')
                            returnhome()

                    else:
                        logging.critical("Tools 10: Failed to relaunch GC after reset settings")
                        print("Failed to relaunch GC after Re setting settings ")
                        print(data)
                        ws['I113'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                else:
                    logging.critical("Tools 10: Failed to locate reset settings only")
                    print("Failed to locate reset settings only")
                    print(data)
                    ws['I113'].value = 'FAIL'
                    wb.save('Test Reports/Automated Test cases.xlsx')
                    returnhome()

            else:
                logging.critical("Tools 10: Failed to locate back up and restore")
                print("Failed to locate back up and restore")
                print(data)
                ws['I113'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 10: Failed to turn tile view off")
            print("Failed to turn tile view off")
            print(data)
            ws['I113'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 10: Failed to navigate to tile view mode in settings")
        print("Failed to navigate to tile view mode setting")
        print(data)
        ws['I113'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()


def factory(queue): # STATUS: TESTED
    # Factory resets all settings
    start = 0
    launch = 300
    init()
    pag.press("enter", presses=2, interval=1)
    pag.write("guideautomation1@outlook.com", interval=0.5)
    pag.press("enter", interval=1)
    pag.write("test", interval=0.5)
    pag.press("enter", presses=3, interval=1)
    pag.write("guideconnect1", interval=0.5)
    pag.press("enter", presses=2, interval=1)
    data = readFromQueue(queue)
    while 'OK' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'OK' in data:
        returnhome()
        delay()
        pag.press("t", interval=1)
        pag.press("enter", interval=1)
        pag.press("b", interval=1)
        data = readFromQueue(queue)
        if 'Backup and restore' in data:
            pag.press("enter", interval=1)
            pag.press("r", presses=2, interval=1)
            pag.press("enter", interval=1)
            pag.press("f", interval=1)
            data = readFromQueue(queue)
            if 'Full factory reset' in data:
                pag.press("enter", interval=1)
                pag.press("right", interval=1)
                pag.press("enter", presses=2, interval=1)

                while 'Emails' not in data:
                    data = readFromQueue(queue)
                    if start < launch:
                        pag.sleep(10)
                        start += 10
                        if start == launch:
                            print(data, ": Timed out after", launch, "Seconds")
                            returnhome()
                            break

                if 'Emails' in data:
                    delay()
                    pag.press("enter", interval=1)
                    data = readFromQueue(queue)
                    if 'Yes' in data:
                        print("Factory 1: Factory Reset Settings \n"
                              "Result: PASS \n"
                              "Current String:", data)
                        ws['I114'].value = 'PASS'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

                    else:
                        logging.critical("Tools 11: Failed to Factory Reset settings")
                        print("Factory 1: Factory Reset Settings \n"
                              "Result: FAIL \n"
                              "Current String:", data)
                        ws['I114'].value = 'FAIL'
                        wb.save('Test Reports/Automated Test cases.xlsx')
                        returnhome()

            else:
                logging.critical("Tools 11: Failed to locate factory reset")
                print("Factory 1: Failed to locate factory reset")
                print(data)
                ws['I114'].value = 'FAIL'
                wb.save('Test Reports/Automated Test cases.xlsx')
                returnhome()

        else:
            logging.critical("Tools 11: Failed to locate back up and restore")
            print("Factory 1: Failed to locate back up and restore")
            print(data)
            ws['I114'].value = 'FAIL'
            wb.save('Test Reports/Automated Test cases.xlsx')
            returnhome()

    else:
        logging.critical("Tools 11: Failed to log into email account")
        print("Factory 1: Failed to log into email account")
        print(data)
        ws['I114'].value = 'FAIL'
        wb.save('Test Reports/Automated Test cases.xlsx')
        returnhome()