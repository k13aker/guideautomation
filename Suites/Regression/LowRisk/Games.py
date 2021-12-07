import logging
from Misc.MyFunctions import *
from openpyxl import load_workbook
timeout = 120

# wb = load_workbook('')
# ws = wb['']


def Hangman(queue):
    # Launches Hangman
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", presses=4, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_HANGMAN' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_HANGMAN' in data:
            pag.sleep(5)
            print('Games 1 - Hangman: Can you launch hangman? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical("Games 1 - Failed to launch Hangman")
            print('Games 1 - Hangman: Can you launch hangman? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 1 - Failed to open games")
        print("Games 1: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def Sudoku(queue):
    # TEST CASE 2: LAUNCH SUDOKU
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", interval=1)
        pag.press("enter", presses=3, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)

        while 'GMENU_GAMES_SUDOKU' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_SUDOKU' in data:
            pag.sleep(5)
            print('Games 2 - Sudoku: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical("Games 2 - Failed to launch Sudoku")
            print('Games 2 - Sudoku: Can you launch Sudoku? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 2 - Failed to open games")
        print("Games 2 - Sudoku: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()


def Blackjack(queue):
    # TEST CASE 3: LAUNCH BLACKJACK
    delay()
    start = 0
    pag.press("right", presses=6, interval=1)  # navigating to entertainment module
    pag.press("enter", interval=1)
    pag.press("right", presses=3, interval=1)
    data = readFromQueue(queue)

    while 'Games' not in data:
        data = readFromQueue(queue)
        if start < timeout:
            pag.sleep(10)
            start += 10
            if start == timeout:
                print(data, ": Timed out after", timeout, "Seconds")
                returnhome()
                break

    if 'Games' in data:
        pag.press("enter", interval=1)
        pag.press("right", presses=2, interval=1)
        pag.press("enter", presses=2, interval=1)
        pag.sleep(1)
        data = readFromQueue(queue)
        pag.sleep(5)

        while 'GMENU_GAMES_BLACKJACK' not in data:
            data = readFromQueue(queue)
            if start < timeout:
                pag.sleep(10)
                start += 10
                if start == timeout:
                    print(data, ": Timed out after", timeout, "Seconds")
                    returnhome()
                    break

        if 'GMENU_GAMES_BLACKJACK' in data:
            print('Games 3 - Blackjack: Can you launch Sudoku? \n'
                  '>>> Result: PASS \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'PASS'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

        else:
            logging.critical('Games 3 - Failed to launch Blackjack')
            print('Games 3 - Blackjack: Can you launch Blackjack? \n'
                  '>>> Result: FAIL \n'
                  '>>> Current String:', data)
            # ws['D39'].value = 'FAIL'
            # wb.save('Test Reports/Excel automation.xlsx')
            returnhome()

    else:
        logging.critical("Games 3 - Blackjack: Failed to open games")
        print("Games 3 - Blackjack: Failed to open games")
        print(data)
        # ws['D39'].value = 'FAIL'
        # wb.save('Test Reports/Excel automation.xlsx')
        returnhome()